# core/pipeline.py
from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Sequence
from collections import Counter, defaultdict

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border
from openpyxl.worksheet.views import Selection
from pyxlsb import open_workbook as open_xlsb_workbook

from core.utils import normalize_text, text_contains, text_eq


@dataclass
class PipelineResult:
    ok: bool
    outputs: List[Path]
    logs: List[str]

    transfer_in_done: int = 0
    transfer_in_hold: int = 0
    transfer_out_done: int = 0
    transfer_out_hold: int = 0
    transfer_out_auto_skip: int = 0


@dataclass
class ScanResult:
    # ê¸°ë³¸ ìƒíƒœ
    ok: bool = False
    logs: List[str] = field(default_factory=list)

    # í•™êµ/ì—°ë„ ì •ë³´
    school_name: str = ""
    year_str: str = ""
    year_int: int = 0

    # ê²½ë¡œë“¤
    project_root: Path = Path(".")
    input_dir: Path = Path(".")
    output_dir: Path = Path(".")
    template_register: Optional[Path] = None
    template_notice: Optional[Path] = None
    db_path: Optional[Path] = None

    # ì¸í’‹ íŒŒì¼
    freshmen_file: Optional[Path] = None
    teacher_file: Optional[Path] = None
    transfer_file: Optional[Path] = None
    withdraw_file: Optional[Path] = None

    # í•™ìƒëª…ë¶€ ê´€ë ¨
    need_roster: bool = False              # ì „ì…/ì „ì¶œ ì¤‘ í•˜ë‚˜ë¼ë„ ìˆìœ¼ë©´ True
    roster_path: Optional[Path] = None
    roster_year: Optional[int] = None
    roster_info: Optional[Dict[str, Any]] = None
    roster_basis_date: Optional[date] = None  # í•™ìƒëª…ë¶€ ê¸°ì¤€ì¼(íŒŒì¼ ìˆ˜ì •ì¼ or ì‚¬ìš©ìê°€ ìˆ˜ì •í•œ ê°’)

    # UI í”Œë˜ê·¸
    needs_open_date: bool = False          # ì „ì¶œ ìˆìœ¼ë©´ True â†’ ê°œí•™ì¼ í•„ìš”
    missing_fields: List[str] = field(default_factory=list)
    can_execute: bool = False
    can_execute_after_input: bool = False


FRESHMEN_KEYWORDS = ["ì‹ ì…ìƒ", "ì‹ ì…"]
TEACHER_KEYWORDS  = ["êµì‚¬", "êµì›"]
TRANSFER_KEYWORDS = ["ì „ì…ìƒ", "ì „ì…"]
WITHDRAW_KEYWORDS = ["ì „ì¶œìƒ", "ì „ì¶œ"]



NOTICE_ORDER = [
    "ì‹ ê·œë“±ë¡ - ë©”ì¼",
    "ì‹ ê·œë“±ë¡ - ë¬¸ì",
    "êµì§ì› ë“±ë¡ - ë©”ì¼",
    "ë°˜ì´ë™ - ë©”ì¼",
    "ë°˜ì´ë™ - ë©”ì¼ (ì‹ ì…ìƒ, êµì§ì› ë“±ë¡ & ë°˜ì´ë™)",
    "ë°˜ì´ë™ - ë¬¸ì",
    "2-6í•™ë…„ ëª…ë‹¨ ë³´ë‚´ ì˜¨ ê²½ìš° - ë©”ì¼",
    "2-6í•™ë…„ ë°˜í¸ì„± ìë£Œ ì¬ìš”ì²­ - ë¬¸ì",
]


HANGUL_RE = re.compile(r"[ê°€-í£]")
EN_RE = re.compile(r"[A-Za-z]")


# ìŠ¬ë¡¯ë³„ í—¤ë” í‚¤ì›Œë“œ (ëŠìŠ¨í•œ ë§¤ì¹­)
FRESHMEN_HEADER_SLOTS = {
    "no":    ["no", "ë²ˆí˜¸"],
    "grade": ["í•™ë…„"],
    "class": ["ë°˜", "í•™ê¸‰"],
    "num":   ["ë²ˆí˜¸", "ë²ˆ"],
    "name":  ["ì„±ëª…", "ì´ë¦„", "í•™ìƒì´ë¦„"],
}

TRANSFER_HEADER_SLOTS = {
    "no":    ["no", "ë²ˆí˜¸"],
    "grade": ["í•™ë…„"],
    "class": ["ë°˜", "í•™ê¸‰"],
    "number":["ë²ˆí˜¸", "ë²ˆ", "ì¶œì„ë²ˆí˜¸"],
    "name":  ["ì„±ëª…", "ì´ë¦„"],
    "remark":["ë¹„ê³ ", "ë©”ëª¨", "íŠ¹ì´ì‚¬í•­"],
}

WITHDRAW_HEADER_SLOTS = {
    "no":    ["no", "ë²ˆí˜¸"],
    "grade": ["í•™ë…„"],
    "class": ["ë°˜", "í•™ê¸‰"],
    "name":  ["ì„±ëª…", "ì´ë¦„"],
    "remark":["ë¹„ê³ ", "ë©”ëª¨", "íŠ¹ì´ì‚¬í•­"],
}

TEACHER_HEADER_SLOTS = {
    "no":      ["no", "ë²ˆí˜¸"],
    "position":["ì§ìœ„", "ë‹´ë‹¹", "ì§ìœ„ë‹´ë‹¹"],
    "name":    ["ì„±ëª…", "ì´ë¦„", "ì„ ìƒë‹˜ì´ë¦„", "êµì‚¬ëª…", "êµì›ëª…"],
    "learn":   ["í•™ìŠµìš©idì‹ ì²­", "í•™ìŠµìš©id", "í•™ìŠµìš©", "í•™ìŠµìš©ì•„ì´ë””"],
    "admin":   ["ê´€ë¦¬ìš©idì‹ ì²­", "ê´€ë¦¬ìš©id", "ê´€ë¦¬ìš©", "ê´€ë¦¬ìš©ì•„ì´ë””"],
}


EXAMPLE_NAMES_RAW = ["í™ê¸¸ë™", "ì´ìˆœì‹ ", "ìœ ê´€ìˆœ", "ì„êº½ì •"]
EXAMPLE_NAMES_NORM = {normalize_text(n) for n in EXAMPLE_NAMES_RAW}
EXAMPLE_KEYWORDS = ["ì˜ˆì‹œ"]  # í–‰ ì•ˆ ì–´ëŠ ì…€ì´ë¼ë„ 'ì˜ˆì‹œ' í¬í•¨ë˜ë©´ ì˜ˆì‹œë¡œ ì²˜ë¦¬


FILL_TRANSFER = PatternFill("solid", fgColor="F8CBAD")  # ì˜…ì€ ì£¼í™©
FILL_DUP      = PatternFill("solid", fgColor="FFFF00")  # ë…¸ë‘
FILL_GREY     = PatternFill("solid", fgColor="D9D9D9")  # íšŒìƒ‰




# ========== L0: infra / excel utils ==========

def ensure_xlsx_only(p: Path) -> None:
    if p.suffix.lower() != ".xlsx":
        raise ValueError(f"[ì˜¤ë¥˜] íŒŒì¼ í˜•ì‹ì´ .xlsxê°€ ì•„ë‹™ë‹ˆë‹¤: {p.name}")

def backup_if_exists(out_path: Path) -> Optional[Path]:
    """ê¸°ì¡´ íŒŒì¼ì´ ìˆìœ¼ë©´ ì‘ì—…/_backupìœ¼ë¡œ ì´ë™."""
    out_path = Path(out_path)
    if not out_path.exists():
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = out_path.parent / "_backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    dest = backup_dir / f"{out_path.stem}_{ts}{out_path.suffix}"
    out_path.replace(dest)
    return dest

def safe_load_workbook(xlsx_path: Path, data_only: bool = True):
    try:
        return load_workbook(xlsx_path, data_only=data_only)
    except TypeError as e:
        msg = str(e)
        if "openpyxl.packaging.custom" not in msg or "NoneType" not in msg:
            raise

        buffer = BytesIO()
        with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(
            buffer, "w", compression=zipfile.ZIP_DEFLATED
        ) as zout:
            for item in zin.infolist():
                if item.filename == "docProps/custom.xml":
                    root = ET.fromstring(zin.read(item.filename))
                    ns = "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
                    tag = f"{{{ns}}}property"
                    for prop in list(root.findall(tag)):
                        name = prop.get("name")
                        if name is None or str(name).strip() == "":
                            root.remove(prop)
                    new_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    zout.writestr(item, new_xml)
                else:
                    zout.writestr(item, zin.read(item.filename))

        buffer.seek(0)
        return load_workbook(buffer, data_only=data_only)

    except IndexError as e:
        # ìŠ¤íƒ€ì¼ ì¸ë±ìŠ¤ ê¼¬ì—¬ì„œ ë‚˜ëŠ” openpyxl ë²„ê·¸ íšŒí”¼ìš©
        # í…œí”Œë¦¿ ì €ì¥ì— ì“°ì¼ ì¼ ìˆëŠ” ì¼€ì´ìŠ¤(data_only=False)ëŠ” ê·¸ëŒ€ë¡œ ì˜¬ë ¤ë³´ë‚´ê³ ,
        # ì¸í’‹ ì½ê¸°ìš©(data_only=True)ì¼ ë•Œë§Œ read_only ëª¨ë“œë¡œ ë‹¤ì‹œ ì‹œë„
        if not data_only:
            raise
        return load_workbook(xlsx_path, data_only=data_only, read_only=True)

def header_map(ws, header_row: int = 1):
    mapping = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        key = str(cell.value)
        key = key.replace("\u00A0", " ")
        key = re.sub(r"\s+", "", key)
        key = key.replace(".", "")
        mapping[key] = cell.column
    return mapping

def write_text_cell(ws, row: int, col: int, value: Any):
    """
    ê°’ì€ ê·¸ëŒ€ë¡œ ë¬¸ìì—´ë¡œ ë„£ê³ , ì…€ íƒ€ì…/ì„œì‹ì€ í…ìŠ¤íŠ¸ë¡œ ê°•ì œ.
    - 3-1, 01, 010-1234 ê°™ì€ ê²ƒë“¤ ë‚ ì§œ/ìˆ«ìë¡œ ì•ˆ ë°”ë€Œê²Œ ë§‰ê¸° ìœ„í•¨.
    """
    cell = ws.cell(row=row, column=col)
    cell.value = "" if value is None else str(value)
    cell.data_type = "s"
    cell.number_format = "@"
    return cell

def find_last_data_row(ws, key_col: int, start_row: int) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is not None and str(v).strip() != "":
            last = r
    return last

def clear_sheet_rows(ws, start_row=2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

def move_sheet_after(wb, sheet_name: str, after_name: str):
    if sheet_name not in wb.sheetnames or after_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    wb._sheets.remove(ws)
    idx = wb.sheetnames.index(after_name)
    wb._sheets.insert(idx + 1, ws)

def delete_rows_below(ws, last_keep_row: int):
    if ws.max_row > last_keep_row:
        ws.delete_rows(last_keep_row + 1, ws.max_row - last_keep_row)

def clear_format_workbook_from_row(wb, start_row: int = 2):
    """
    ëª¨ë“  ì‹œíŠ¸ì—ì„œ:
    - start_rowë¶€í„° ì‹¤ì œ ë°ì´í„°ê°€ ìˆëŠ” ë§ˆì§€ë§‰ í–‰ê¹Œì§€ ìŠ¤ìº”
    - ê·¸ ì•„ë˜ í–‰ë“¤ì— ëŒ€í•´ì„œë§Œ ì„œì‹(fill, border) ì œê±°
    """
    for ws in wb.worksheets:
        last_data_row = 0
        max_row = ws.max_row
        max_col = ws.max_column or 1

        # ì‹¤ì œ ë°ì´í„° ë§ˆì§€ë§‰ í–‰ ì°¾ê¸°
        for r in range(start_row, max_row + 1):
            row_has_value = False
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip() != "":
                    row_has_value = True
                    break
            if row_has_value:
                last_data_row = r

        if last_data_row == 0:
            continue

        # ë§ˆì§€ë§‰ ë°ì´í„° í–‰ ì•„ë˜ë¶€í„° ì„œì‹ ì œê±°
        for r in range(last_data_row + 1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()

def reset_view_to_a1(wb):
    """
    - ëª¨ë“  ì‹œíŠ¸: í™”ë©´ì€ A1, ì»¤ì„œëŠ” A2
    - ëª¨ë“  ì‹œíŠ¸: 1í–‰ ê³ ì •(freeze_panes = A2)
    - ëª¨ë“  ì‹œíŠ¸: ê·¸ë£¹ ì„ íƒ(tabSelected) í•´ì œ
    - í†µí•©ë¬¸ì„œ: ì²« ë²ˆì§¸ ì‹œíŠ¸ë§Œ ì„ íƒ + í™œì„±
    """
    # 1) ê³µí†µ ë·°/ê³ ì • ì„¤ì •
    for ws in wb.worksheets:
        sv = ws.sheet_view

        # í™”ë©´/ì»¤ì„œ
        sv.topLeftCell = "A1"
        sv.activeCell = "A2"
        sv.selection = [Selection(activeCell="A2", sqref="A2")]

        # 1í–‰ ê³ ì •
        ws.freeze_panes = "A2"

        # ì‹œíŠ¸ ê·¸ë£¹ ì„ íƒ í’€ê¸°
        if hasattr(sv, "tabSelected"):
            sv.tabSelected = False

    # 2) ì²« ë²ˆì§¸ ì‹œíŠ¸ë§Œ ì„ íƒ + í™œì„±
    first_ws = wb.worksheets[0]
    if hasattr(first_ws.sheet_view, "tabSelected"):
        first_ws.sheet_view.tabSelected = True

    wb.active = 0

    # 3) í†µí•©ë¬¸ì„œ ë·°ë„ ì²« ì‹œíŠ¸ ê¸°ì¤€ìœ¼ë¡œ í†µì¼
    if getattr(wb, "views", None):
        views = wb.views
        if views:
            views[0].activeTab = 0
            views[0].firstSheet = 0





# ========== L1: domain utils (names / headers / examples) ==========

def normalize_name(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-zê°€-í£\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""

    has_ko = bool(HANGUL_RE.search(s))
    has_en = bool(EN_RE.search(s))

    if has_ko and not has_en:
        return s.replace(" ", "")

    if has_en and not has_ko:
        parts = [p for p in s.split(" ") if p]
        parts = [p.lower().capitalize() for p in parts]
        return "".join(parts)

    if has_ko and has_en:
        def _fix_en(m: re.Match) -> str:
            tok = m.group(0).lower()
            return tok[0].upper() + tok[1:] if tok else tok
        s2 = re.sub(r"[A-Za-z]+", _fix_en, s)
        return s2.replace(" ", "")

    return ""

def normalize_name_key(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-zê°€-í£\s]", "", s)
    s = re.sub(r"\s+", "", s)
    return s.casefold()

def english_casefold_key(name: str) -> str:
    if name is None:
        return ""
    return str(name).strip().casefold()

def dedup_suffix_letters(n: int) -> str:
    if n <= 0:
        return ""
    out = ""
    while n > 0:
        n -= 1
        out = chr(ord("A") + (n % 26)) + out
        n //= 26
    return out

def apply_suffix_for_duplicates(names: List[str]) -> List[str]:
    total = {}
    for nm in names:
        key = english_casefold_key(nm)
        total[key] = total.get(key, 0) + 1

    seen = {}
    out = []
    for nm in names:
        key = english_casefold_key(nm)
        if total.get(key, 0) <= 1:
            out.append(nm)
            continue
        seen[key] = seen.get(key, 0) + 1
        out.append(nm + dedup_suffix_letters(seen[key]))
    return out

def _strip_korean_suffix_for_notice(raw_name: Any) -> str:
    """
    ì•ˆë‚´íŒŒì¼ ë™ëª…ì´ì¸ íŒì •ìš©:
    - 'ê¹€ì„œí˜„A', 'ê¹€ì„œí˜„B' ê°™ì´ 'í•œê¸€+ëŒ€ë¬¸ì' íŒ¨í„´ì´ë©´ ë’¤ì˜ ëŒ€ë¬¸ìë§Œ ë–¼ê³  ë³¸ ì´ë¦„ë§Œ ë‚¨ê¹€
    - ì˜ì–´ ì´ë¦„(James, Anna ë“±)ì€ ê·¸ëŒ€ë¡œ ìœ ì§€ (ë§¨ ëì´ ëŒ€ë¬¸ìì—¬ë„ í•œê¸€ì´ ì—†ìœ¼ë©´ ì•ˆ ê¹ìŒ)
    """
    if raw_name is None:
        return ""
    s = str(raw_name).strip()
    if not s:
        return ""

    has_ko = bool(HANGUL_RE.search(s))

    # í•œê¸€ì´ ìˆê³ , ëì´ ëŒ€ë¬¸ìì´ê³ , ê¸¸ì´ê°€ ì–´ëŠ ì •ë„ ì´ìƒì´ë©´ suffixë¡œ ê°„ì£¼
    if has_ko and re.search(r"[A-Z]$", s) and len(s) >= 3:
        # ë§¨ ë ì—°ì† ëŒ€ë¬¸ìë§Œ ì œê±° (ì˜ˆ: 'ê¹€ì„œí˜„A', 'ê¹€ì„œí˜„AB')
        s = re.sub(r"[A-Z]+$", "", s).strip()

    return s

def notice_name_key(raw_name: Any) -> str:
    """
    ì•ˆë‚´íŒŒì¼ ë™ëª…ì´ì¸ íŒì •ìš© ìµœì¢… í‚¤:
    - í•œê¸€+A/B/C suffix ì œê±° í›„ normalize_name_key ì ìš©
    """
    base = _strip_korean_suffix_for_notice(raw_name)
    return normalize_name_key(base)

def _normalize_header_cell(val: Any) -> str:
    """
    ì—‘ì…€ í—¤ë” ì…€ ì •ê·œí™”:
    - None â†’ ""
    - ê³µë°±/ì¤„ë°”ê¿ˆ/nbsp ì œê±°
    - ë§ˆì¹¨í‘œ(.) ì œê±°
    - ì†Œë¬¸ì ë³€í™˜
    """
    if val is None:
        s = ""
    else:
        s = str(val)

    # nbsp â†’ ì¼ë°˜ ê³µë°±
    s = s.replace("\u00A0", " ")
    # ëª¨ë“  ê³µë°± ì œê±°
    s = re.sub(r"\s+", "", s)
    # ë§ˆì¹¨í‘œ ì œê±°
    s = s.replace(".", "")
    # ì†Œë¬¸ì
    s = s.lower()
    return s

def _build_header_slot_map(ws, header_row: int, slots: Dict[str, List[str]]) -> Dict[str, int]:
    """
    slots ì •ì˜(FRESHMEN_HEADER_SLOTS ë“±)ë¥¼ ê¸°ì¤€ìœ¼ë¡œ
    ì‹¤ì œ ì—‘ì…€ í—¤ë” í–‰ì—ì„œ ê° slotì´ ì–´ëŠ ì»¬ëŸ¼ì— ìˆëŠ”ì§€ ì°¾ì•„ì„œ
    {slot: col_idx} í˜•íƒœë¡œ ë°˜í™˜.
    """
    # header_map: {í—¤ë”í…ìŠ¤íŠ¸ -> column index}
    hm = header_map(ws, header_row)

    # í—¤ë” í…ìŠ¤íŠ¸ë¥¼ ì •ê·œí™”í•´ì„œ ë¹„êµ (ê³µë°±/ë§ˆì¹¨í‘œ ì œê±°, ì†Œë¬¸ì)
    norm_to_col: Dict[str, int] = {}
    for raw_key, col in hm.items():
        norm_key = _normalize_header_cell(raw_key)
        if norm_key:
            norm_to_col[norm_key] = col

    result: Dict[str, int] = {}

    for slot, patterns in slots.items():
        for pat in patterns:
            pat_norm = _normalize_header_cell(pat)
            if not pat_norm:
                continue
            # í—¤ë” ì •ê·œí™” ë¬¸ìì—´ ì•ˆì— íŒ¨í„´ì´ í¬í•¨ë˜ë©´ ë§¤ì¹­
            for header_norm, col in norm_to_col.items():
                if pat_norm in header_norm:
                    result[slot] = col
                    break
            if slot in result:
                break

    return result

def _detect_header_row_generic(ws, slots: Dict[str, List[str]],
                               max_search_row: int = 15,
                               max_col: int = 10,
                               min_match_slots: int = 3) -> int:
    """
    slots: {slot_name: [pattern1, pattern2, ...]}
    í•œ í–‰ì—ì„œ slotì´ ëª‡ ê°œ ë§¤ì¹­ë˜ëŠ”ì§€ ë³´ê³ , min_match_slots ì´ìƒì´ë©´ í—¤ë” í›„ë³´ë¡œ ë³¸ë‹¤.
    """
    best_row: Optional[int] = None
    best_score: int = 0

    for row in ws.iter_rows(min_row=1, max_row=max_search_row):
        row_idx = row[0].row
        vals = [_normalize_header_cell(c.value) for c in row[:max_col]]

        matched_slots = set()
        for slot, patterns in slots.items():
            for pat in patterns:
                pat_norm = _normalize_header_cell(pat)
                if not pat_norm:
                    continue
                if any(pat_norm in v for v in vals if v):
                    matched_slots.add(slot)
                    break  # ìŠ¬ë¡¯ ë‹¹ 1íšŒë§Œ ì¹´ìš´íŠ¸

        score = len(matched_slots)
        if score > best_score:
            best_score = score
            best_row = row_idx

    if best_row is None or best_score < min_match_slots:
        raise ValueError(
            "[ì˜¤ë¥˜] ì¸í’‹ íŒŒì¼ì—ì„œ í—¤ë”ë¥¼ ìë™ìœ¼ë¡œ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. "
            "í—¤ë” í–‰ì— í•™ë…„/ë°˜/ì´ë¦„/NO ë“±ì˜ í‚¤ì›Œë“œê°€ ë™ì‹œì— 3ê°œ ì´ìƒ ìˆì–´ì•¼ í•©ë‹ˆë‹¤."
        )

    return best_row

def detect_header_row_freshmen(ws) -> int:
    return _detect_header_row_generic(ws, FRESHMEN_HEADER_SLOTS,
                                      max_search_row=15, max_col=10, min_match_slots=3)

def detect_header_row_transfer(ws) -> int:
    return _detect_header_row_generic(ws, TRANSFER_HEADER_SLOTS,
                                      max_search_row=15, max_col=10, min_match_slots=3)

def detect_header_row_withdraw(ws) -> int:
    return _detect_header_row_generic(ws, WITHDRAW_HEADER_SLOTS,
                                      max_search_row=15, max_col=10, min_match_slots=3)

def detect_header_row_teacher(ws) -> int:
    # êµì‚¬ëŠ” NO / ì´ë¦„ / í•™ë…„/ë°˜ / ì‹ ì²­ ì»¬ëŸ¼ ì¤‘ ìµœì†Œ 3ìŠ¬ë¡¯ ì´ìƒ
    return _detect_header_row_generic(ws, TEACHER_HEADER_SLOTS,
                                      max_search_row=15, max_col=10, min_match_slots=3)

def _row_is_empty(ws, row: int, max_col: Optional[int] = None) -> bool:
    if max_col is None:
        max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and str(v).strip() != "":
            return False
    return True

def _row_has_example_keyword(ws, row: int, max_col: Optional[int] = None) -> bool:
    if max_col is None:
        max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        s = normalize_text(str(v))
        if not s:
            continue
        for kw in EXAMPLE_KEYWORDS:
            if kw in s:
                return True
    return False

def _cell_is_example_name(value: Any) -> bool:
    if value is None:
        return False
    s = normalize_text(str(value))
    return bool(s) and s in EXAMPLE_NAMES_NORM

def detect_example_and_data_start(
    ws,
    header_row: int,
    name_col: int,
    max_search_row: Optional[int] = None,
    max_col: Optional[int] = None,
) -> Tuple[List[int], int]:
    """
    í—¤ë” ì•„ë˜ì—ì„œ ì˜ˆì‹œ í–‰(0ê°œ ì´ìƒ)ê³¼ ì‹¤ì œ ë°ì´í„° ì‹œì‘ í–‰ì„ ìë™ ê°ì§€í•œë‹¤.

    - header_row ë°”ë¡œ ì•„ë˜ í–‰ë¶€í„° ìŠ¤ìº”
    - ì™„ì „ ë¹ˆ í–‰ì€ ê±´ë„ˆëœ€
    - 'ì˜ˆì‹œ' í‚¤ì›Œë“œê°€ ìˆê±°ë‚˜ ì´ë¦„ ì¹¸ì´ ì˜ˆì‹œ ì´ë¦„ì´ë©´ â†’ ì˜ˆì‹œ í–‰
    - ê·¸ ì™¸ ì²« ë²ˆì§¸ ë¹„-ì˜ˆì‹œ í–‰ â†’ ì‹¤ì œ ë°ì´í„° ì‹œì‘ í–‰
    """
    if max_search_row is None:
        max_search_row = ws.max_row

    example_rows: List[int] = []
    r = header_row + 1

    while r <= max_search_row:
        # 1) ì™„ì „ ë¹ˆ í–‰ì€ ìŠ¤í‚µ
        if _row_is_empty(ws, r, max_col=max_col):
            r += 1
            continue

        # 2) í–‰ ì•ˆì— 'ì˜ˆì‹œ' í‚¤ì›Œë“œ ìˆìœ¼ë©´ ì˜ˆì‹œ
        if _row_has_example_keyword(ws, r, max_col=max_col):
            example_rows.append(r)
            r += 1
            continue

        # 3) ì´ë¦„ ì¹¸ì´ ì˜ˆì‹œ ì´ë¦„ì´ë©´ ì˜ˆì‹œ
        v_name = ws.cell(row=r, column=name_col).value
        if _cell_is_example_name(v_name):
            example_rows.append(r)
            r += 1
            continue

        # 4) ì—¬ê¸°ê¹Œì§€ ì•ˆ ê±¸ë¦¬ë©´ â†’ ì‹¤ì œ ë°ì´í„° ì‹œì‘
        return example_rows, r

    raise ValueError(
        f"[ì˜¤ë¥˜] ë°ì´í„° ì‹œì‘ í–‰ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. í—¤ë”({header_row}í–‰) ì•„ë˜ì— ì˜ˆì‹œë‚˜ ì‹¤ì œ ë°ì´í„°ë¡œ ë³´ì´ëŠ” í–‰ì´ ì—†ìŠµë‹ˆë‹¤."
    )

def detect_input_layout(xlsx_path: Path, kind: str) -> Dict[str, Any]:
    """
    UIì—ì„œ ì¸í’‹ íŒŒì¼ êµ¬ì¡°ë¥¼ ë¯¸ë¦¬ ë³´ì—¬ì¤„ ë•Œ ì‚¬ìš©.
    kind: 'freshmen' | 'transfer' | 'withdraw' | 'teacher'
    ë°˜í™˜:
      {
        "header_row": int,
        "example_rows": [int, ...],
        "data_start_row": int,
      }
    """
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    kind_norm = (kind or "").strip().lower()

    # 1) í—¤ë” ìë™ ê°ì§€
    if kind_norm == "freshmen":
        header_row = detect_header_row_freshmen(ws)
        # ğŸ”¹ í—¤ë” ì´ë¦„ ê¸°ì¤€ìœ¼ë¡œ ì‹¤ì œ ì´ë¦„ ì»¬ëŸ¼ ì°¾ê¸°
        slot_cols = _build_header_slot_map(ws, header_row, FRESHMEN_HEADER_SLOTS)
        name_col = slot_cols.get("name", 5)  # ëª» ì°¾ìœ¼ë©´ ê¸°ì¡´ Eì—´ fallback
   
    elif kind_norm == "transfer":
        header_row = detect_header_row_transfer(ws)
        name_col = 5  # Eì—´: ì„±ëª…
    elif kind_norm == "withdraw":
        header_row = detect_header_row_withdraw(ws)
        name_col = 4  # Dì—´: ì„±ëª…
    elif kind_norm == "teacher":
        header_row = detect_header_row_teacher(ws)
        name_col = 3  # Cì—´: ì„ ìƒë‹˜ ì´ë¦„
    else:
        raise ValueError(f"[ì˜¤ë¥˜] ì§€ì›í•˜ì§€ ì•ŠëŠ” kind ê°’ì…ë‹ˆë‹¤: {kind}")

    # 2) ì˜ˆì‹œ í–‰ + ì‹¤ì œ ë°ì´í„° ì‹œì‘ í–‰ ìë™ ê°ì§€
    example_rows, data_start_row = detect_example_and_data_start(
        ws,
        header_row=header_row,
        name_col=name_col,
    )

    return {
        "header_row": header_row,
        "example_rows": example_rows,
        "data_start_row": data_start_row,
    }

def normalize_withdraw_class(raw, grade: int) -> str:
    """
    ì „ì¶œ ëª…ë‹¨ Cì—´(ë°˜) ë¬¸ìì—´ì„ í†µì¼ëœ í˜•ì‹ìœ¼ë¡œ ì •ê·œí™”:
    - í•™ë…„ ì •ë³´ëŠ” ë¬´ì‹œí•˜ê³ , ë¬¸ìì—´ì—ì„œ 'ë§ˆì§€ë§‰ ìˆ«ì ë¬¶ìŒ'ì„ ë°˜ ë²ˆí˜¸ë¡œ ì‚¬ìš©í•œë‹¤.
      ì˜ˆ)
        '1-10'          -> grade-10ë°˜
        '1í•™ë…„10ë°˜'     -> grade-10ë°˜
        '1 í•™ë…„ 10 ë°˜'  -> grade-10ë°˜
        '10ë°˜' / '10'   -> grade-10ë°˜
        '3-5ë°˜'         -> grade-5ë°˜
    """
    if raw is None:
        return ""

    s = str(raw).strip()
    if not s:
        return ""

    # ê³µë°±/ì „ê° ê³µë°± ì •ë¦¬
    s = s.replace("\u3000", " ").replace("\u00A0", " ")
    s = re.sub(r"\s+", "", s)

    # ë¬¸ìì—´ ì•ˆì˜ ìˆ«ì ë©ì–´ë¦¬ë“¤ ì „ë¶€ ì¶”ì¶œ
    nums = re.findall(r"\d+", s)
    if not nums:
        # ìˆ«ìê°€ ì „í˜€ ì—†ìœ¼ë©´ ê·¸ëƒ¥ ì›ë³¸ ë°˜í™˜ (ìµœì†Œí•œ ì´ìƒí•œ ê°’ì´ë¼ëŠ” ê±¸ ëˆˆìœ¼ë¡œ ë³´ê²Œ)
        return s

    # "ë§ˆì§€ë§‰ ìˆ«ì ë¬¶ìŒ"ì„ ë°˜ ë²ˆí˜¸ë¡œ ì‚¬ìš©
    class_no = int(nums[-1])

    return f"{grade}-{class_no}ë°˜"





# ========== L2: input readers (ì‹ ì…/ì „ì…/ì „ì¶œ/êµì‚¬) ==========

def read_freshmen_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    """
    ì‹ ì…ìƒ íŒŒì¼ì„ ì½ì–´ì„œ
    [{"grade": í•™ë…„(int), "class": ë°˜(str), "number": ë²ˆí˜¸(str), "name": ì´ë¦„(str)}, ...]
    í˜•íƒœë¡œ ë°˜í™˜.

    - í•™ë…„ / ë°˜ / ì´ë¦„ì€ í•„ìˆ˜
    - ë²ˆí˜¸ëŠ” ì—†ì–´ë„ ë¨(ë¹ˆ ë¬¸ìì—´ë¡œ ì²˜ë¦¬)
    - ì»¬ëŸ¼ ìœ„ì¹˜ëŠ” ê³ ì •(B,C,D,E)ì´ ì•„ë‹ˆë¼ í—¤ë” ì´ë¦„ ê¸°ì¤€ìœ¼ë¡œ íƒìƒ‰
    """
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # 1) í—¤ë” ìë™ ê°ì§€ (í•„ìš” ì‹œ)
    if header_row is None:
        header_row = detect_header_row_freshmen(ws)

    # 1-1) í—¤ë”ì—ì„œ ì‹¤ì œ ì»¬ëŸ¼ ìœ„ì¹˜ ì°¾ê¸°
    slot_cols = _build_header_slot_map(ws, header_row, FRESHMEN_HEADER_SLOTS)
    col_grade = slot_cols.get("grade")
    col_class = slot_cols.get("class")
    # ë²ˆí˜¸ëŠ” num(ë²ˆí˜¸/ë²ˆ) ìš°ì„ , ì—†ìœ¼ë©´ no(ë²ˆí˜¸)ë¼ë„ ì‚¬ìš©
    col_num = slot_cols.get("num") or slot_cols.get("no")
    col_name = slot_cols.get("name")

    missing = []
    if col_grade is None:
        missing.append("í•™ë…„")
    if col_class is None:
        missing.append("ë°˜")
    if col_name is None:
        missing.append("ì„±ëª…/ì´ë¦„")

    # í•™ë…„/ë°˜/ì´ë¦„ì€ í•„ìˆ˜, ë²ˆí˜¸ëŠ” ì—†ì–´ë„ ë¨
    if missing:
        raise ValueError(
            "[ì˜¤ë¥˜] ì‹ ì…ìƒ íŒŒì¼ í—¤ë”ì—ì„œ "
            + ", ".join(missing)
            + " ì—´ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. 'í•™ë…„', 'ë°˜', 'ì´ë¦„' í—¤ë”ë¥¼ ì¶”ê°€í•˜ê±°ë‚˜ ìˆ˜ì •í•´ ì£¼ì„¸ìš”."
        )

    # 2) ì˜ˆì‹œ/ë°ì´í„° ì‹œì‘ í–‰ ìë™ ê°ì§€ (ì‚¬ìš©ìê°€ data_start_row ì§ì ‘ ì¤€ ê²½ìš° ìš°ì„ )
    if data_start_row is None:
        _, data_start_row = detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=col_name,
        )

    out: List[Dict[str, Any]] = []
    row = data_start_row
    while True:
        grade = ws.cell(row=row, column=col_grade).value
        cls   = ws.cell(row=row, column=col_class).value
        num   = ws.cell(row=row, column=col_num).value if col_num is not None else None
        name  = ws.cell(row=row, column=col_name).value

        # 1) í–‰ ì „ì²´ê°€ ë¹„ì–´ ìˆìœ¼ë©´ ì¢…ë£Œ
        if all(v is None or str(v).strip() == "" for v in [grade, cls, num, name]):
            break

        # 2) í•„ìˆ˜ê°’(í•™ë…„, ë°˜, ì´ë¦„) ì²´í¬
        if any(v is None or str(v).strip() == "" for v in [grade, cls, name]):
            raise ValueError(
                f"[ì˜¤ë¥˜] ì‹ ì…ìƒ íŒŒì¼ {row}í–‰ì—ì„œ í•™ë…„/ë°˜/ì´ë¦„ ì¤‘ ë¹„ì–´ ìˆëŠ” ê°’ì´ ìˆìŠµë‹ˆë‹¤."
            )

        # 3) í•™ë…„ì—ì„œ ìˆ«ìë§Œ ì¶”ì¶œ
        grade_s = str(grade).strip()
        m = re.search(r"\d+", grade_s)
        if not m:
            raise ValueError(
                f"[ì˜¤ë¥˜] ì‹ ì…ìƒ íŒŒì¼ {row}í–‰ í•™ë…„ì—ì„œ ìˆ«ìë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤: {grade_s!r}"
            )
        grade_i = int(m.group(0))

        cls_s = str(cls).strip()
        num_s = "" if (num is None or str(num).strip() == "") else str(num).strip()
        name_n = normalize_name(name)
        if not name_n:
            raise ValueError(
                f"[ì˜¤ë¥˜] ì‹ ì…ìƒ íŒŒì¼ {row}í–‰ ì´ë¦„ ì •ê·œí™” ê²°ê³¼ê°€ ë¹„ì–´ ìˆìŠµë‹ˆë‹¤."
            )

        out.append(
            {
                "grade": grade_i,
                "class": cls_s,
                "number": num_s,
                "name": name_n,
            }
        )
        row += 1

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, x)

    out.sort(
        key=lambda r: (
            r["grade"],
            _safe_int(r["class"]),
            _safe_int(r["number"]),
        )
    )
    return out

def read_transfer_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    """
    ì „ì…ìƒ ì—‘ì…€ì—ì„œ í•™ë…„/ë°˜/ë²ˆí˜¸/ì´ë¦„ë§Œ ì½ì–´ì˜¨ë‹¤.
    - IDëŠ” ì „í˜€ ì‚¬ìš©í•˜ì§€ ì•ŠëŠ”ë‹¤ (ìŠ¬ë¡¯ë„ ë‘ì§€ ì•ŠìŒ).
    - í—¤ë” ë§¤í•‘ì€ TRANSFER_HEADER_SLOTSë§Œ ì‚¬ìš©.
    """
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # 1) í—¤ë” í–‰ íƒì§€
    if header_row is None:
        header_row = detect_header_row_transfer(ws)

    # 2) í—¤ë” â†’ ì»¬ëŸ¼ ë§¤í•‘ (ìŠ¬ë¡¯ì€ TRANSFER_HEADER_SLOTSì— ì •ì˜ëœ ê²ƒë§Œ ì‚¬ìš©)
    slot_cols = _build_header_slot_map(ws, header_row, TRANSFER_HEADER_SLOTS)

    col_grade = slot_cols.get("grade")
    col_class = slot_cols.get("class")
    col_num   = slot_cols.get("num")   # ë²ˆí˜¸ëŠ” ìˆìœ¼ë©´ ì‚¬ìš©, ì—†ì–´ë„ ë¨
    col_name  = slot_cols.get("name")

    missing = []
    if col_grade is None:
        missing.append("í•™ë…„")
    if col_class is None:
        missing.append("ë°˜")
    if col_name is None:
        missing.append("ì´ë¦„")

    # ë²ˆí˜¸ëŠ” í•„ìˆ˜ ì•„ë‹˜
    if missing:
        raise ValueError(
            "[ì˜¤ë¥˜] ì „ì…ìƒ íŒŒì¼ í—¤ë”ì—ì„œ "
            + ", ".join(missing)
            + " ì—´ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. í—¤ë”ëª…ì„ í™•ì¸í•´ ì£¼ì„¸ìš”."
        )

    # 3) ë°ì´í„° ì‹œì‘ í–‰
    if data_start_row is None:
        _, data_start_row = detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=col_name,
        )

    out: List[Dict[str, Any]] = []
    row = data_start_row

    while True:
        grade = ws.cell(row=row, column=col_grade).value
        cls   = ws.cell(row=row, column=col_class).value
        num   = ws.cell(row=row, column=col_num).value if col_num is not None else None
        name  = ws.cell(row=row, column=col_name).value

        # ì™„ì „ ë¹ˆ ì¤„ì´ë©´ ì¢…ë£Œ (í•™ë…„/ë°˜/ë²ˆí˜¸/ì´ë¦„ ì „ë¶€ ë¹„ì–´ ìˆìœ¼ë©´)
        if all(
            v is None or str(v).strip() == ""
            for v in [grade, cls, num, name]
        ):
            break

        # í•„ìˆ˜ê°’ ì²´í¬: í•™ë…„/ë°˜/ì´ë¦„ë§Œ í•„ìˆ˜
        if any(
            v is None or str(v).strip() == ""
            for v in [grade, cls, name]
        ):
            raise ValueError(
                f"[ì˜¤ë¥˜] ì „ì…ìƒ íŒŒì¼ {row}í–‰ì—ì„œ í•™ë…„/ë°˜/ì´ë¦„ ì¤‘ ë¹„ì–´ ìˆëŠ” ê°’ì´ ìˆìŠµë‹ˆë‹¤."
            )

        # í•™ë…„ ìˆ«ì ì¶”ì¶œ
        grade_s = str(grade).strip()
        m = re.search(r"\d+", grade_s)
        if not m:
            raise ValueError(
                f"[ì˜¤ë¥˜] ì „ì…ìƒ íŒŒì¼ {row}í–‰ í•™ë…„ì—ì„œ ìˆ«ìë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤: {grade_s!r}"
            )
        grade_i = int(m.group(0))

        cls_s = str(cls).strip()
        num_s = "" if (num is None or str(num).strip() == "") else str(num).strip()
        name_n = normalize_name(name)

        out.append(
            {
                "grade": grade_i,
                "class": cls_s,
                "number": num_s,
                "name": name_n,
                # âš ï¸ IDëŠ” ì „ì…ì—ì„œ ì ˆëŒ€ ì“°ì§€ ì•ŠëŠ”ë‹¤ -> í‚¤ ìì²´ë¥¼ ë§Œë“¤ì§€ ì•ŠìŒ
            }
        )
        row += 1

    # í•„ìš”í•˜ë©´ ì—¬ê¸°ì„œ grade / class / number ê¸°ì¤€ ì •ë ¬ ê°€ëŠ¥
    return out

def read_teacher_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    """
    êµì‚¬ íŒŒì¼ ì½ê¸°.
    í—¤ë” ì´ë¦„ì„ ê¸°ì¤€ìœ¼ë¡œ ì§ìœ„/ì´ë¦„/í•™ìŠµìš©IDì‹ ì²­/ê´€ë¦¬ìš©IDì‹ ì²­ ì»¬ëŸ¼ì„ ì°¾ëŠ”ë‹¤.
    """
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # 1) í—¤ë” ìë™ ê°ì§€
    if header_row is None:
        header_row = detect_header_row_teacher(ws)

    # 2) í—¤ë” â†’ ì»¬ëŸ¼ ë§¤í•‘
    slot_cols = _build_header_slot_map(ws, header_row, TEACHER_HEADER_SLOTS)
    col_pos   = slot_cols.get("position")  # ì§ìœ„/ë‹´ë‹¹ (ì—†ì–´ë„ ë¨)
    col_name  = slot_cols.get("name")      # ì´ë¦„ (í•„ìˆ˜)
    col_learn = slot_cols.get("learn")     # í•™ìŠµìš© ID ì‹ ì²­ (ì—†ìœ¼ë©´ False)
    col_admin = slot_cols.get("admin")     # ê´€ë¦¬ìš© ID ì‹ ì²­ (ì—†ìœ¼ë©´ False)

    if col_name is None:
        raise ValueError(
            "[ì˜¤ë¥˜] êµì‚¬ íŒŒì¼ í—¤ë”ì—ì„œ ì´ë¦„ ì—´ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. 'ì„±ëª…' ë˜ëŠ” 'ì´ë¦„' í—¤ë”ë¥¼ í™•ì¸í•´ ì£¼ì„¸ìš”."
        )

    # 3) ì˜ˆì‹œ/ë°ì´í„° ì‹œì‘ í–‰ ìë™ ê°ì§€
    if data_start_row is None:
        _, data_start_row = detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=col_name,
        )

    out: List[Dict[str, Any]] = []
    row = data_start_row
    while True:
        # í˜„ì¬ í–‰ ê°’ë“¤ ì½ê¸°
        def _get(col_idx: Optional[int]):
            if col_idx is None:
                return None
            return ws.cell(row=row, column=col_idx).value

        pos    = _get(col_pos)
        name   = _get(col_name)
        v_learn = _get(col_learn)
        v_admin = _get(col_admin)

        # ì™„ì „ ë¹ˆ ì¤„ì´ë©´ ì¢…ë£Œ
        if all(
            v is None or str(v).strip() == ""
            for v in [pos, name, v_learn, v_admin]
        ):
            break

        # ì´ë¦„ ì—†ìœ¼ë©´ ê·¸ í–‰ì€ ê±´ë„ˆëœ€
        if name is None or str(name).strip() == "":
            row += 1
            continue

        name_n = normalize_name(name)
        if not name_n:
            row += 1
            continue

        learn_apply = False
        admin_apply = False
        if col_learn is not None:
            learn_apply = not (v_learn is None or str(v_learn).strip() == "")
        if col_admin is not None:
            admin_apply = not (v_admin is None or str(v_admin).strip() == "")

        out.append(
            {
                "position": "" if pos is None else str(pos).strip(),
                "name": name_n,
                "learn_apply": learn_apply,
                "admin_apply": admin_apply,
            }
        )
        row += 1

    return out

def read_withdraw_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    """
    ì „ì¶œìƒ íŒŒì¼ ì½ê¸°.
    í—¤ë” ì´ë¦„ì„ ê¸°ì¤€ìœ¼ë¡œ í•™ë…„/ë°˜/ì´ë¦„ ì»¬ëŸ¼ì„ ì°¾ëŠ”ë‹¤.
    """
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # 1) í—¤ë” ìë™ ê°ì§€
    if header_row is None:
        header_row = detect_header_row_withdraw(ws)

    # 2) í—¤ë” â†’ ì»¬ëŸ¼ ë§¤í•‘
    slot_cols = _build_header_slot_map(ws, header_row, WITHDRAW_HEADER_SLOTS)
    col_grade = slot_cols.get("grade")
    col_class = slot_cols.get("class")
    col_name  = slot_cols.get("name")

    missing = []
    if col_grade is None:
        missing.append("í•™ë…„")
    if col_class is None:
        missing.append("ë°˜")
    if col_name is None:
        missing.append("ì„±ëª…/ì´ë¦„")

    if missing:
        raise ValueError(
            "[ì˜¤ë¥˜] ì „ì¶œìƒ íŒŒì¼ í—¤ë”ì—ì„œ "
            + ", ".join(missing)
            + " ì—´ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. í—¤ë”ëª…ì„ í™•ì¸í•´ ì£¼ì„¸ìš”."
        )

    # 3) ì˜ˆì‹œ/ë°ì´í„° ì‹œì‘ í–‰ ìë™ ê°ì§€
    if data_start_row is None:
        _, data_start_row = detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=col_name,
        )

    out: List[Dict[str, Any]] = []
    row = data_start_row
    while True:
        grade = ws.cell(row=row, column=col_grade).value
        cls   = ws.cell(row=row, column=col_class).value
        name  = ws.cell(row=row, column=col_name).value

        vals = [grade, cls, name]
        # ì™„ì „ ë¹ˆ ì¤„ì´ë©´ ì¢…ë£Œ
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        # ì¼ë¶€ë§Œ ë¹„ì–´ ìˆìœ¼ë©´ ì˜¤ë¥˜
        if any(v is None or str(v).strip() == "" for v in vals):
            raise ValueError(
                f"[ì˜¤ë¥˜] ì „ì¶œìƒ íŒŒì¼ {row}í–‰ì— í•™ë…„/ë°˜/ì´ë¦„ ì¤‘ ë¹„ì–´ ìˆëŠ” ê°’ì´ ìˆìŠµë‹ˆë‹¤."
            )

        # í•™ë…„ì—ì„œ ìˆ«ìë§Œ ì¶”ì¶œ (1, 2í•™ë…„, "3" ë‹¤ ì»¤ë²„)
        grade_s = str(grade).strip()
        m = re.search(r"\d+", grade_s)
        if not m:
            raise ValueError(
                f"[ì˜¤ë¥˜] ì „ì¶œìƒ íŒŒì¼ {row}í–‰ í•™ë…„ì—ì„œ ìˆ«ìë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤: {grade_s!r}"
            )
        grade_i = int(m.group(0))

        cls_s = normalize_withdraw_class(cls, grade_i)
        if not cls_s:
            raise ValueError(
                f"[ì˜¤ë¥˜] ì „ì¶œìƒ íŒŒì¼ {row}í–‰ ë°˜ ì •ê·œí™” ê²°ê³¼ê°€ ë¹„ì–´ ìˆìŠµë‹ˆë‹¤."
            )

        name_n = normalize_name(name)
        if not name_n:
            raise ValueError(
                f"[ì˜¤ë¥˜] ì „ì¶œìƒ íŒŒì¼ {row}í–‰ ì´ë¦„ ì •ê·œí™” ê²°ê³¼ê°€ ë¹„ì–´ ìˆìŠµë‹ˆë‹¤."
            )

        out.append({"grade": grade_i, "class": cls_s, "name": name_n})
        row += 1

    return out





# ========== L3: roster / transfer / withdraw core logic ==========

def parse_roster_year_from_filename(roster_path: Path) -> Optional[int]:
    stem = roster_path.stem
    s = stem.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()

    m = re.search(r"(\d{4})\s*í•™\s*ë…„ë„", s)
    if m:
        return int(m.group(1))

    m2 = re.search(r"(19\d{2}|20\d{2})", s)
    if m2:
        return int(m2.group(1))

    return None

def load_roster_sheet(dirs: Dict[str, Path], school_name: str):
    """
    í•™ìƒëª…ë¶€(.xlsx, íŒŒì¼ëª…ì— 'í•™ìƒëª…ë¶€' í¬í•¨)ë¥¼ í•™êµ í´ë”ì—ì„œ ì°¾ì•„ì„œ
    - ì²« ë²ˆì§¸ ì‹œíŠ¸ë¥¼ openpyxl ì›Œí¬ì‹œíŠ¸ë¡œ ë°˜í™˜
    - íŒŒì¼ ê²½ë¡œ
    - íŒŒì¼ëª… ê¸°ì¤€ ì¶”ì • í•™ë…„ë„ (ì—†ìœ¼ë©´ None)
    ë¥¼ ëŒë ¤ì¤€ë‹¤.
    """
    root_dir = dirs["SCHOOL_ROOT"]

    kw = (school_name or "").strip()
    if not kw:
        raise ValueError("[ì˜¤ë¥˜] í•™êµëª…ì´ ë¹„ì–´ ìˆì–´ í•™ìƒëª…ë¶€ í´ë”ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")

    # ğŸ”¹ í•™êµ í´ë”ë¥¼ í¬í•¨ ë§¤ì¹­ìœ¼ë¡œ ì°¾ê¸°
    matches = [
        p
        for p in root_dir.iterdir()
        if p.is_dir() and text_contains(p.name, kw)
    ]

    if not matches:
        raise ValueError(
            f"[ì˜¤ë¥˜] í•™ìƒëª…ë¶€ë¥¼ ì°¾ì„ í•™êµ í´ë”ë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. "
            f"(ì‘ì—… í´ë” ë‚´ '{school_name}' í¬í•¨ í´ë” ì—†ìŒ)"
        )

    if len(matches) > 1:
        raise ValueError(
            f"[ì˜¤ë¥˜] í•™ìƒëª…ë¶€ë¥¼ ì°¾ì„ í•™êµ í´ë” í›„ë³´ê°€ ì—¬ëŸ¬ ê°œì…ë‹ˆë‹¤: "
            + ", ".join(p.name for p in matches)
        )

    school_root = matches[0]

    candidates: List[Path] = [
        p
        for p in school_root.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".xlsx"
        and "í•™ìƒëª…ë¶€" in p.stem
        and not p.name.startswith("~$")
    ]
    if not candidates:
        raise ValueError("[ì˜¤ë¥˜] í•™ìƒëª…ë¶€(.xlsx, íŒŒì¼ëª…ì— 'í•™ìƒëª…ë¶€') íŒŒì¼ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.")

    # ê°€ì¥ ìµœê·¼ ìˆ˜ì • íŒŒì¼ ì‚¬ìš©
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    roster_path = candidates[0]

    wb = safe_load_workbook(roster_path, data_only=True)
    ws = wb.worksheets[0]
    roster_year = parse_roster_year_from_filename(roster_path)

    return ws, roster_path, roster_year

def parse_class_str(s: str) -> Optional[Tuple[int, str]]:
    if s is None:
        return None
    m = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", str(s))
    if not m:
        return None
    return int(m.group(1)), m.group(2).strip()

def extract_id_prefix4(uid: str) -> Optional[int]:
    if uid is None:
        return None
    s = str(uid).strip()
    if len(s) >= 4 and s[:4].isdigit():
        return int(s[:4])
    return None

def analyze_roster_once(roster_ws, input_year: int) -> Dict:
    hm = header_map(roster_ws, 1)
    need = ["í˜„ì¬ë°˜", "ì´ì „ë°˜", "í•™ìƒì´ë¦„", "ì•„ì´ë””"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[ì˜¤ë¥˜] í•™ìƒëª…ë¶€ì— '{k}' í—¤ë”ê°€ ì—†ìŠµë‹ˆë‹¤.")

    c_class = hm["í˜„ì¬ë°˜"]
    c_name  = hm["í•™ìƒì´ë¦„"]
    c_id    = hm["ì•„ì´ë””"]

    prefixes_by_grade = defaultdict(list)
    name_counter_by_grade = defaultdict(Counter)
    prefixes_grade1 = []

    for r in range(2, roster_ws.max_row + 1):
        clv = roster_ws.cell(r, c_class).value
        nmv = roster_ws.cell(r, c_name).value
        idv = roster_ws.cell(r, c_id).value
        if clv is None or nmv is None:
            continue

        parsed = parse_class_str(clv)
        if parsed is None:
            continue
        g, _cls = parsed

        nm = normalize_name(nmv)
        if not nm:
            continue
        name_counter_by_grade[g][nm] += 1

        p4 = extract_id_prefix4(idv)
        if p4 is not None:
            prefixes_by_grade[g].append(p4)
            if g == 1:
                prefixes_grade1.append(p4)

    prefix_mode_by_grade = {}
    for g, arr in prefixes_by_grade.items():
        if arr:
            prefix_mode_by_grade[g] = Counter(arr).most_common(1)[0][0]

    roster_time = "unknown"
    ref_shift = 0
    if prefixes_grade1:
        mode1 = Counter(prefixes_grade1).most_common(1)[0][0]
        if mode1 == input_year:
            roster_time = "this_year"
            ref_shift = 0
        elif mode1 == input_year - 1:
            roster_time = "last_year"
            ref_shift = -1
        else:
            roster_time = "unknown"
            ref_shift = 0

    return {
        "roster_time": roster_time,
        "ref_grade_shift": ref_shift,
        "prefix_mode_by_roster_grade": prefix_mode_by_grade,
        "name_count_by_roster_grade": name_counter_by_grade,
    }

def build_transfer_ids(
    transfer_rows: List[Dict],
    roster_info: Dict,
    input_year: int,
) -> Tuple[List[Dict], List[Dict], Dict[int, int]]:
    shift = roster_info["ref_grade_shift"]
    prefix_mode = roster_info["prefix_mode_by_roster_grade"]
    name_counts = roster_info["name_count_by_roster_grade"]

    done: List[Dict] = []
    hold: List[Dict] = []
    final_prefix_by_current_grade: Dict[int, int] = {}
    seen_in_transfer_by_grade = defaultdict(Counter)

    grade1_rows = [tr for tr in transfer_rows if tr["grade"] == 1]
    if grade1_rows:
        g1_names = [tr["name"] for tr in grade1_rows]
        g1_names_sfx = apply_suffix_for_duplicates(g1_names)
        for tr, nm_sfx in zip(grade1_rows, g1_names_sfx):
            uid = f"{input_year}{nm_sfx}"
            done.append({**tr, "id": uid})

    other_rows = [tr for tr in transfer_rows if tr["grade"] != 1]

    for tr in other_rows:
        g_cur = tr["grade"]
        g_roster = g_cur + shift

        pref = prefix_mode.get(g_roster)
        if pref is None:
            hold.append({**tr, "ë³´ë¥˜ì‚¬ìœ ": f"ëª…ë¶€ í•™ë…„({g_roster})ì—ì„œ ID prefix ìµœë¹ˆê°’ ì‚°ì¶œ ë¶ˆê°€"})
            continue

        final_prefix_by_current_grade[g_cur] = pref

        nm = tr["name"]
        base_cnt = name_counts.get(g_roster, Counter()).get(nm, 0)

        seen_in_transfer_by_grade[g_cur][nm] += 1
        add_seq = seen_in_transfer_by_grade[g_cur][nm]

        need_suffix = (base_cnt > 0)
        suffix = dedup_suffix_letters(add_seq) if need_suffix else ""

        uid = f"{pref}{nm}{suffix}"

        is_dup_with_roster = base_cnt > 0  # ğŸ”¸ ëª…ë¶€ ê¸°ì¤€ ë™ëª…ì´ì¸ ì—¬ë¶€

        done.append({**tr, "id": uid})

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, str(x))

    done.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))
    hold.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))

    return done, hold, final_prefix_by_current_grade

def build_withdraw_outputs(
    roster_ws,
    withdraw_rows: List[Dict],
    school_start_date: date,
    work_date: date,
    roster_info: Optional[Dict] = None,
) -> Tuple[List[Dict], List[Dict]]:
    """
    í•™ìƒëª…ë¶€ + ì „ì¶œ ëª…ë‹¨ ê¸°ë°˜ í‡´ì›/ë³´ë¥˜ ë¦¬ìŠ¤íŠ¸ ìƒì„±.
    - í‡´ì›ì¼ì: ì‘ì—…ì¼ < ê°œí•™ì¼ â†’ ê°œí•™ì¼, ê·¸ ì™¸ì—ëŠ” ì‘ì—…ì¼ ê¸°ì¤€
    """
    done: List[Dict] = []
    hold: List[Dict] = []

    # í‡´ì›ì¼ì(íŒŒì¼ ì „ì²´ ê³µí†µ)
    eff = school_start_date if work_date < school_start_date else work_date

    hm = header_map(roster_ws, 1)
    need = ["í˜„ì¬ë°˜", "ì´ì „ë°˜", "í•™ìƒì´ë¦„", "ì•„ì´ë””"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[ì˜¤ë¥˜] í•™ìƒëª…ë¶€ì— '{k}' í—¤ë”ê°€ ì—†ìŠµë‹ˆë‹¤.")

    col_now   = hm["í˜„ì¬ë°˜"]
    col_prev  = hm["ì´ì „ë°˜"]
    col_name  = hm["í•™ìƒì´ë¦„"]
    col_id    = hm["ì•„ì´ë””"]

    # scanì—ì„œ ë„˜ê²¨ì¤€ í•™ë…„ë„ ì •ë³´ëŠ” ë¡œê·¸/ì°¸ê³ ìš©ìœ¼ë¡œë§Œ ì‚¬ìš©
    roster_time = (roster_info or {}).get("roster_time", "this_year")

    # ì¸ë±ìŠ¤ë“¤
    roster_map: Dict[str, List[Dict]] = {}
    roster_by_grade_name: Dict[str, List[Dict]] = {}
    seen_grade_name_ids = set()  # (grade, name_key, id_str)

    def _strip_name_suffix(name_key: str) -> str:
        """
        ì´ë¦„ í‚¤ì—ì„œ ë’¤ìª½ A/B/C ê°™ì€ ì•ŒíŒŒë²³ suffix ì œê±°.
        ì˜ˆ: 'í™ê¸¸ë™A' -> 'í™ê¸¸ë™', 'í™ê¸¸ë™' -> 'í™ê¸¸ë™'
        """
        if not name_key:
            return ""
        return re.sub(r"[A-Za-z]+$", "", name_key)

    def _find_suffix_candidates_for_grade(grade: int, base_name_key: str) -> List[Dict]:
        if grade is None or grade <= 0 or not base_name_key:
            return []

        prefix = f"{grade}|"
        out: List[Dict] = []

        for k, rows in roster_by_grade_name.items():
            if not k.startswith(prefix):
                continue
            _, nk = k.split("|", 1)
            if _strip_name_suffix(nk) == base_name_key:
                out.extend(rows)

        return out


    def _index_class_map(class_val, now_class_val, name_key: str, idv, name_disp: str):
        """ë°˜+ì´ë¦„ ì™„ì „ ë§¤ì¹­ìš© ì¸ë±ìŠ¤ (í˜„ì¬ë°˜/ì´ì „ë°˜ ë‘˜ ë‹¤ ë§¤ì¹­, í˜„ì¬ë°˜ì€ ë”°ë¡œ ë³´ê´€)"""
        if class_val is None:
            return
        c = str(class_val).strip()
        if not c:
            return

        c_now = "" if now_class_val is None else str(now_class_val).strip()

        key1 = f"{c}|{name_key}"
        roster_map.setdefault(key1, []).append(
            {
                "class": c,         # ë§¤ì¹­ì— ì‚¬ìš©ëœ ë°˜(í˜„ì¬/ì´ì „ ë‘˜ ì¤‘ í•˜ë‚˜)
                "now_class": c_now, # í•™ìƒëª…ë¶€ Aì—´ í˜„ì¬ë°˜ (í‡´ì›ë°˜ëª…ìœ¼ë¡œ ì“¸ ê°’)
                "name_key": name_key,
                "name_disp": name_disp,
                "id": "" if idv is None else str(idv).strip(),
            }
        )

    def _index_grade_map(class_val, now_class_val, name_key: str, idv, name_disp: str):
        """í•™ë…„+ì´ë¦„ fallbackìš© ì¸ë±ìŠ¤ (ì „ì¶œìš© í•™ë…„ì€ í˜„ì¬ë°˜ ê¸°ì¤€ìœ¼ë¡œ ì¡ëŠ”ë‹¤)"""
        if class_val is None:
            return
        c = str(class_val).strip()
        if not c:
            return

        parsed = parse_class_str(c)
        if parsed is None:
            return
        g = parsed[0]

        id_str = "" if idv is None else str(idv).strip()
        dedup_key = (g, name_key, id_str)
        if dedup_key in seen_grade_name_ids:
            return
        seen_grade_name_ids.add(dedup_key)

        c_now = "" if now_class_val is None else str(now_class_val).strip()

        key2 = f"{g}|{name_key}"
        roster_by_grade_name.setdefault(key2, []).append(
            {
                "class": c,
                "now_class": c_now,
                "name_key": name_key,
                "name_disp": name_disp,
                "id": id_str,
                "grade": g,
            }
        )

    # í•™ìƒëª…ë¶€ ì¸ë±ìŠ¤ ìƒì„±
    for r in range(2, roster_ws.max_row + 1):
        nmv = roster_ws.cell(r, col_name).value
        if nmv is None:
            continue
        name_disp = normalize_name(nmv)
        name_key  = normalize_name_key(nmv)
        if not name_key:
            continue

        idv  = roster_ws.cell(r, col_id).value
        nowv = roster_ws.cell(r, col_now).value   # Aì—´: í˜„ì¬ë°˜
        prevv = roster_ws.cell(r, col_prev).value # ì´ì „ë°˜

        now_class_val = nowv
        _index_class_map(nowv,  now_class_val, name_key, idv, name_disp)
        _index_class_map(prevv, now_class_val, name_key, idv, name_disp)

        base_class_val = nowv or prevv
        _index_grade_map(base_class_val, now_class_val, name_key, idv, name_disp)

    # ì „ì¶œ í–‰ ì²˜ë¦¬
    for w in withdraw_rows:
        g_cur = w["grade"]
        w_name_disp = w["name"]
        w_name_key  = normalize_name_key(w_name_disp)
        if not w_name_key:
            hold.append(
                {
                    "í•™ë…„": g_cur,
                    "ë°˜": w["class"],
                    "ì„±ëª…": w_name_disp,
                    "ë³´ë¥˜ì‚¬ìœ ": "ì„±ëª… ì •ê·œí™”(í‚¤) ê²°ê³¼ê°€ ë¹„ì–´ ìˆìŒ",
                }
            )
            continue

        # ì „ì¶œ ëª…ë‹¨ì˜ ë°˜(Cì—´)ì€ ì´ë¯¸ normalize_withdraw_classë¡œ í†µì¼ëœ ìƒíƒœë¼ê³  ê°€ì •
        w_class_full = w["class"]
        key = f"{w_class_full}|{w_name_key}"
        matches = roster_map.get(key, [])

        if len(matches) == 0:
            # ê°™ì€ í•™ë…„/ë‹¤ìŒ í•™ë…„ì—ì„œ ì´ë¦„ë§Œ ì¼ì¹˜í•˜ëŠ” í›„ë³´ ì°¾ì•„ë³´ê¸° (exact key ê¸°ì¤€)
            cand0 = roster_by_grade_name.get(f"{g_cur}|{w_name_key}", [])
            cand1 = roster_by_grade_name.get(f"{g_cur+1}|{w_name_key}", [])
            cand = cand0 + cand1
            if len(cand) == 1:
                # ì´ë¦„ í‚¤(w_name_key) ê¸°ì¤€ìœ¼ë¡œ ìœ ì¼í•˜ê²Œ ì°¾ì€ ê²½ìš°
                matches = cand
            else:
                # ğŸ”¹ ë™ëª…ì´ì¸(A/B/C suffix) ì²˜ë¦¬: ì´ë¦„ì—ì„œ suffix ì œê±° í›„ ë‹¤ì‹œ íƒìƒ‰
                base_name_key = _strip_name_suffix(w_name_key)

                # 1) í˜„ì¬ í•™ë…„ ê¸°ì¤€
                suffix_cand0 = _find_suffix_candidates_for_grade(g_cur, base_name_key)
                if len(suffix_cand0) == 1:
                    # í•™ë…„+ì´ë¦„ ê¸°ì¤€ìœ¼ë¡œ ìœ ì¼ â†’ ìë™ ë§¤ì¹­
                    matches = suffix_cand0
                elif len(suffix_cand0) >= 2:
                    # ê°™ì€ í•™ë…„ ì•ˆì—ì„œ ë™ëª…ì´ì¸ ì—¬ëŸ¬ ëª… â†’ ë³´ë¥˜
                    hold.append(
                        {
                            "í•™ë…„": g_cur,
                            "ë°˜": w["class"],
                            "ì„±ëª…": w_name_disp,
                            "ë³´ë¥˜ì‚¬ìœ ": (
                                "ë³´ë¥˜: í•™ìƒëª…ë¶€ì—ì„œ ë™ëª…ì´ì¸(A,B,C ë“±)ìœ¼ë¡œ êµ¬ë¶„ëœ ì´ë¦„ â€“ "
                                "ìë™ ë§¤ì¹­í•˜ì§€ ì•Šê³  ìˆ˜ë™ í™•ì¸ì´ í•„ìš”í•©ë‹ˆë‹¤."
                            ),
                        }
                    )
                    continue
                else:
                    # 2) g+1 í•™ë…„ ê¸°ì¤€ (ëª…ë¶€ ì‹œì ê³¼ ì „ì¶œ ëª…ë‹¨ í•™ë…„ ì°¨ì´ ë³´ì •ìš©)
                    suffix_cand1 = _find_suffix_candidates_for_grade(g_cur + 1, base_name_key)
                    if len(suffix_cand1) == 1:
                        matches = suffix_cand1
                    elif len(suffix_cand1) >= 2:
                        hold.append(
                            {
                                "í•™ë…„": g_cur,
                                "ë°˜": w["class"],
                                "ì„±ëª…": w_name_disp,
                                "ë³´ë¥˜ì‚¬ìœ ": (
                                    "ë³´ë¥˜: í•™ìƒëª…ë¶€ì—ì„œ ë™ëª…ì´ì¸(A,B,C ë“±)ìœ¼ë¡œ êµ¬ë¶„ëœ ì´ë¦„ â€“ "
                                    "ìë™ ë§¤ì¹­í•˜ì§€ ì•Šê³  ìˆ˜ë™ í™•ì¸ì´ í•„ìš”í•©ë‹ˆë‹¤."
                                ),
                            }
                        )
                        continue
                    else:
                        # suffix ê¸°ì¤€ìœ¼ë¡œë„ í›„ë³´ ì—†ìŒ â†’ ê¸°ì¡´ cand ê¸°ë°˜ ì‚¬ìœ  ì‚¬ìš©
                        if len(cand) == 0:
                            reason = (
                                "ìë™ ì œì™¸: í•™ìƒëª…ë¶€ì— ì¡´ì¬í•˜ì§€ ì•ŠëŠ” í•™ìƒ â€“ "
                                "ì„œë²„ ë¯¸ë“±ë¡/í•™ë…„ ë¶ˆì¼ì¹˜ ë“±ìœ¼ë¡œ ì¶”ì •ë˜ë©° í‡´ì› ì²˜ë¦¬ ëŒ€ìƒì—ì„œ ì œì™¸í–ˆìŠµë‹ˆë‹¤. "
                                "(ë°˜ ë§¤ì¹­ ì‹¤íŒ¨, g ë˜ëŠ” g+1 íƒìƒ‰)"
                            )
                        else:
                            reason = (
                                f"ë³´ë¥˜: í•™ë…„+ì´ë¦„ í›„ë³´ê°€ 2ê±´ ì´ìƒ({len(cand)}ê±´) â€“ ìˆ˜ë™ í™•ì¸ í•„ìš”. "
                                "(ë°˜ ë§¤ì¹­ ì‹¤íŒ¨, g ë˜ëŠ” g+1 íƒìƒ‰)"
                            )
                        hold.append(
                            {
                                "í•™ë…„": g_cur,
                                "ë°˜": w["class"],
                                "ì„±ëª…": w_name_disp,
                                "ë³´ë¥˜ì‚¬ìœ ": reason,
                            }
                        )
                        continue

        if len(matches) > 1:
            hold.append(
                {
                    "í•™ë…„": g_cur,
                    "ë°˜": w["class"],
                    "ì„±ëª…": w_name_disp,
                    "ë³´ë¥˜ì‚¬ìœ ": f"ì¤‘ë³µ ë§¤ì¹­({len(matches)}ê±´)",
                }
            )
            continue

        m = matches[0]

        # í•™ìƒëª…ë¶€ Aì—´(í˜„ì¬ë°˜) ê¸°ì¤€ìœ¼ë¡œ í‡´ì›ë°˜ëª… ê²°ì •
        now_class = m.get("now_class") or m.get("class") or w.get("class")
        withdraw_class = now_class

        done.append(
            {
                "í‡´ì›ë°˜ëª…": withdraw_class,
                "í•™ìƒì´ë¦„": w_name_disp,
                "ì•„ì´ë””": m["id"],
                "í‡´ì›ì¼ì": eff,
            }
        )

    return done, hold





# ========== L4: output writers (ë“±ë¡/ì•ˆë‚´/í‡´ì›) ==========

def write_withdraw_to_register(wb, done_rows: List[Dict], hold_rows: List[Dict]):
    # ğŸ”¹ í‡´ì› ì™„ë£Œ ì‹œíŠ¸: í•­ìƒ ì‚¬ìš© (ì—†ìœ¼ë©´ ìƒì„±)
    ws_done = wb["í‡´ì›"] if "í‡´ì›" in wb.sheetnames else wb.create_sheet("í‡´ì›")

    # í‡´ì› ì™„ë£Œ ì •ë ¬ (í‡´ì›ë°˜ëª… â†’ í•™ìƒì´ë¦„ ì˜¤ë¦„ì°¨ìˆœ)
    done_rows = sorted(
        done_rows,
        key=lambda r: (
            str(r.get("í‡´ì›ë°˜ëª…", "")).strip(),
            str(r.get("í•™ìƒì´ë¦„", "")).strip(),
        ),
    )

    clear_sheet_rows(ws_done, 2)

    r = 2
    for row in done_rows:
        write_text_cell(ws_done, r, 1, row["í‡´ì›ë°˜ëª…"])
        write_text_cell(ws_done, r, 2, row["í•™ìƒì´ë¦„"])
        write_text_cell(ws_done, r, 3, row["ì•„ì´ë””"])
        ws_done.cell(r, 4).value = row["í‡´ì›ì¼ì"]       # ë‚ ì§œëŠ” date ê°ì²´ ê·¸ëŒ€ë¡œ
        ws_done.cell(r, 4).number_format = "yyyy-mm-dd"
        r += 1

    # ğŸ”¹ ë³´ë¥˜ 0ëª… ì²˜ë¦¬: ì‹œíŠ¸ ì•„ì˜ˆ ë§Œë“¤ì§€ ì•ŠìŒ (ìˆìœ¼ë©´ ì‚­ì œ)
    ws_hold = None
    if hold_rows:
        # ë³´ë¥˜ ì •ë ¬ (í•™ë…„ â†’ ë°˜ â†’ ì„±ëª…)
        hold_rows = sorted(
            hold_rows,
            key=lambda r: (
                str(r.get("í•™ë…„", "")).strip(),
                str(r.get("ë°˜", "")).strip(),
                str(r.get("ì„±ëª…", "")).strip(),
            ),
        )

        ws_hold = wb["í‡´ì›_ë³´ë¥˜"] if "í‡´ì›_ë³´ë¥˜" in wb.sheetnames else wb.create_sheet("í‡´ì›_ë³´ë¥˜")
        # í—¤ë”ëŠ” í…œí”Œë¦¿ì— ìˆë‹¤ê³  ê°€ì •í•˜ê³ , 2í–‰ë¶€í„°ë§Œ ë¹„ìš°ê¸°
        clear_sheet_rows(ws_hold, 2)

        r = 2
        for row in hold_rows:
            write_text_cell(ws_hold, r, 1, row.get("í•™ë…„", ""))
            write_text_cell(ws_hold, r, 2, row.get("ë°˜", ""))
            write_text_cell(ws_hold, r, 3, row.get("ì„±ëª…", ""))
            write_text_cell(ws_hold, r, 4, row.get("ë³´ë¥˜ì‚¬ìœ ", ""))
            r += 1

        move_sheet_after(wb, "í‡´ì›_ë³´ë¥˜", "í‡´ì›")
    else:
        # í…œí”Œë¦¿ì— ì´ë¯¸ ìˆëŠ” ê²½ìš°ëŠ” ì‚­ì œ
        if "í‡´ì›_ë³´ë¥˜" in wb.sheetnames:
            wb.remove(wb["í‡´ì›_ë³´ë¥˜"])

    from openpyxl.styles import Font, Alignment

    def _format_sheet(ws):
        for rr in range(1, ws.max_row + 1):
            for cc in range(1, ws.max_column + 1):
                cell = ws.cell(rr, cc)
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    _format_sheet(ws_done)
    if ws_hold is not None:
        _format_sheet(ws_hold)

def school_kind_from_name(school_name: str) -> Tuple[str, str]:
    s = (school_name or "").strip()
    if not s:
        return "", ""
    last = s[-1]
    if last == "ì´ˆ":
        return "ì´ˆë“±ë¶€", "ì´ˆ"
    if last == "ì¤‘":
        return "ì¤‘ë“±ë¶€", "ì¤‘"
    if last == "ê³ ":
        return "ê³ ë“±ë¶€", "ê³ "
    return "", ""

def write_transfer_hold_sheet(wb, hold_rows: List[Dict]):
    sheet_name = "ì „ì…ìƒ_ë³´ë¥˜"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    ws.delete_rows(1, ws.max_row)

    write_text_cell(ws, 1, 1, "í•™ë…„")
    write_text_cell(ws, 1, 2, "ë°˜")
    write_text_cell(ws, 1, 3, "ë²ˆí˜¸")
    write_text_cell(ws, 1, 4, "ì„±ëª…")
    write_text_cell(ws, 1, 5, "ë³´ë¥˜ì‚¬ìœ ")

    r = 2
    for row in hold_rows:
        write_text_cell(ws, r, 1, row.get("grade", ""))
        write_text_cell(ws, r, 2, row.get("class", ""))
        write_text_cell(ws, r, 3, row.get("number", ""))
        write_text_cell(ws, r, 4, row.get("name", ""))
        write_text_cell(ws, r, 5, row.get("ë³´ë¥˜ì‚¬ìœ ", ""))
        r += 1

def make_register_class_name(grade_i: int, class_value: Any) -> str:
    """
    ë“±ë¡íŒŒì¼ [í•™ìƒìë£Œ] ìˆ˜ê°•ë°˜ í‘œê¸° í†µì¼:
    - "1-1", "01-01" ê°™ì´ 'í•™ë…„-ë°˜'ì´ ì´ë¯¸ ë“¤ì–´ì˜¨ ê²½ìš° â†’ ë‘˜ ë‹¤ 0 ì œê±° â†’ "1-1"
    - "1", "01" ì²˜ëŸ¼ ìˆ«ìë§Œ ìˆëŠ” ê²½ìš° â†’ í•™ë…„-ë°˜ ì¡°í•© â†’ "1-1"
    - ê·¸ ì™¸ ë¬¸ìì—´ì€ ì¼ë‹¨ gradeì™€ ê·¸ëƒ¥ ë¶™ì„.
    """
    if class_value is None:
        return ""

    s = str(class_value).strip()
    if not s:
        return ""

    # 1) ì´ë¯¸ "í•™ë…„-ë°˜" í˜•íƒœì¸ ê²½ìš° (01-01, 1-01, 01-1 ë“±)
    m = re.match(r"^\s*0*(\d+)\s*-\s*0*(\d+)\s*$", s)
    if m:
        g = int(m.group(1))
        c = int(m.group(2))
        return f"{g}-{c}"  # â†’ 01-01, 1-01, 01-1 ì „ë¶€ "1-1"

    # 2) ìˆ«ìë§Œ ìˆëŠ” ê²½ìš° ("1", "01", "03" ë“±) â†’ ê°™ì€ í•™ë…„ ì•ˆì—ì„œ ë°˜ ë²ˆí˜¸ë¡œ í•´ì„
    m2 = re.match(r"^\s*0*(\d+)\s*$", s)
    if m2:
        c = int(m2.group(1))
        return f"{grade_i}-{c}"

    # 3) ê·¸ ì™¸ ë³µì¡í•œ ë¬¸ìì—´ì´ë©´ ì¼ë‹¨ ìˆëŠ” ê·¸ëŒ€ë¡œ ë¶™ì¸ë‹¤
    return f"{grade_i}-{s}"

def fill_register(
    template_path: Path,
    out_path: Path,
    school_name: str,
    year: str,
    freshmen_rows: List[Dict],
    transfer_done_rows: List[Dict],
    teacher_rows: List[Dict],
    transfer_hold_rows: Optional[List[Dict]] = None,
    withdraw_done_rows: Optional[List[Dict]] = None,
    withdraw_hold_rows: Optional[List[Dict]] = None,
) -> None:
    ensure_xlsx_only(template_path)

    wb = load_workbook(template_path)
    ws_students = wb["í•™ìƒìë£Œ"]
    ws_staff = wb["ì§ì›ì •ë³´"]
    ws_groups = wb["ê·¸ë£¹ë°˜ì •ë³´"]

    # =========================
    # [í•™ìƒìë£Œ] ì»¬ëŸ¼ ë§¤í•‘
    # =========================
    hm = header_map(ws_students, 1)
    need = ["No", "í•™ìƒì´ë¦„", "ID", "í•™êµêµ¬ë¶„", "í•™êµ", "í•™ë…„", "ìˆ˜ê°•ë°˜"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[ì˜¤ë¥˜] í…œí”Œë¦¿ [í•™ìƒìë£Œ]ì— '{k}' í—¤ë”ê°€ ì—†ìŠµë‹ˆë‹¤.")

    col_no = hm["No"]
    col_name = hm["í•™ìƒì´ë¦„"]
    col_id = hm["ID"]
    col_kind = hm["í•™êµêµ¬ë¶„"]
    col_school = hm["í•™êµ"]
    col_grade = hm["í•™ë…„"]
    col_class = hm["ìˆ˜ê°•ë°˜"]

    # ê¸°ì¡´ ë°ì´í„° clear
    for r in range(2, ws_students.max_row + 1):
        for c in [col_no, col_name, col_id, col_kind, col_school, col_grade, col_class]:
            ws_students.cell(row=r, column=c).value = None

    kind_full, kind_prefix = school_kind_from_name(school_name)

    def write_student_row(r: int, no: int, name: str, uid: str, grade_i: int, cls_name: str):
        write_text_cell(ws_students, r, col_no, no)
        write_text_cell(ws_students, r, col_name, name)
        write_text_cell(ws_students, r, col_id, uid)
        write_text_cell(ws_students, r, col_kind, kind_full if kind_full else "")
        write_text_cell(ws_students, r, col_school, school_name)
        write_text_cell(ws_students, r, col_grade, f"{kind_prefix}{grade_i}" if kind_prefix else "")
        write_text_cell(ws_students, r, col_class, cls_name)

    write_row = 2
    running_no = 1

    # ì‹ ì…ìƒ ID: í•™ë…„ë„ + ì´ë¦„(ì¤‘ë³µ suffix í¬í•¨)
    fn_names = [r["name"] for r in freshmen_rows]
    fn_names_sfx = apply_suffix_for_duplicates(fn_names)
    fn_ids = [f"{year}{nm}" for nm in fn_names_sfx]

    for i, fr in enumerate(freshmen_rows):
        r = write_row + i
        write_student_row(
            r=r,
            no=running_no,
            name=fr["name"],
            uid=fn_ids[i],
            grade_i=fr["grade"],
            cls_name=make_register_class_name(fr["grade"], fr["class"]),
            )
        running_no += 1
    write_row += len(freshmen_rows)

    # ì „ì…ìƒ(ì™„ë£Œ)
    for tr in transfer_done_rows:
        r = write_row
        write_student_row(
            r=r,
            no=running_no,
            name=tr["name"],
            uid=tr["id"],
            grade_i=tr["grade"],
            cls_name=make_register_class_name(tr["grade"], tr["class"]),
        )
        running_no += 1
        write_row += 1

    # ì„ ìƒë‹˜(í•™ìŠµìš©) â†’ í•™ìƒìë£Œì— "ì„ ìƒë‹˜ë°˜"
    teachers_learn = [t for t in teacher_rows if t["learn_apply"]]
    t_names = [t["name"] for t in teachers_learn]
    t_names_sfx = apply_suffix_for_duplicates(t_names)
    t_ids = [f"{nm}1" for nm in t_names_sfx]

    for j, t in enumerate(teachers_learn):
        r = write_row + j
        write_student_row(
            r=r,
            no=running_no,
            name=t["name"],
            uid=t_ids[j],
            grade_i=1,
            cls_name="ì„ ìƒë‹˜ë°˜",
        )
        running_no += 1
    write_row += len(teachers_learn)

    # =========================
    # [ì§ì›ì •ë³´]
    # =========================
    hm2 = header_map(ws_staff, 1)
    hm2_lower = {k.lower(): v for k, v in hm2.items()}

    need2 = ["no", "ì´ë¦„", "ì•„ì´ë””", "ê¶Œí•œë¶€ì—¬"]
    for k in need2:
        if k.lower() not in hm2_lower:
            raise ValueError(f"[ì˜¤ë¥˜] í…œí”Œë¦¿ [ì§ì›ì •ë³´]ì— '{k}' í—¤ë”ê°€ ì—†ìŠµë‹ˆë‹¤.")

    col_s_no = hm2_lower["no"]
    col_s_name = hm2_lower["ì´ë¦„"]
    col_s_id = hm2_lower["ì•„ì´ë””"]
    col_s_role = hm2_lower["ê¶Œí•œë¶€ì—¬"]

    for r in range(2, ws_staff.max_row + 1):
        for c in [col_s_no, col_s_name, col_s_id, col_s_role]:
            ws_staff.cell(row=r, column=c).value = None  # í…œí”Œë¦¿ í´ë¦¬ì–´ëŠ” ê·¸ëƒ¥ ë‘¬ë„ ë¨

    teachers_admin = [t for t in teacher_rows if t["admin_apply"]]
    a_names = [t["name"] for t in teachers_admin]
    a_names_sfx = apply_suffix_for_duplicates(a_names)

    staff_write = 2

    for i, t in enumerate(teachers_admin):
        r = staff_write + i
        write_text_cell(ws_staff, r, col_s_no, i + 1)
        write_text_cell(ws_staff, r, col_s_name, t["name"])
        write_text_cell(ws_staff, r, col_s_id, a_names_sfx[i])
        write_text_cell(ws_staff, r, col_s_role, "ì„ ìƒë‹˜")

    # =========================
    # [ê·¸ë£¹ë°˜ì •ë³´]
    # =========================
    hm_g = header_map(ws_groups, 1)
    need_g = ["ê·¸ë£¹ëª…", "ë°˜ëª…", "ìˆ˜ê°•ë£Œ", "ë‹´ì„ëª…", "FullMode"]
    for k in need_g:
        if k not in hm_g:
            raise ValueError(f"[ì˜¤ë¥˜] í…œí”Œë¦¿ [ê·¸ë£¹ë°˜ì •ë³´]ì— '{k}' í—¤ë”ê°€ ì—†ìŠµë‹ˆë‹¤.")

    col_g_group = hm_g["ê·¸ë£¹ëª…"]
    col_g_class = hm_g["ë°˜ëª…"]
    col_g_fee = hm_g["ìˆ˜ê°•ë£Œ"]
    col_g_teacher = hm_g["ë‹´ì„ëª…"]
    col_g_full = hm_g["FullMode"]

    for r in range(2, ws_groups.max_row + 1):
        for c in [col_g_group, col_g_class, col_g_fee, col_g_teacher, col_g_full]:
            ws_groups.cell(row=r, column=c).value = None

    class_set = set()
    last_student_row = find_last_data_row(ws_students, key_col=col_no, start_row=2)
    for r in range(2, last_student_row + 1):
        v = ws_students.cell(row=r, column=col_class).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            class_set.add(s)

    def parse_grade_class(class_name: str):
        """
        '1-1', '3-10' í˜•íƒœì—ì„œ (í•™ë…„, ë°˜)ì„ ì •ìˆ˜ë¡œ ì¶”ì¶œ.
        ë°˜ì´ ìˆ«ìê°€ ì•„ë‹ˆë©´ (í•™ë…„, None).
        """
        s = str(class_name)

        # ìˆ«ì-ìˆ«ì (ì˜ˆ: 1-10)
        m = re.match(r"^\s*(\d+)\s*-\s*(\d+)\s*$", s)
        if m:
            return int(m.group(1)), int(m.group(2))

        # ìˆ«ì-ë¬¸ì (ì˜ˆ: 1-Aë°˜)
        m = re.match(r"^\s*(\d+)\s*-\s*(.+)\s*$", s)
        if m:
            return int(m.group(1)), None

        return None, None


    def group_name_from_class(class_name: str) -> str:
        g, _ = parse_grade_class(class_name)
        if g is None:
            return "ê¸°íƒ€"
        return f"{g}í•™ë…„"


    def class_sort_key(class_name: str):
        if class_name == "ì„ ìƒë‹˜ë°˜":
            return (2, 0, 0, "zzz")

        g, c = parse_grade_class(class_name)

        if g is None:
            return (1, 0, 0, str(class_name))

        # ë°˜ ë²ˆí˜¸ ìˆ«ì ì •ë ¬
        class_order = c if c is not None else 9999

        return (0, g, class_order, str(class_name))

    class_list = sorted(class_set, key=class_sort_key)

    start_r = 2
    r = start_r
    for cls_name in class_list:
        write_text_cell(ws_groups, r, col_g_group, group_name_from_class(cls_name))
        write_text_cell(ws_groups, r, col_g_class, cls_name)
        ws_groups.cell(r, col_g_fee).value = None
        write_text_cell(ws_groups, r, col_g_teacher, "ì„ ìƒë‹˜")
        write_text_cell(ws_groups, r, col_g_full, "Y")
        r += 1

    # ì „ì… ë³´ë¥˜ ì‹œíŠ¸
    if transfer_hold_rows:
        write_transfer_hold_sheet(wb, transfer_hold_rows)

    # ì „ì¶œ ì™„ë£Œ/ë³´ë¥˜ ì‹œíŠ¸
    if (withdraw_done_rows is not None) and (withdraw_hold_rows is not None):
        write_withdraw_to_register(wb, withdraw_done_rows, withdraw_hold_rows)

    # ì›Œí¬ë¶ ì „ì²´: ë¹ˆ í–‰ ì•„ë˜ ì„œì‹ ì œê±° + A1ë¡œ í†µì¼
    clear_format_workbook_from_row(wb, start_row=2)
    reset_view_to_a1(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    backup_if_exists(out_path)
    wb.save(out_path)

def _parse_grade_class_from_register(raw: Any) -> Tuple[Optional[int], str]:
    """
    registerì˜ ìˆ˜ê°•ë°˜ ì»¬ëŸ¼ í•´ì„ìš©.

    - "1-3", "01-05" ê°™ì€ ë¬¸ìì—´ â†’ (1, "3"), (1, "05")
    - ì—‘ì…€ì´ ë‚ ì§œë¡œ ì¸ì‹í•œ ê°’(01-01 â†’ datetime/date)ë„
      ë‹¤ì‹œ "mm-dd" í˜•íƒœë¡œ ë³µì›í•´ì„œ ì²˜ë¦¬.
    - íŒ¨í„´ì´ ì•ˆ ë§ìœ¼ë©´ (None, ì›ë³¸ ë¬¸ìì—´)ì„ ê·¸ëŒ€ë¡œ ëŒë ¤ì¤˜ì„œ
      ì•ˆë‚´ë¬¸ì—ëŠ” ì›ë˜ í…ìŠ¤íŠ¸ê°€ ì°íˆê²Œ í•¨.
    """
    if raw is None:
        return None, ""

    # 1) ì—‘ì…€ì´ 01-01ì„ ë‚ ì§œë¡œ ë°”ê¿”ë²„ë¦° ê²½ìš°: datetime/date â†’ "mm-dd"
    if isinstance(raw, (datetime, date)):
        s = raw.strftime("%m-%d")
    else:
        s = str(raw).strip()

    if not s:
        return None, ""

    # "1-3", "01-05", "01-01" ê°™ì€ íŒ¨í„´
    m = re.match(r"^\s*0*(\d+)\s*-\s*0*(\d+)\s*$", s)
    if not m:
        # íŒ¨í„´ì´ ì•„ì˜ˆ ì•ˆ ë§ìœ¼ë©´ ë¬¸ìì—´ ê·¸ëŒ€ë¡œ ì“°ê²Œ
        return None, s

    grade = int(m.group(1))           # ì•ìë¦¬ 0 ë‚ ë¦¼ â†’ 01 â†’ 1
    cls_str = str(int(m.group(2)))    # ë°˜ë„ 01 â†’ 1

    return grade, cls_str

def build_notice_student_sheet(
    ws_notice,
    register_students_ws,
    transfer_ids: set,
    transfer_dup_ids: set,   # ğŸ”¸ ì¶”ê°€

):
    """
    ì•ˆë‚´íŒŒì¼ - í•™ìƒ ID,PW(í•™ìŠµìš©)
    í—¤ë” 3í–‰: No., í•™ë…„, ë°˜, í•™ìƒì´ë¦„, ID, PW
    ë°ì´í„° 4í–‰ë¶€í„°
    """
    hm_r = header_map(register_students_ws, 1)

    # ğŸ”¹ í•„ìˆ˜ í—¤ë”: ì˜ˆì „ì²˜ëŸ¼ No í¬í•¨
    need_r = ["No", "í•™ìƒì´ë¦„", "ID", "ìˆ˜ê°•ë°˜"]
    for k in need_r:
        if k not in hm_r:
            raise ValueError(f"[ì˜¤ë¥˜] ë“±ë¡ì‘ì—…íŒŒì¼ [í•™ìƒìë£Œ]ì— '{k}' í—¤ë”ê°€ ì—†ìŠµë‹ˆë‹¤.")

    c_r_name = hm_r["í•™ìƒì´ë¦„"]
    c_r_id   = hm_r["ID"]
    c_r_cls  = hm_r["ìˆ˜ê°•ë°˜"]
    c_r_no   = hm_r["No"]

    header_row = 3
    start_row = 4

    # 0) ê¸°ì¡´ ë°ì´í„° 'ê°’'ë§Œ ì§€ìš°ê¸° (ê·¸ëŒ€ë¡œ ìœ ì§€)
    max_row = ws_notice.max_row
    if max_row >= start_row:
        for r in range(start_row, max_row + 1):
            for c in range(1, 7):  # No~PW
                ws_notice.cell(row=r, column=c).value = None

    # 1) 1ì°¨ íŒ¨ìŠ¤: ì•ˆë‚´ì— ë“¤ì–´ê°ˆ í•™ìƒ ëª©ë¡ + (í•™ë…„, ì´ë¦„í‚¤) ì¹´ìš´íŠ¸
    tmp_rows: List[Dict[str, Any]] = []
    name_counter: Counter[tuple] = Counter()

    # ğŸ”µ last_rë¥¼ ë‹¤ì‹œ No ê¸°ì¤€ìœ¼ë¡œ
    last_r = find_last_data_row(register_students_ws, key_col=c_r_no, start_row=2)
    for r in range(2, last_r + 1):
        nm  = register_students_ws.cell(r, c_r_name).value
        uid = register_students_ws.cell(r, c_r_id).value
        cls = register_students_ws.cell(r, c_r_cls).value

        cls_str = "" if cls is None else str(cls).strip()
        if cls_str == "ì„ ìƒë‹˜ë°˜":
            continue

        if (nm is None or str(nm).strip() == "") and (uid is None or str(uid).strip() == ""):
            continue

        nm_s  = "" if nm  is None else str(nm).strip()
        uid_s = "" if uid is None else str(uid).strip()
        if not uid_s:
            # ID ì—†ëŠ” í–‰ì€ ì•ˆë‚´ íŒŒì¼ì—ì„œë„ ì œì™¸
            continue

        grade, cls_only = _parse_grade_class_from_register(cls)
        if grade is None:
            g_disp = ""
            cls_disp = cls_str
        else:
            g_disp = grade
            cls_disp = cls_only

        name_key = notice_name_key(nm_s)
        key = (grade, name_key)

        tmp_rows.append(
            {
                "grade": g_disp,
                "class_disp": cls_disp,
                "name": nm_s,
                "id": uid_s,
                "key": key,
                "is_transfer": uid_s in transfer_ids,
                "is_transfer_dup_with_roster": uid_s in transfer_dup_ids,  # ğŸ”¸ ì¶”ê°€
            }
        )

        if grade is not None and name_key:
            name_counter[key] += 1

    # 2) 2ì°¨ íŒ¨ìŠ¤: ì‹¤ì œ ì•ˆë‚´ ì‹œíŠ¸ì— ì“°ë©´ì„œ ìƒ‰ì¹ 
    cur_row = start_row
    running_no = 1

    for rec in tmp_rows:
        key = rec["key"]
        # 1) ê¸°ë³¸: ë“±ë¡íŒŒì¼ ë‚´ë¶€ (í•™ë…„, ì´ë¦„í‚¤) ì¹´ìš´íŠ¸ë¡œ ë™ëª…ì´ì¸ íŒì •
        dup_flag = name_counter.get(key, 0) >= 2

        # 2) ì „ì…ìƒì´ê³ , ëª…ë¶€ ê¸°ì¤€ ë™ëª…ì´ì¸ìœ¼ë¡œ íŒì •ëœ ê²½ìš° â†’ ë¬´ì¡°ê±´ ë™ëª…ì´ì¸ ì²˜ë¦¬
        if rec["is_transfer"] and rec.get("is_transfer_dup_with_roster"):
            dup_flag = True

        # 3) A~F ì—´ ì „ì²´ë¥¼ í…ìŠ¤íŠ¸ë¡œ ì“°ë©´ì„œ ì…€ ê°ì²´ í™•ë³´
        cell_no    = write_text_cell(ws_notice, cur_row, 1, running_no)        # No.
        cell_grade = write_text_cell(ws_notice, cur_row, 2, rec["grade"])      # í•™ë…„
        cell_class = write_text_cell(ws_notice, cur_row, 3, rec["class_disp"]) # ë°˜
        cell_name  = write_text_cell(ws_notice, cur_row, 4, rec["name"])       # ì´ë¦„
        cell_id    = write_text_cell(ws_notice, cur_row, 5, rec["id"])         # ID

        # PW ì»¬ëŸ¼(6ì—´)ì´ ìˆìœ¼ë©´ í•„ìš” ì‹œ ê°’ ë„£ê¸° (ì—†ìœ¼ë©´ None / ë¹ˆê°’)
        cell_pw    = write_text_cell(ws_notice, cur_row, 6, "1234")

        # ì´ í–‰ì—ì„œ ìƒ‰ì¹ í•  ëŒ€ìƒ ì…€ ì „ì²´ (A~F)
        row_cells = [cell_no, cell_grade, cell_class, cell_name, cell_id, cell_pw]

        # 4) ì „ì…ìƒ: í–‰ ì „ì²´ ì£¼í™©
        if rec["is_transfer"]:
            for cell in row_cells:
                cell.fill = FILL_TRANSFER

        # 5) ë™ëª…ì´ì¸: í–‰ ì „ì²´ ë…¸ë‘ (ì „ì…+ë™ëª…ì´ì¸ì´ë©´ ë…¸ë‘ìœ¼ë¡œ ë®ì–´ì”€)
        if dup_flag:
            for cell in row_cells:
                cell.fill = FILL_DUP

        running_no += 1
        cur_row += 1

def build_notice_teacher_sheet(
    ws_notice,
    teacher_rows: List[Dict],
    learn_ids: Optional[List[str]] = None,
    admin_ids: Optional[List[str]] = None,
):
    """
    ì•ˆë‚´íŒŒì¼ - ì„ ìƒë‹˜ID,PW(ê´€ë¦¬ìš©,í•™ìŠµìš©)
    í—¤ë” 3í–‰, ë°ì´í„° 4í–‰ë¶€í„°.
    - No, ì§ìœ„, ì„ ìƒë‹˜ì´ë¦„: teacher_rowsì˜ position/name ê·¸ëŒ€ë¡œ
    - ê´€ë¦¬ìš©ID: ë“±ë¡íŒŒì¼ [ì§ì›ì •ë³´]ì—ì„œ ê°€ì ¸ì˜¨ ID (fallback: name)
    - í•™ìŠµìš©ID: ë“±ë¡íŒŒì¼ [í•™ìƒìë£Œ] ì„ ìƒë‹˜ë°˜ì—ì„œ ê°€ì ¸ì˜¨ ID (fallback: name+'1')
    - ì‹ ì²­ ì•ˆ í•œ ì¹¸ì€ íšŒìƒ‰ ì²˜ë¦¬
    """
    header_row = 3
    start_row = 4

    # ì§ìœ„(Bì—´) ì»¬ëŸ¼ í­ í™•ì¥ (ê¸´ ì§ìœ„/ë‹´ë‹¹ ëª…ì¹­ ì˜ë¦¬ì§€ ì•Šë„ë¡)
    try:
        ws_notice.column_dimensions["B"].width = 16.6
    except Exception:
        pass

    # ì „ì²´ êµì‚¬ ì¤‘ ì‹ ì²­ì ìˆ˜
    admin_total = sum(1 for t in teacher_rows if t.get("admin_apply"))
    learn_total = sum(1 for t in teacher_rows if t.get("learn_apply"))

    admin_ids_list = admin_ids or []
    learn_ids_list = learn_ids or []

    # ë“±ë¡íŒŒì¼ì—ì„œ ê°€ì ¸ì˜¨ ID ê¸¸ì´ê°€ ì‹ ì²­ì ìˆ˜ì™€ ë§ìœ¼ë©´ ê·¸ëŒ€ë¡œ ì‚¬ìš©
    use_admin_from_reg = admin_total > 0 and len(admin_ids_list) >= admin_total
    use_learn_from_reg = learn_total > 0 and len(learn_ids_list) >= learn_total

    idx_admin = 0
    idx_learn = 0

    r_out = start_row
    no = 1
    for t in teacher_rows:
        pos = "" if t.get("position") is None else str(t.get("position")).strip()
        nm  = "" if t.get("name") is None else str(t.get("name")).strip()
        if not nm and not pos and (not t.get("learn_apply")) and (not t.get("admin_apply")):
            continue

        admin_apply = bool(t.get("admin_apply"))
        learn_apply = bool(t.get("learn_apply"))

        # ----- ê´€ë¦¬ìš© ID: ë“±ë¡íŒŒì¼ ìš°ì„  -----
        admin_id = ""
        if admin_apply:
            if use_admin_from_reg:
                admin_id = admin_ids_list[idx_admin]
            else:
                admin_id = nm  # fallback: ì˜ˆì „ ë°©ì‹
            idx_admin += 1
        admin_pw = "t1234" if admin_id else ""

        # ----- í•™ìŠµìš© ID: ë“±ë¡íŒŒì¼ ìš°ì„  -----
        learn_id = ""
        if learn_apply:
            if use_learn_from_reg:
                learn_id = learn_ids_list[idx_learn]
            else:
                learn_id = f"{nm}1"  # fallback: ì˜ˆì „ ë°©ì‹
            idx_learn += 1
        learn_pw = "1234" if learn_id else ""

        # A: No. / B: ì§ìœ„ / C: ì„ ìƒë‹˜ì´ë¦„ / D: êµ¬ë¶„ìš© ë¹ˆ ì¹¸
        # E: ê´€ë¦¬ìš© ID / F: PW / G: êµ¬ë¶„ìš© ë¹ˆ ì¹¸ / H: í•™ìŠµìš© ID / I: PW
        write_text_cell(ws_notice, r_out, 1, no)
        write_text_cell(ws_notice, r_out, 2, pos)
        write_text_cell(ws_notice, r_out, 3, nm)
        write_text_cell(ws_notice, r_out, 5, admin_id)
        write_text_cell(ws_notice, r_out, 6, admin_pw)
        write_text_cell(ws_notice, r_out, 8, learn_id)
        write_text_cell(ws_notice, r_out, 9, learn_pw)

        # íšŒìƒ‰ ì²˜ë¦¬(ì‹ ì²­ ì•ˆ í•œ ì˜ì—­)
        if not admin_apply:
            for c in [5, 6]:
                ws_notice.cell(r_out, c).fill = FILL_GREY

        if not learn_apply:
            for c in [8, 9]:
                ws_notice.cell(r_out, c).fill = FILL_GREY

        no += 1
        r_out += 1

    delete_rows_below(ws_notice, r_out - 1)

def build_notice_file(
    template_notice_path: Path,
    out_notice_path: Path,
    out_register_path: Path,
    teacher_file_path: Optional[Path],
    transfer_done_rows: List[Dict],
) -> None:
    ensure_xlsx_only(template_notice_path)
    ensure_xlsx_only(out_register_path)

    wb_notice = safe_load_workbook(template_notice_path, data_only=False)
    wb_reg = load_workbook(out_register_path)

    if "í•™ìƒìë£Œ" not in wb_reg.sheetnames:
        raise ValueError("[ì˜¤ë¥˜] ë“±ë¡ì‘ì—…íŒŒì¼ì— 'í•™ìƒìë£Œ' ì‹œíŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤.")

    ws_reg_students = wb_reg["í•™ìƒìë£Œ"]

    def _norm_sheetname(s: str) -> str:
        if s is None:
            return ""
        s = str(s)
        s = s.replace("\u00A0", " ")
        s = re.sub(r"\s+", "", s)
        return s

    def _pick_sheet_by_keywords(wb, keywords: List[str]) -> str:
        keys = [_norm_sheetname(k) for k in keywords]
        for name in wb.sheetnames:
            n = _norm_sheetname(name)
            if all(k in n for k in keys):
                return name
        raise ValueError(
            "[ì˜¤ë¥˜] ì•ˆë‚´ í…œí”Œë¦¿ì—ì„œ í•„ìš”í•œ ì‹œíŠ¸ë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.\n"
            f"- keywords: {keywords}\n"
            f"- sheetnames: {wb.sheetnames}"
        )
    
    sh_student = _pick_sheet_by_keywords(wb_notice, ["í•™ìƒ", "PW", "í•™ìŠµìš©"])
    sh_teacher = _pick_sheet_by_keywords(wb_notice, ["ì„ ìƒë‹˜", "PW"])

    ws_notice_students = wb_notice[sh_student]
    ws_notice_teachers = wb_notice[sh_teacher]

    # 1) ë“±ë¡ì‘ì—…íŒŒì¼ í•™ìƒìë£Œ ì‹œíŠ¸
    ws_reg_students = wb_reg["í•™ìƒìë£Œ"]

    # 1-1) ì „ì… ì™„ë£Œ í•™ìƒ ID set (í•™ìƒ ì•ˆë‚´ ì‹œíŠ¸ ìƒ‰ì¹ ìš©)
    transfer_ids: set[str] = set()
    transfer_dup_ids: set[str] = set()   # ğŸ”¸ ëª…ë¶€ ê¸°ì¤€ ë™ëª…ì´ì¸ ì „ì… IDë§Œ

    for tr in transfer_done_rows:
        uid = tr.get("id")
        if not uid:
            continue
        uid_s = str(uid).strip()
        transfer_ids.add(uid_s)

        if tr.get("dup_with_roster"):    # 1ë‹¨ê³„ì—ì„œ ë¶™ì¸ í”Œë˜ê·¸
            transfer_dup_ids.add(uid_s)

    # 1-2) í•™ìƒ ì•ˆë‚´ ì‹œíŠ¸ ìƒì„±
    build_notice_student_sheet(
        ws_notice=ws_notice_students,
        register_students_ws=ws_reg_students,
        transfer_ids=transfer_ids,
        transfer_dup_ids=transfer_dup_ids,  # ğŸ”¸ ì¶”ê°€

    )

    # --- ë“±ë¡ì‘ì—…íŒŒì¼ì—ì„œ ì‹¤ì œ ID ê°€ì ¸ì˜¤ê¸° (êµì‚¬ ì•ˆë‚´ìš©) ---

    # 2) í•™ìƒìë£Œ ì‹œíŠ¸ì—ì„œ ì„ ìƒë‹˜ë°˜ í•™ìŠµìš© ID
    learn_ids_from_register: Optional[List[str]] = None
    try:
        hm_r = header_map(ws_reg_students, 1)
        col_r_class = hm_r["ìˆ˜ê°•ë°˜"]
        col_r_id    = hm_r["ID"]
        tmp_learn: List[str] = []
        max_row = ws_reg_students.max_row or 1
        for row in range(2, max_row + 1):
            cls_val = ws_reg_students.cell(row=row, column=col_r_class).value
            id_val  = ws_reg_students.cell(row=row, column=col_r_id).value
            if cls_val is None and id_val is None:
                continue
            if str(cls_val).strip() == "ì„ ìƒë‹˜ë°˜" and id_val:
                tmp_learn.append(str(id_val).strip())
        if tmp_learn:
            learn_ids_from_register = tmp_learn
    except Exception:
        learn_ids_from_register = None

    # 3) ì§ì›ì •ë³´ ì‹œíŠ¸ì—ì„œ ê´€ë¦¬ìš© ID
    admin_ids_from_register: Optional[List[str]] = None
    try:
        if "ì§ì›ì •ë³´" in wb_reg.sheetnames:
            ws_reg_staff = wb_reg["ì§ì›ì •ë³´"]
            hm_s = header_map(ws_reg_staff, 1)
            col_s_id = hm_s["ì•„ì´ë””"]
            tmp_admin: List[str] = []
            max_row = ws_reg_staff.max_row or 1
            for row in range(2, max_row + 1):
                id_val = ws_reg_staff.cell(row=row, column=col_s_id).value
                if not id_val:
                    continue
                tmp_admin.append(str(id_val).strip())
            if tmp_admin:
                admin_ids_from_register = tmp_admin
    except Exception:
        admin_ids_from_register = None

    # 4) êµì‚¬ ì•ˆë‚´ ì‹œíŠ¸ ìƒì„±
    teacher_rows = read_teacher_rows(teacher_file_path) if teacher_file_path else []
    build_notice_teacher_sheet(
        ws_notice=ws_notice_teachers,
        teacher_rows=teacher_rows,
        learn_ids=learn_ids_from_register,
        admin_ids=admin_ids_from_register,
)

    out_notice_path.parent.mkdir(parents=True, exist_ok=True)
    backup_if_exists(out_notice_path)

    # ì•ˆë‚´ íŒŒì¼ë„ ì›Œí¬ë¶ ê³µí†µ ê·œì¹™ ì ìš©
    clear_format_workbook_from_row(wb_notice, start_row=4)
    reset_view_to_a1(wb_notice)

    wb_notice.save(out_notice_path)

def render_mail_text(
    mail_template_text: str,
    school_name: str,
    domain: str,
) -> str:
    """
    í…ìŠ¤íŠ¸ íŒŒì¼ ë‚´ë¶€:
    - 'OOì´ˆ'/'OOì¤‘'/'OOê³ ' ê°™ì€ í‘œí˜„ â†’ school_name
    - 'OOOOO.readinggate.com' â†’ domain
    """
    txt = mail_template_text or ""
    if school_name:
        txt = txt.replace("OOì´ˆ", school_name).replace("OOì¤‘", school_name).replace("OOê³ ", school_name)
    if domain:
        txt = re.sub(r"[A-Za-z0-9\-]+\.readinggate\.com", domain, txt)
    return txt






# ========== L5: orchestrator (scan / execute / run) ==========

def get_project_dirs(work_root: Path) -> Dict[str, Path]:
    """
    ì‘ì—… í´ë”(work_root) êµ¬ì¡°:

    work_root/
      â”œâ”€ â—resources/  (ë˜ëŠ” ì´ë¦„ì— 'resources' í¬í•¨ëœ ì•„ë¬´ í´ë” 1ê°œ)
      â”‚    â”œâ”€ DB/
      â”‚    â”œâ”€ templates/
      â”‚    â””â”€ notices/
      â”œâ”€ Aì´ˆë“±í•™êµ/
      â”œâ”€ Bì¤‘í•™êµ/
      â””â”€ ...
    """
    work_root = work_root.resolve()

    # ì´ë¦„ì— 'resources' ê°€ ë“¤ì–´ê°„ í´ë”ë“¤ì„ ëª¨ë‘ ìˆ˜ì§‘
    candidates = [
        p for p in work_root.iterdir()
        if p.is_dir() and "resources" in p.name.lower()
    ]

    if len(candidates) == 0:
        # ì•„ë¬´ê²ƒë„ ì—†ìœ¼ë©´ ê¸°ë³¸ê°’: work_root/resources
        resources_root = work_root / "resources"
    elif len(candidates) == 1:
        resources_root = candidates[0]
    else:
        # ì—¬ëŸ¬ ê°œë©´ ì• ë§¤í•˜ë‹ˆê¹Œ ë°”ë¡œ ì—ëŸ¬
        names = [p.name for p in candidates]
        raise ValueError(
            f"[ì˜¤ë¥˜] ì‘ì—… í´ë” ë‚´ì— 'resources'ë¥¼ í¬í•¨í•œ í´ë”ê°€ ì—¬ëŸ¬ ê°œ ìˆìŠµë‹ˆë‹¤: {names}"
        )

    return {
        "WORK_ROOT": work_root,
        "RESOURCES_ROOT": resources_root,
        "DB": resources_root / "DB",
        "TEMPLATES": resources_root / "templates",
        "NOTICES": resources_root / "notices",
        "SCHOOL_ROOT": work_root,  # í•™êµ í´ë”ëŠ” work_root ë°”ë¡œ ì•„ë˜
    }

def find_templates(format_dir: Path) -> Tuple[Optional[Path], Optional[Path], List[str]]:
    """
    [templates] í´ë” í…œí”Œë¦¿ 2ê°œ ì‹ë³„:
    - ë“±ë¡ í…œí”Œë¦¿: íŒŒì¼ëª…ì— 'ë“±ë¡' í¬í•¨
    - ì•ˆë‚´ í…œí”Œë¦¿: íŒŒì¼ëª…ì— 'ì•ˆë‚´' í¬í•¨
    """
    format_dir = Path(format_dir).resolve()
    if not format_dir.exists():
        return None, None, [f"[ì˜¤ë¥˜] [templates] í´ë”ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {format_dir}"]

    xlsx_files = [
        p for p in format_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    ]
    if not xlsx_files:
        return None, None, [f"[ì˜¤ë¥˜] [templates] í´ë”ì— .xlsx íŒŒì¼ì´ ì—†ìŠµë‹ˆë‹¤: {format_dir}"]

    reg = [p for p in xlsx_files if "ë“±ë¡" in p.stem]
    notice = [p for p in xlsx_files if "ì•ˆë‚´" in p.stem]

    errors: List[str] = []
    if len(reg) == 0:
        errors.append("[ì˜¤ë¥˜] [templates] í´ë”ì—ì„œ 'ë“±ë¡' í…œí”Œë¦¿ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. (íŒŒì¼ëª…ì— 'ë“±ë¡' í¬í•¨)")
    elif len(reg) > 1:
        errors.append("[ì˜¤ë¥˜] [templates] í´ë”ì— 'ë“±ë¡' í…œí”Œë¦¿ì´ ì—¬ëŸ¬ ê°œì…ë‹ˆë‹¤.")

    if len(notice) == 0:
        errors.append("[ì˜¤ë¥˜] [templates] í´ë”ì—ì„œ 'ì•ˆë‚´' í…œí”Œë¦¿ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. (íŒŒì¼ëª…ì— 'ì•ˆë‚´' í¬í•¨)")
    elif len(notice) > 1:
        errors.append("[ì˜¤ë¥˜] [templates] í´ë”ì— 'ì•ˆë‚´' í…œí”Œë¦¿ì´ ì—¬ëŸ¬ ê°œì…ë‹ˆë‹¤.")

    if errors:
        return None, None, errors

    return reg[0], notice[0], []

def scan_work_root(work_root: Path) -> Dict[str, Any]:
    """
    ì‘ì—… ë£¨íŠ¸ì—ì„œ resources/DB, resources/templates, resources/notices, í•™êµ í´ë” ìƒíƒœë¥¼ ì ê²€í•œë‹¤.
    app.pyëŠ” ì—¬ê¸°ì„œ ë‹¤ìŒ í‚¤ë“¤ì„ ê¸°ëŒ€í•˜ê³  ìˆìŒ:

      - ok: bool
      - errors: List[str]
      - message: str
      - school_folders: List[str]
      - notice_titles: List[str]

      - db_ok: bool
      - errors_db: List[str]
      - db_file: Optional[Path]

      - format_ok: bool
      - errors_format: List[str]
      - register_template: Optional[Path]
      - notice_template: Optional[Path]
    """
    work_root = work_root.resolve()
    dirs = get_project_dirs(work_root)

    # ì „ì²´ ì—ëŸ¬
    errors: List[str] = []

    # -------------------------
    # 0. resources ë£¨íŠ¸
    # -------------------------
    res_root = dirs["RESOURCES_ROOT"].resolve()

    # í•™êµ í´ë” ëª©ë¡ (resources í´ë” ì œì™¸)
    school_folders = [
        p.name
        for p in work_root.iterdir()
        if p.is_dir()
        and p.resolve() != res_root
        and not p.name.startswith(".")
    ]
    school_folders.sort()

    # -------------------------
    # 1. DB í´ë” ì ê²€
    # -------------------------
    db_ok = False
    errors_db: List[str] = []
    db_file: Optional[Path] = None

    db_dir = dirs["DB"]
    if not db_dir.exists():
        errors_db.append("[ì˜¤ë¥˜] resources/DB í´ë”ê°€ ì—†ìŠµë‹ˆë‹¤.")
    else:
        db_files = [
            p for p in db_dir.glob("*.xlsb")
            if "í•™êµì „ì²´ëª…ë‹¨" in p.stem and not p.name.startswith("~$")
        ]
        if len(db_files) == 0:
            errors_db.append("[ì˜¤ë¥˜] DB í´ë”ì— 'í•™êµì „ì²´ëª…ë‹¨' xlsb íŒŒì¼ì´ ì—†ìŠµë‹ˆë‹¤.")
        elif len(db_files) > 1:
            errors_db.append("[ì˜¤ë¥˜] DB í´ë”ì— 'í•™êµì „ì²´ëª…ë‹¨' xlsb íŒŒì¼ì´ 2ê°œ ì´ìƒì…ë‹ˆë‹¤.")
        else:
            db_ok = True
            db_file = db_files[0]

    # -------------------------
    # 2. templates(ì–‘ì‹) í´ë” ì ê²€
    # -------------------------
    format_ok = False
    errors_format: List[str] = []
    register_template: Optional[Path] = None
    notice_template: Optional[Path] = None

    tpl_dir = dirs["TEMPLATES"]
    if not tpl_dir.exists():
        errors_format.append("[ì˜¤ë¥˜] resources/templates í´ë”ê°€ ì—†ìŠµë‹ˆë‹¤.")
    else:
        reg_files = [
            p for p in tpl_dir.glob("*.xlsx")
            if "ë“±ë¡" in p.stem and not p.name.startswith("~$")
        ]
        notice_files = [
            p for p in tpl_dir.glob("*.xlsx")
            if "ì•ˆë‚´" in p.stem and not p.name.startswith("~$")
        ]

        if len(reg_files) != 1:
            errors_format.append("templates í´ë”ì— 'ë“±ë¡' í…œí”Œë¦¿ íŒŒì¼ì´ ì •í™•íˆ 1ê°œ ìˆì–´ì•¼ í•©ë‹ˆë‹¤.")
        else:
            register_template = reg_files[0]

        if len(notice_files) != 1:
            errors_format.append("templates í´ë”ì— 'ì•ˆë‚´' í…œí”Œë¦¿ íŒŒì¼ì´ ì •í™•íˆ 1ê°œ ìˆì–´ì•¼ í•©ë‹ˆë‹¤.")
        else:
            notice_template = notice_files[0]

        if not errors_format:
            format_ok = True

    # -------------------------
    # 3. notices í´ë” ì ê²€
    # -------------------------
    notice_dir = dirs["NOTICES"]
    notice_titles: List[str] = []

    if not notice_dir.exists():
        errors.append("[ì˜¤ë¥˜] resources/notices í´ë”ê°€ ì—†ìŠµë‹ˆë‹¤.")
    else:
        txt_files = [p for p in notice_dir.glob("*.txt") if p.is_file()]
        if not txt_files:
            errors.append("[ì˜¤ë¥˜] notices í´ë”ì— .txt íŒŒì¼ì´ ì—†ìŠµë‹ˆë‹¤.")
        else:
            notice_titles = sorted({p.stem.strip() for p in txt_files})

    # -------------------------
    # 4. ì „ì²´ ì—ëŸ¬ í•©ì¹˜ê¸°
    # -------------------------
    errors.extend(errors_db)
    errors.extend(errors_format)

    ok = len(errors) == 0
    message = (
        "[OK] resources(DB/templates/notices)ê°€ ì •ìƒì ìœ¼ë¡œ ì¤€ë¹„ë˜ì—ˆìŠµë‹ˆë‹¤."
        if ok else ""
    )

    return {
        "ok": ok,
        "errors": errors,
        "message": message,
        "school_folders": school_folders,
        "notice_titles": notice_titles,

        # DB ìƒíƒœ (app.pyì—ì„œ ì‚¬ìš©)
        "db_ok": db_ok,
        "errors_db": errors_db,
        "db_file": db_file,

        # ì–‘ì‹ ìƒíƒœ (app.pyì—ì„œ ì‚¬ìš©)
        "format_ok": format_ok,
        "errors_format": errors_format,
        "register_template": register_template,
        "notice_template": notice_template,
    }

def find_single_input_file(input_dir: Path, keywords: Sequence[str]) -> Optional[Path]:
    if not input_dir.exists():
        return None

    kw_list: List[str] = []
    for k in keywords:
        k = "" if k is None else str(k).strip()
        if k:
            kw_list.append(k)

    if not kw_list:
        return None

    candidates: List[Path] = []
    for p in input_dir.iterdir():
        if not (p.is_file() and p.suffix.lower() == ".xlsx"):
            continue
        if p.name.startswith("~$"):
            continue
        if any(text_contains(p.name, kw) for kw in kw_list):
            candidates.append(p)

    if len(candidates) == 0:
        return None
    if len(candidates) > 1:
        raise ValueError(f"[ì˜¤ë¥˜] {kw_list} í¬í•¨ .xlsx íŒŒì¼ì´ 2ê°œ ì´ìƒ: {[c.name for c in candidates]}")
    return candidates[0]

def choose_template_register(format_dir: Path, year_str: str = "") -> Path:
    reg, notice, errors = find_templates(format_dir)
    if errors:
        raise ValueError(errors[0])
    assert reg is not None
    return reg

def choose_template_notice(format_dir: Path, year_str: str = "") -> Path:
    reg, notice, errors = find_templates(format_dir)
    if errors:
        raise ValueError(errors[-1])
    assert notice is not None
    return notice

def choose_db_xlsb(db_dir: Path) -> Path:
    if not db_dir.exists():
        raise ValueError(f"[ì˜¤ë¥˜] DB í´ë”ê°€ ì—†ìŠµë‹ˆë‹¤: {db_dir}")

    xlsb_files = [
        p for p in db_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsb" and not p.name.startswith("~$")
    ]
    if not xlsb_files:
        raise ValueError("[ì˜¤ë¥˜] DB í´ë”ì— .xlsb íŒŒì¼ì´ ì—†ìŠµë‹ˆë‹¤.")
    xlsb_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return xlsb_files[0]

def search_schools_in_db(work_root: Path, keyword: str, limit: int = 30) -> List[str]:
    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)
    db_path = choose_db_xlsb(dirs["DB"])

    kw = (keyword or "").strip()
    if not kw:
        return []

    kw_norm = normalize_text(kw)

    results: List[str] = []
    seen = set()

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            return []
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 4:
                    continue
                v = row[4].v  # Eì—´
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue

                if kw_norm and (kw_norm in normalize_text(s)) and s not in seen:
                    seen.add(s)
                    results.append(s)
                    if len(results) >= limit:
                        break

    return results

def school_exists_in_db(db_dir: Path, school_name: str) -> Path:
    db_path = choose_db_xlsb(db_dir)

    target = (school_name or "").strip()
    if not target:
        raise ValueError("[ì˜¤ë¥˜] í•™êµëª…ì´ ë¹„ì–´ ìˆìŠµë‹ˆë‹¤(DB ê²€ì¦ ë¶ˆê°€).")

    target_norm = normalize_text(target)
    found = False

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            raise ValueError("[ì˜¤ë¥˜] DB xlsbì— ì‹œíŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤.")
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 4:
                    continue
                v = row[4].v  # Eì—´
                if v is None:
                    continue
                cell = str(v).strip()
                if not cell:
                    continue
                cell_norm = normalize_text(cell)
                if target_norm and cell_norm and (target_norm in cell_norm):
                    found = True
                    break

    if not found:
        raise ValueError(f"[ì˜¤ë¥˜] DB(Eì—´ 9í–‰~)ì—ì„œ í•™êµëª… '{target}' í¬í•¨ í•­ëª©ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.")

    return db_path

def _normalize_domain(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    s = re.sub(r"^https?://", "", s, flags=re.I)
    s = s.split("/")[0].strip()
    return s

def get_school_domain_from_db(db_dir: Path, school_name: str) -> Optional[str]:
    """
    DB xlsbì—ì„œ:
    - Eì—´: í•™êµëª… ë§¤ì¹­
    - Fì—´: í™ˆí˜ì´ì§€(ë¦¬ë”©ê²Œì´íŠ¸ ì „ìš© ë„ë©”ì¸) ë°˜í™˜
    ì—†ìœ¼ë©´ None
    """
    db_path = choose_db_xlsb(db_dir)
    target = (school_name or "").strip()
    if not target:
        return None
    target_norm = normalize_text(target)

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            return None
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 5:
                    continue
                ev = row[4].v  # E
                if ev is None:
                    continue
                ecell = str(ev).strip()
                if not ecell:
                    continue
                if target_norm and (target_norm in normalize_text(ecell)):
                    fv = row[5].v  # F
                    dom = _normalize_domain("" if fv is None else str(fv))
                    return dom if dom else None
    return None

def load_notice_templates(work_root: Path) -> dict[str, str]:
    dirs = get_project_dirs(work_root)
    notice_dir = dirs["NOTICES"]

    if not notice_dir.exists():
        return {}

    result = {}

    for p in notice_dir.glob("*.txt"):
        if not p.is_file():
            continue
        try:
            text = p.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = p.read_text(encoding="utf-8-sig")

        result[p.stem.strip()] = text.strip()

    return result

def domain_missing_message(school_name: str) -> str:
    _, kind_prefix = school_kind_from_name(school_name)
    kind_disp = kind_prefix if kind_prefix else "í•™êµ"
    return f"{kind_disp} (ì‚¬ìš©ìê°€ ì‘ì—…ì¤‘ì¸) ì˜ ë„ë©”ì¸ ì£¼ì†Œê°€ ì¡´ì¬í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤. í•™êµ ì „ì²´ ëª…ë‹¨ íŒŒì¼ì„ í™•ì¸í•˜ì„¸ìš”."

def scan_pipeline(
    work_root: Path,
    school_name: str,
    school_start_date: date,
    work_date: date,
    roster_basis_date: Optional[date] = None,
) -> ScanResult:
    logs: List[str] = []

    def log(msg: str):
        logs.append(msg)

    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)

    school_name = (school_name or "").strip()
    year_str = str(school_start_date.year).strip()
    year_int = school_start_date.year

    sr = ScanResult(
        ok=False,
        logs=logs,
        school_name=school_name,
        year_str=year_str,
        year_int=0,
        project_root=work_root,
        input_dir=Path("."),
        output_dir=Path("."),
        template_register=None,
        template_notice=None,
        db_path=None,
        freshmen_file=None,
        teacher_file=None,
        transfer_file=None,
        withdraw_file=None,
        need_roster=False,
        roster_path=None,
        roster_year=None,
        roster_info=None,
        needs_open_date=False,
        missing_fields=[],
        can_execute=False,
        can_execute_after_input=False,
    )

    try:
        if not school_name:
            raise ValueError("[ì˜¤ë¥˜] í•™êµëª…ì´ ë¹„ì–´ ìˆìŠµë‹ˆë‹¤.")
        year_int = int(year_str)
        sr.year_int = year_int

        db_path = school_exists_in_db(dirs["DB"], school_name)
        sr.db_path = db_path
        log(f"[OK] DB ê²€ì¦ í†µê³¼ | ì‚¬ìš© íŒŒì¼: {db_path.name}")

        # ğŸ”¹ í•™êµ í´ë”: ì„ íƒí•œ í•™êµëª…ì´ í¬í•¨ëœ í´ë” ì°¾ê¸° (ì •ê·œí™” í¬í•¨ ë§¤ì¹­)
        root_dir = dirs["SCHOOL_ROOT"]

        kw = (school_name or "").strip()
        if not kw:
            raise ValueError("[ì˜¤ë¥˜] í•™êµëª…ì´ ë¹„ì–´ ìˆì–´ í•™êµ í´ë”ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")

        matches = [
            p
            for p in root_dir.iterdir()
            if p.is_dir() and text_contains(p.name, kw)
        ]

        if not matches:
            raise ValueError(
                f"[ì˜¤ë¥˜] ì‘ì—… í´ë” ì•ˆì—ì„œ '{school_name}' ì´(ê°€) í¬í•¨ëœ í•™êµ í´ë”ë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤."
            )

        if len(matches) > 1:
            raise ValueError(
                f"[ì˜¤ë¥˜] '{school_name}' ì´(ê°€) í¬í•¨ëœ í´ë”ê°€ ì—¬ëŸ¬ ê°œì…ë‹ˆë‹¤: "
                + ", ".join(p.name for p in matches)
            )

        school_dir = matches[0]

        # ğŸ”¹ ë¡œê·¸: ì–´ë–¤ í´ë”ë¡œ ë§¤ì¹­ëëŠ”ì§€ ëª…í™•íˆ ì°ì–´ì¤Œ
        log(f"[OK] í•™êµ í´ë” ë§¤ì¹­: {school_dir.name}")

        input_dir = school_dir
        output_dir = school_dir / "ì‘ì—…"

        sr.input_dir = input_dir
        sr.output_dir = output_dir

        # ğŸ”¹ ë¡œê·¸: í•™êµ í´ë” ì•ˆ íŒŒì¼ ëª©ë¡ ì¶œë ¥ (ì•ˆì „ ì²˜ë¦¬)
        try:
            file_list = [p.name for p in input_dir.iterdir() if p.is_file()]
            log(f"[DEBUG] input files: {file_list}")
        except Exception as e:
            log(f"[WARN] í•™êµ í´ë” íŒŒì¼ ëª©ë¡ ì¡°íšŒ ì¤‘ ì˜¤ë¥˜: {e}")

        # ğŸ‘‰ ì‹ ì…ìƒ íŒŒì¼ì€ ì´ì œ "í•„ìˆ˜" ì•„ë‹˜. None í—ˆìš©.
        freshmen_file = find_single_input_file(input_dir, FRESHMEN_KEYWORDS)
        teacher_file  = find_single_input_file(input_dir, TEACHER_KEYWORDS)
        transfer_file = find_single_input_file(input_dir, TRANSFER_KEYWORDS)
        withdraw_file = find_single_input_file(input_dir, WITHDRAW_KEYWORDS)

        # ğŸ‘‰ ì‹ ì…/ì „ì…/ì „ì¶œ/êµì‚¬ í‚¤ì›Œë“œ ë“¤ì–´ê°„ íŒŒì¼ì´ í•˜ë‚˜ë„ ì—†ìœ¼ë©´ ì—ëŸ¬
        if not any([freshmen_file, teacher_file, transfer_file, withdraw_file]):
            raise ValueError(
                "[ì˜¤ë¥˜] ì‹ ì…ìƒ/ì „ì…ìƒ/ì „ì¶œìƒ/êµì‚¬ í‚¤ì›Œë“œê°€ ë“¤ì–´ê°„ ì…ë ¥ xlsx íŒŒì¼ì„ í•˜ë‚˜ë„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. "
                "í•™ìƒëª…ë¶€ë§Œ ìˆëŠ” ê²½ìš°ì—ëŠ” ì „ì…Â·ì „ì¶œÂ·êµì‚¬ ì¤‘ í•˜ë‚˜ ì´ìƒì˜ íŒŒì¼ì´ í•„ìš”í•©ë‹ˆë‹¤."
            )

        sr.freshmen_file = freshmen_file
        sr.teacher_file = teacher_file
        sr.transfer_file = transfer_file
        sr.withdraw_file = withdraw_file

        log(f"[OK] ì‹ ì…ìƒ: {freshmen_file.name}" if freshmen_file else "[SKIP] ì‹ ì…ìƒ íŒŒì¼ ì—†ìŒ (í‚¤ì›Œë“œ: ì‹ ì…ìƒ/ì‹ ì…)")
        log(f"[OK] êµì‚¬: {teacher_file.name}" if teacher_file else "[SKIP] êµì‚¬ íŒŒì¼ ì—†ìŒ (í‚¤ì›Œë“œ: êµì‚¬/êµì›)")
        log(f"[OK] ì „ì…ìƒ: {transfer_file.name}" if transfer_file else "[SKIP] ì „ì…ìƒ íŒŒì¼ ì—†ìŒ (í‚¤ì›Œë“œ: ì „ì…ìƒ/ì „ì…)")
        log(f"[OK] ì „ì¶œìƒ: {withdraw_file.name}" if withdraw_file else "[SKIP] ì „ì¶œìƒ íŒŒì¼ ì—†ìŒ (í‚¤ì›Œë“œ: ì „ì¶œìƒ/ì „ì¶œ)")

        template_register = choose_template_register(dirs["TEMPLATES"], year_str)
        sr.template_register = template_register
        log(f"[OK] ì–‘ì‹(ë“±ë¡): {template_register.name}")

        template_notice = choose_template_notice(dirs["TEMPLATES"], year_str)
        sr.template_notice = template_notice
        log(f"[OK] ì–‘ì‹(ì•ˆë‚´): {template_notice.name}")

        need_roster = bool(transfer_file) or bool(withdraw_file)
        sr.need_roster = need_roster

        if need_roster:
            roster_ws, roster_path, roster_year = load_roster_sheet(dirs, school_name)
            sr.roster_path = roster_path
            sr.roster_year = roster_year
            log(f"[OK] í•™ìƒëª…ë¶€: {roster_path.name}")

            try:
                modified_date = datetime.fromtimestamp(roster_path.stat().st_mtime).date()

                # 1) ê¸°ë³¸ê°’: íŒŒì¼ ë§ˆì§€ë§‰ ìˆ˜ì •ì¼
                auto_basis = modified_date

                # 2) ì‚¬ìš©ìê°€ UIì—ì„œ ë°”ê¿”ì„œ ë‚´ë ¤ì¤€ ê°’ì´ ìˆìœ¼ë©´ ê·¸ê²ƒì„ ìš°ì„  ì‚¬ìš©
                if roster_basis_date is not None and roster_basis_date != auto_basis:
                    sr.roster_basis_date = roster_basis_date
                    log(
                        f"[INFO] í•™ìƒëª…ë¶€ ë§ˆì§€ë§‰ ìˆ˜ì •ì¼ì€ {auto_basis.isoformat()} ì´ì§€ë§Œ, "
                        f"ì‚¬ìš©ìê°€ ëª…ë¶€ ê¸°ì¤€ì¼ì„ {roster_basis_date.isoformat()} ë¡œ ìˆ˜ì •í–ˆìŠµë‹ˆë‹¤."
                    )
                else:
                    sr.roster_basis_date = auto_basis
                    log(
                        f"[INFO] í•™ìƒëª…ë¶€ ë§ˆì§€ë§‰ ìˆ˜ì •ì¼({auto_basis.isoformat()})ì„ "
                        "ëª…ë¶€ ê¸°ì¤€ì¼ë¡œ ìë™ ê°ì§€í–ˆìŠµë‹ˆë‹¤."
                    )

                # 3) ì‘ì—…ì¼ê³¼ ë‹¤ë¥¸ ê²½ìš° ì°¸ê³ ìš© ì•ˆë‚´
                if sr.roster_basis_date != work_date:
                    log(
                        "[INFO] í˜„ì¬ ì‘ì—…ì¼ê³¼ ëª…ë¶€ ê¸°ì¤€ì¼ì´ ë‹¤ë¦…ë‹ˆë‹¤. "
                        f"(ì‘ì—…ì¼={work_date.isoformat()}, ëª…ë¶€ ê¸°ì¤€ì¼={sr.roster_basis_date.isoformat()})"
                    )

            except Exception as e:
                log(f"[WARN] í•™ìƒëª…ë¶€ íŒŒì¼ ìˆ˜ì •ì¼ ì¡°íšŒ ì¤‘ ì˜¤ë¥˜: {e}")

                            # 4) í•™ìƒëª…ë¶€ ë‚´ìš© ë¶„ì„ (ID prefix ê¸°ë°˜ ì—°ë„/í•™ë…„ ì‹œí”„íŠ¸ íŒì •)
            try:
                roster_info = analyze_roster_once(roster_ws, input_year=year_int)
                sr.roster_info = roster_info

                rt = roster_info.get("roster_time", "unknown")
                shift = roster_info.get("ref_grade_shift", 0)
                log(
                    f"[INFO] í•™ìƒëª…ë¶€ ë¶„ì„ ê²°ê³¼: roster_time={rt}, ref_grade_shift={shift} "
                    "(ID prefix ê¸°ë°˜ íŒì •)"
                )
            except Exception as e:
                log(f"[WARN] í•™ìƒëª…ë¶€ ë¶„ì„(analyze_roster_once) ì¤‘ ì˜¤ë¥˜: {e}")
                sr.roster_info = None

            # 2) ID prefix ê¸°ë°˜ í•™ë…„ë„ ì¶”ì • (ì°¸ê³ ìš© ì•ˆë‚´)
            try:
                expected_year = year_int
                roster_info = analyze_roster_once(roster_ws, input_year=expected_year)
                id_roster_time = roster_info.get("roster_time")  # this_year / last_year / unknown

                if id_roster_time == "this_year":
                    log(f"[INFO] í•™ìƒëª…ë¶€ ID íŒ¨í„´ ê¸°ì¤€ìœ¼ë¡œ {expected_year}í•™ë…„ë„ ëª…ë¶€ë¡œ ì¶”ì •ë©ë‹ˆë‹¤.")
                elif id_roster_time == "last_year":
                    log(f"[INFO] í•™ìƒëª…ë¶€ ID íŒ¨í„´ ê¸°ì¤€ìœ¼ë¡œ {expected_year-1}í•™ë…„ë„ ëª…ë¶€ë¡œ ì¶”ì •ë©ë‹ˆë‹¤.")
                else:
                    log("[INFO] í•™ìƒëª…ë¶€ ID íŒ¨í„´ ê¸°ì¤€ í•™ë…„ë„ ì¶”ì •ì´ ë¶ˆí™•ì‹¤í•©ë‹ˆë‹¤(unknown).")

                # 3) 'ëª…ë¶€ ê¸°ì¤€ì¼' + ê°œí•™ì¼ ê¸°ì¤€ìœ¼ë¡œ ì‹¤ì œ ì‚¬ìš©í•  í•™ë…„ë„ ê²°ì •
                #    - UIì—ì„œ ì‚¬ìš©ìê°€ ì…ë ¥í•œ ê°’(roster_basis_date)ì´ ìˆìœ¼ë©´ ê·¸ê±¸ ìµœìš°ì„ ìœ¼ë¡œ ì‚¬ìš©
                basis_date = roster_basis_date or sr.roster_basis_date or work_date
                sr.roster_basis_date = basis_date  # ìµœì¢… ê¸°ì¤€ì¼ì„ ScanResultì—ë„ ë°˜ì˜

                if basis_date < school_start_date:
                    roster_time = "last_year"
                    ref_shift = -1
                else:
                    roster_time = "this_year"
                    ref_shift = 0

                roster_info["roster_time"] = roster_time          # ìš°ë¦¬ê°€ ì‹¤ì œë¡œ ì“¸ í•™ë…„ë„
                roster_info["ref_grade_shift"] = ref_shift        # g_roster = g_cur + ref_shift
                roster_info["id_roster_time"] = id_roster_time    # ID íŒ¨í„´ ê¸°ì¤€ ê°’ì€ ì°¸ê³ ìš©

                sr.roster_info = roster_info

                log(
                    "[INFO] ëª…ë¶€ ê¸°ì¤€ì¼/ê°œí•™ì¼ ê¸°ì¤€ìœ¼ë¡œ "
                    f"'{ 'ì‘ë…„' if roster_time == 'last_year' else 'ì˜¬í•´' } í•™ë…„ë„ ëª…ë¶€'ë¡œ ê°„ì£¼í•©ë‹ˆë‹¤. "
                    f"(ref_grade_shift={ref_shift})"
                )

                # ID ì¶”ì •ê°’ê³¼ ì‹¤ì œ ì‚¬ìš© í•™ë…„ë„ê°€ ë‹¤ë¥´ë©´ ê²½ê³ ë§Œ
                if id_roster_time in ("this_year", "last_year") and id_roster_time != roster_time:
                    log(
                        "[WARN] í•™ìƒëª…ë¶€ ID íŒ¨í„´ ê¸°ì¤€ í•™ë…„ë„ ì¶”ì •ì´ "
                        "ëª…ë¶€ ê¸°ì¤€ì¼/ê°œí•™ì¼ ê¸°ì¤€ ì˜ˆìƒ í•™ë…„ë„ì™€ ë‹¤ë¥¼ ìˆ˜ ìˆìŠµë‹ˆë‹¤. "
                        "ëª…ë¶€ê°€ ìµœì‹ ì¸ì§€ í•œ ë²ˆ ë” í™•ì¸í•´ ì£¼ì„¸ìš”."
                    )
            except Exception as e:
                log(f"[WARN] í•™ìƒëª…ë¶€ í•™ë…„ë„/ID íŒ¨í„´ ì¶”ì • ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {e}")
        else:
            log("[SKIP] ì „ì…/ì „ì¶œ íŒŒì¼ì´ ì—†ì–´ í•™ìƒëª…ë¶€ ë¡œë“œë¥¼ ìŠ¤í‚µ")

        needs_open_date = bool(withdraw_file)
        sr.needs_open_date = needs_open_date
        if needs_open_date:
            log("[INFO] ì „ì¶œìƒ íŒŒì¼ ê°ì§€ â†’ ê°œí•™ì¼(í‡´ì›ì¼ì ê³„ì‚°ìš©) ì…ë ¥ í•„ìš”")
        else:
            log("[INFO] ê°œí•™ì¼ ì…ë ¥ ë¶ˆí•„ìš”")

        # -------------------------
        # 3-3 ì‹¤í–‰ ê°€ëŠ¥ ì—¬ë¶€ í”Œë˜ê·¸ ì •ë¦¬
        # -------------------------
        missing_fields: List[str] = []

        if sr.db_path is None:
            missing_fields.append("DB íŒŒì¼")
        if sr.template_register is None:
            missing_fields.append("ë“±ë¡ í…œí”Œë¦¿")
        if sr.template_notice is None:
            missing_fields.append("ì•ˆë‚´ í…œí”Œë¦¿")

        # ğŸ”» ì‹ ì…ìƒ íŒŒì¼ì€ ì„ íƒ ì‚¬í•­ìœ¼ë¡œ ë³€ê²½
        #    (ì…ë ¥ xlsxê°€ ì•„ì˜ˆ ì—†ì„ ë•ŒëŠ” ìœ„ì—ì„œ ì´ë¯¸ ValueErrorë¡œ ë§‰ê³  ìˆìŒ)
        # if sr.freshmen_file is None:
        #     missing_fields.append("ì‹ ì…ìƒ ëª…ë‹¨(.xlsx)")

        if sr.need_roster and sr.roster_path is None:
            missing_fields.append("í•™ìƒëª…ë¶€(.xlsx)")

        sr.missing_fields = missing_fields
        sr.needs_open_date = bool(sr.withdraw_file)

        # ìµœì¢… ì‹¤í–‰ ê°€ëŠ¥ ì—¬ë¶€ ì¬ê³„ì‚°
        base_ok = True
        if sr.need_roster and sr.roster_path is None:
            base_ok = False

        sr.can_execute_after_input = base_ok
        sr.can_execute = base_ok and (len(sr.missing_fields) == 0)

        sr.ok = True
        log("[DONE] ìŠ¤ìº” ì™„ë£Œ")
        return sr
    
    except Exception as e:
        log(f"[ERROR] {e}")
        sr.ok = False
        return sr

def _extract_layout(layout_overrides: Dict[str, Any], kind: str, default_header: int):
    """
    layout_overrides[kind]ê°€
      - dict: {"header_row": x, "data_start_row": y, ...}
      - int : y (data_start_rowë§Œ)
      - None: ìë™ ê°ì§€
    ì´ëŸ° ì¼€ì´ìŠ¤ë¥¼ ëª¨ë‘ ì²˜ë¦¬í•´ì„œ (header_row, data_start_row) íŠœí”Œë¡œ ë°˜í™˜.
    """
    info = layout_overrides.get(kind)

    # dict í˜•íƒœ (detect_input_layout ê²°ê³¼ ê·¸ëŒ€ë¡œ ë“¤ì–´ì˜¨ ê²½ìš°)
    if isinstance(info, dict):
        header = info.get("header_row") or default_header
        data_start = info.get("data_start_row")
        return header, data_start

    # ìˆ«ì í•˜ë‚˜ë§Œ ë“¤ì–´ì˜¨ ê²½ìš° â†’ headerëŠ” ê¸°ë³¸ê°’ ìœ ì§€
    if isinstance(info, (int, float)):
        return default_header, int(info)

    # ì•„ë¬´ ê²ƒë„ ì—†ìœ¼ë©´ ìë™ ê°ì§€
    return default_header, None

def execute_pipeline(
    scan: ScanResult,
    work_date: date,
    school_start_date: Optional[date] = None,
    layout_overrides: Optional[Dict[str, int]] = None,
) -> PipelineResult:
    """
    scan ê²°ê³¼ë¥¼ ê¸°ë°˜ìœ¼ë¡œ ë“±ë¡íŒŒì¼ + ì•ˆë‚´íŒŒì¼ì„ í•œ ë²ˆì— ìƒì„±.
    - ì‹ ì…ìƒë§Œ ìˆì–´ë„ ë™ì‘
    - ì „ì…/ì „ì¶œ/êµì‚¬ íŒŒì¼ì´ ì—†ìœ¼ë©´ ê·¸ ë¶€ë¶„ì€ ìë™ìœ¼ë¡œ ìŠ¤í‚µ
    - ì „ì¶œì€ í•™ìƒëª…ë¶€ + ê°œí•™ì¼ì´ ëª¨ë‘ ìˆì–´ì•¼ ì²˜ë¦¬
    """
    # ìŠ¤ìº” ë¡œê·¸ ì´ì–´ì„œ ì‚¬ìš©
    logs: List[str] = list(scan.logs)

    def log(msg: str):
        logs.append(msg)

    layout_overrides = layout_overrides or {}

    try:
        if not scan.ok:
            raise ValueError("[ì˜¤ë¥˜] scan ê²°ê³¼ê°€ ok=False ì…ë‹ˆë‹¤. ìŠ¤ìº” ë‹¨ê³„ ì˜¤ë¥˜ë¥¼ ë¨¼ì € í™•ì¸í•´ ì£¼ì„¸ìš”.")

        school_name = scan.school_name
        year_str = scan.year_str
        year_int = scan.year_int or int(year_str)

        log(f"[INFO] ì‹¤í–‰ ì‹œì‘ | í•™êµ={school_name}, í•™ë…„ë„={year_str}")
        log(f"[INFO] ì‘ì—… í´ë”: {scan.output_dir}")

        # -------------------------------------------------
        # 1) ì¸í’‹ íŒŒì¼ ê²½ë¡œ ì •ë¦¬
        # -------------------------------------------------
        freshmen_path = scan.freshmen_file
        teacher_path = scan.teacher_file
        transfer_path = scan.transfer_file
        withdraw_path = scan.withdraw_file

        # -------------------------------------------------
        # 2) ì¸í’‹ ì½ê¸° (ë ˆì´ì•„ì›ƒ override ë°˜ì˜)
        # -------------------------------------------------
        # ì‹ ì…ìƒ (ìˆì„ ë•Œë§Œ)
        freshmen_rows: List[Dict] = []
        if freshmen_path:
            fr_header, fr_start = _extract_layout(layout_overrides, "freshmen", default_header=2)
            log(
                "[DEBUG] ì‹ ì…ìƒ layout: "
                f"header_row={fr_header}, data_start_row={fr_start if fr_start is not None else 'auto'}"
            )
            freshmen_rows = read_freshmen_rows(
                freshmen_path,
                header_row=fr_header,
                data_start_row=fr_start,
            )
            log(f"[OK] ì‹ ì…ìƒ {len(freshmen_rows)}ëª… ë¡œë“œ")
        else:
            log("[INFO] ì‹ ì…ìƒ íŒŒì¼ ì—†ìŒ â†’ ì‹ ì…ìƒ ë“±ë¡ì€ ìŠ¤í‚µí•©ë‹ˆë‹¤.")

        # êµì‚¬
        if teacher_path:
            t_header, t_start = _extract_layout(layout_overrides, "teacher", default_header=3)
            log(
                "[DEBUG] êµì‚¬ layout: "
                f"header_row={t_header}, data_start_row={t_start if t_start is not None else 'auto'}"
            )
            teacher_rows = read_teacher_rows(
                teacher_path,
                header_row=t_header,
                data_start_row=t_start,
            )
            log(f"[OK] êµì‚¬ ì‹ ì²­ {len(teacher_rows)}ê±´ ë¡œë“œ")
        else:
            teacher_rows = []
            log("[INFO] êµì‚¬ íŒŒì¼ ì—†ìŒ â†’ êµì‚¬ ê´€ë ¨ ì²˜ë¦¬ëŠ” ìŠ¤í‚µ")

        # ì „ì…
        if transfer_path:
            tr_header, tr_start = _extract_layout(layout_overrides, "transfer", default_header=2)
            log(
                "[DEBUG] ì „ì…ìƒ layout: "
                f"header_row={tr_header}, data_start_row={tr_start if tr_start is not None else 'auto'}"
            )
            transfer_rows = read_transfer_rows(
                transfer_path,
                header_row=tr_header,
                data_start_row=tr_start,
            )
            log(f"[OK] ì „ì…ìƒ {len(transfer_rows)}ëª… ë¡œë“œ")
        else:
            transfer_rows = []
            log("[INFO] ì „ì…ìƒ íŒŒì¼ ì—†ìŒ â†’ ì „ì… ì²˜ë¦¬ ìŠ¤í‚µ")

        # ì „ì¶œ
        if withdraw_path:
            wd_header, wd_start = _extract_layout(layout_overrides, "withdraw", default_header=2)
            log(
                "[DEBUG] ì „ì¶œìƒ layout: "
                f"header_row={wd_header}, data_start_row={wd_start if wd_start is not None else 'auto'}"
            )
            withdraw_rows = read_withdraw_rows(
                withdraw_path,
                header_row=wd_header,
                data_start_row=wd_start,
            )
            log(f"[OK] ì „ì¶œìƒ {len(withdraw_rows)}ëª… ë¡œë“œ")
        else:
            withdraw_rows = []
            log("[INFO] ì „ì¶œìƒ íŒŒì¼ ì—†ìŒ â†’ ì „ì¶œ ì²˜ë¦¬ ìŠ¤í‚µ")

        # -------------------------------------------------
        # 3) ì „ì… ID ìƒì„± (í•™ìƒëª…ë¶€ê°€ ìˆëŠ” ê²½ìš°ì—ë§Œ)
        # -------------------------------------------------
        transfer_done_rows: List[Dict] = []
        transfer_hold_rows: List[Dict] = []
        prefix_by_grade: Dict[int, int] = {}

        if transfer_rows:
            if not (scan.roster_path and scan.roster_info):
                raise ValueError("[ì˜¤ë¥˜] ì „ì…ìƒì´ ìˆëŠ”ë° í•™ìƒëª…ë¶€ ì •ë³´ê°€ ì—†ìŠµë‹ˆë‹¤. ìŠ¤ìº” ê²°ê³¼ë¥¼ í™•ì¸í•˜ì„¸ìš”.")

            # (í•„ìš”í•˜ë©´ roster_ws ë„˜ê¸°ë„ë¡ í™•ì¥ ê°€ëŠ¥)
            # roster_wb = safe_load_workbook(scan.roster_path, data_only=True)
            # roster_ws = roster_wb.worksheets[0]

            transfer_done_rows, transfer_hold_rows, prefix_by_grade = build_transfer_ids(
                transfer_rows=transfer_rows,
                roster_info=scan.roster_info,
                input_year=year_int,
            )
            log(f"[OK] ì „ì… ID ë§¤ì¹­ ì™„ë£Œ | ì™„ë£Œ {len(transfer_done_rows)}ëª…, ë³´ë¥˜ {len(transfer_hold_rows)}ëª…")
        else:
            log("[INFO] ì „ì…ìƒ ì—†ìŒ â†’ ì „ì… ID ìƒì„± ìŠ¤í‚µ")

        # -------------------------------------------------
        # 4) ì „ì¶œ í‡´ì› ë¦¬ìŠ¤íŠ¸ ìƒì„± (í•™ìƒëª…ë¶€ + ê°œí•™ì¼ + ì‘ì—…ì¼ í•„ìš”)
        # -------------------------------------------------
        withdraw_done_rows: List[Dict] = []
        withdraw_hold_rows: List[Dict] = []
        transfer_out_auto_skip: int = 0

        if withdraw_rows:
            if not scan.roster_path:
                raise ValueError("[ì˜¤ë¥˜] ì „ì¶œìƒì´ ìˆëŠ”ë° í•™ìƒëª…ë¶€ íŒŒì¼ ê²½ë¡œê°€ ì—†ìŠµë‹ˆë‹¤. ìŠ¤ìº” ê²°ê³¼ë¥¼ í™•ì¸í•˜ì„¸ìš”.")
            if not scan.roster_info:
                raise ValueError("[ì˜¤ë¥˜] ì „ì¶œìƒì´ ìˆëŠ”ë° í•™ìƒëª…ë¶€ ì •ë³´(roster_info)ê°€ ì—†ìŠµë‹ˆë‹¤.")
            if school_start_date is None:
                raise ValueError("[ì˜¤ë¥˜] ì „ì¶œ ì²˜ë¦¬ì— í•„ìš”í•œ ê°œí•™ì¼ì´ ì…ë ¥ë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.")

            roster_wb2 = safe_load_workbook(scan.roster_path, data_only=True)
            sheets2 = roster_wb2.worksheets
            if not sheets2:
                raise ValueError(f"[ì˜¤ë¥˜] í•™ìƒëª…ë¶€ì— ì‹œíŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤: {scan.roster_path.name}")
            roster_ws2 = sheets2[0]

            withdraw_done_rows, withdraw_hold_rows = build_withdraw_outputs(
                roster_ws=roster_ws2,
                withdraw_rows=withdraw_rows,
                school_start_date=school_start_date,
                work_date=work_date,
                roster_info=scan.roster_info,
            )

            transfer_out_auto_skip = sum(
                1 for row in withdraw_hold_rows
                if str(row.get("ë³´ë¥˜ì‚¬ìœ ", "")).startswith("ìë™ ì œì™¸")
            )

            log(
                f"[OK] ì „ì¶œ í‡´ì› ë¦¬ìŠ¤íŠ¸ ìƒì„± | "
                f"í‡´ì› {len(withdraw_done_rows)}ëª…, ë³´ë¥˜ {len(withdraw_hold_rows)}ëª… "
                f"(ìë™ ì œì™¸ {transfer_out_auto_skip}ëª… í¬í•¨)"
            )
        else:
            log("[INFO] ì „ì¶œìƒ ì—†ìŒ â†’ í‡´ì› ì²˜ë¦¬ ìŠ¤í‚µ")

        # -------------------------------------------------
        # 5) ë“±ë¡ì‘ì—…íŒŒì¼ ìƒì„±
        # -------------------------------------------------
        if not scan.template_register:
            raise ValueError("[ì˜¤ë¥˜] ë“±ë¡ í…œí”Œë¦¿ ê²½ë¡œê°€ ì—†ìŠµë‹ˆë‹¤. ìŠ¤ìº” ê²°ê³¼ë¥¼ í™•ì¸í•˜ì„¸ìš”.")

        out_register_path = scan.output_dir / f"â˜…{school_name}_ë“±ë¡ì‘ì—…íŒŒì¼(ì‘ì—…ìš©).xlsx"


        fill_register(
            template_path=scan.template_register,
            out_path=out_register_path,
            school_name=school_name,
            year=year_str,
            freshmen_rows=freshmen_rows,
            transfer_done_rows=transfer_done_rows,
            teacher_rows=teacher_rows,
            # ë¦¬ìŠ¤íŠ¸ ê·¸ëŒ€ë¡œ ì „ë‹¬ (ë¹„ì–´ ìˆì–´ë„ OK)
            transfer_hold_rows=transfer_hold_rows,
            withdraw_done_rows=withdraw_done_rows,
            withdraw_hold_rows=withdraw_hold_rows,
        )
        log(f"[OK] ë“±ë¡ì‘ì—…íŒŒì¼ ìƒì„± ì™„ë£Œ: {out_register_path.name}")
        

        # -------------------------------------------------
        # 6) ì•ˆë‚´íŒŒì¼ ìƒì„± (ID/PW)
        # -------------------------------------------------

        if not scan.template_notice:
            raise ValueError("[ì˜¤ë¥˜] ì•ˆë‚´ í…œí”Œë¦¿ ê²½ë¡œê°€ ì—†ìŠµë‹ˆë‹¤. ìŠ¤ìº” ê²°ê³¼ë¥¼ í™•ì¸í•˜ì„¸ìš”.")

        # ì‹¤ì œë¡œ ì‘ì—…ëœ ëŒ€ìƒë§Œ ì œëª©ì— í¬í•¨
        notice_kinds: List[str] = []

        # ì‹ ì…ìƒ: íŒŒì¼ì´ ìˆê³ , ì½ì–´ì˜¨ í–‰ì´ 1ê°œ ì´ìƒì¼ ë•Œë§Œ í¬í•¨
        if freshmen_path and len(freshmen_rows) > 0:
            notice_kinds.append("ì‹ ì…ìƒ")

        # ì „ì…ìƒ: ì‹¤ì œ IDê°€ ìƒì„±ëœ ì¼€ì´ìŠ¤ê°€ ìˆì„ ë•Œë§Œ í¬í•¨
        if len(transfer_done_rows) > 0:
            notice_kinds.append("ì „ì…ìƒ")

        # êµì§ì›: ì‹ ì²­ ë°ì´í„°ê°€ 1ê°œ ì´ìƒì¼ ë•Œë§Œ í¬í•¨
        if len(teacher_rows) > 0:
            notice_kinds.append("êµì§ì›")

        # ë°©ì–´: ì•„ë¬´ë„ ì—†ì„ ì¼ì€ ê±°ì˜ ì—†ì§€ë§Œ, í˜¹ì‹œ ëª¨ë¥¼ ê²½ìš° ëŒ€ë¹„
        if notice_kinds:
            title_middle = ",".join(notice_kinds) + "_ID,PWì•ˆë‚´"
        else:
            title_middle = "ID,PWì•ˆë‚´"

        out_notice_path = scan.output_dir / f"â˜†{school_name}_{title_middle}.xlsx"

        build_notice_file(
            template_notice_path=scan.template_notice,
            out_notice_path=out_notice_path,
            out_register_path=out_register_path,
            teacher_file_path=teacher_path,
            transfer_done_rows=transfer_done_rows,
        )
        log(f"[OK] ì•ˆë‚´íŒŒì¼ ìƒì„± ì™„ë£Œ: {out_notice_path.name}")

        # -------------------------------------------------
        # 7) ê²°ê³¼ ì •ë¦¬
        # -------------------------------------------------
        pr = PipelineResult(
            ok=True,
            outputs=[out_register_path, out_notice_path],
            logs=logs,
        )
        pr.transfer_in_done = len(transfer_done_rows)
        pr.transfer_in_hold = len(transfer_hold_rows)
        pr.transfer_out_done = len(withdraw_done_rows)
        pr.transfer_out_hold = len(withdraw_hold_rows)
        pr.transfer_out_auto_skip = transfer_out_auto_skip

        log("[DONE] ì‹¤í–‰ ì™„ë£Œ")
        return pr

    except Exception as e:
        log(f"[ERROR] {e}")
        return PipelineResult(
            ok=False,
            outputs=[],
            logs=logs,
        )

def run_pipeline(
    work_root: Path,
    school_name: str,
    school_start_date: date,
    work_date: date,
    layout_overrides: Optional[Dict[str, Dict[str, int]]] = None,
    roster_basis_date: Optional[date] = None,
) -> PipelineResult:
    """
    Streamlitì—ì„œ ë¶€ë¥´ëŠ” ì‹¤ì œ ì‹¤í–‰ í•¨ìˆ˜.

    - 1) scan_pipelineìœ¼ë¡œ ì¸í’‹ ìƒíƒœ ì ê²€
    - 2) ë¬¸ì œ ì—†ìœ¼ë©´ execute_pipelineìœ¼ë¡œ ë“±ë¡/ì•ˆë‚´ ì—‘ì…€ íŒŒì¼ ìƒì„±
    - 3) ì•ˆë‚´ ë©”ì¼/ë¬¸ì í…ìŠ¤íŠ¸ëŠ” generate_notice_mail_textì—ì„œë§Œ ì²˜ë¦¬ (txt íŒŒì¼ ìƒì„± ì—†ìŒ)
    """
    logs: List[str] = []

    def log(msg: str):
        logs.append(msg)

    work_root = Path(work_root).resolve()
    school_name = (school_name or "").strip()

    if not school_name:
        log("[ì˜¤ë¥˜] í•™êµëª…ì´ ë¹„ì–´ ìˆìŠµë‹ˆë‹¤.")
        return PipelineResult(ok=False, outputs=[], logs=logs)

    try:
        # 1) ì‚¬ì „ ì ê²€ (scan)
        scan = scan_pipeline(
            work_root=work_root,
            school_name=school_name,
            school_start_date=school_start_date,
            work_date=work_date,
            roster_basis_date=roster_basis_date,
        )

        logs.extend(scan.logs)

        if not scan.ok:
            log("[ì˜¤ë¥˜] ìŠ¤ìº” ë‹¨ê³„ì—ì„œ ì˜¤ë¥˜ê°€ ë°œìƒí•˜ì—¬ ì‹¤í–‰ì„ ì¤‘ë‹¨í•©ë‹ˆë‹¤.")
            return PipelineResult(ok=False, outputs=[], logs=logs)

        if not scan.can_execute:
            msg = ", ".join(scan.missing_fields) if scan.missing_fields else "í•„ìˆ˜ íŒŒì¼ ëˆ„ë½"
            log(f"[ì˜¤ë¥˜] ì‹¤í–‰ ë¶ˆê°€ ìƒíƒœì…ë‹ˆë‹¤. ({msg})")
            return PipelineResult(ok=False, outputs=[], logs=logs)

        # 2) ì‹¤ì œ ì‹¤í–‰ì€ execute_pipelineì— ìœ„ì„
        result = execute_pipeline(
            scan=scan,
            work_date=work_date,
            school_start_date=school_start_date,
            layout_overrides=layout_overrides,
        )

        # scan ë‹¨ê³„ ë¡œê·¸ + ì‹¤í–‰ ë¡œê·¸ í•©ì¹˜ê¸°
        result.logs = logs + result.logs
        return result

    except Exception as e:
        log(f"[ERROR] ì‹¤í–‰ ì¤‘ ì˜ˆì™¸ ë°œìƒ: {e}")
        return PipelineResult(
            ok=False,
            outputs=[],
            logs=logs,
        )

def run_pipeline_partial(
    work_root: Path,
    school_name: str,
    open_date: date,
    mode: str,
) -> PipelineResult:
    """
    UIì˜ 'ë¶€ë¶„ ì‹¤í–‰' ë²„íŠ¼ìš©.
    í˜„ì¬ëŠ” ì•ˆì •ì„±ì„ ìœ„í•´ ì „ì²´ íŒŒì´í”„ë¼ì¸ì„ ì¬ìƒì„±í•˜ëŠ” ë°©ì‹ìœ¼ë¡œ ë™ì‘.
    mode: 'freshmen'|'teacher'|'transfer'|'withdraw' (ì§€ê¸ˆì€ ì•„ì§ êµ¬ë¶„í•˜ì§€ ì•ŠìŒ)
    """
    # ë¶€ë¶„ ì‹¤í–‰ ë¶„ê¸° ë¡œì§ì€ ë‚˜ì¤‘ì— ì§„ì§œ ë‚˜ëˆ„ê³ ,
    # ì§€ê¸ˆì€ ì „ì²´ run_pipelineì„ ê·¸ëŒ€ë¡œ ëŒë¦°ë‹¤.
    return run_pipeline(
        work_root=work_root,
        school_name=school_name,
        school_start_date=open_date,
        work_date=open_date,
        roster_basis_date=None,
    )

 