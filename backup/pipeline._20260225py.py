# core/pipeline.py
from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Sequence
from collections import Counter, defaultdict

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border
from openpyxl.worksheet.views import Selection
from pyxlsb import open_workbook as open_xlsb_workbook

from core.utils import normalize_text, text_contains, text_eq
from dataclasses import dataclass, field

# =========================
# Result types
# =========================
@dataclass
class PipelineResult:
    ok: bool
    outputs: List[Path]
    logs: List[str]

    transfer_in_done: int = 0
    transfer_in_hold: int = 0
    transfer_out_done: int = 0
    transfer_out_hold: int = 0
    transfer_out_auto_skip: int = 0
    
@dataclass
class ScanResult:
    # 기본 상태
    ok: bool = False
    logs: List[str] = field(default_factory=list)

    # 학교/연도 정보
    school_name: str = ""
    year_str: str = ""
    year_int: int = 0

    # 경로들
    project_root: Path = Path(".")
    input_dir: Path = Path(".")
    output_dir: Path = Path(".")
    template_register: Optional[Path] = None
    template_notice: Optional[Path] = None
    db_path: Optional[Path] = None

    # 인풋 파일
    freshmen_file: Optional[Path] = None
    teacher_file: Optional[Path] = None
    transfer_file: Optional[Path] = None
    withdraw_file: Optional[Path] = None

    # 학생명부 관련
    need_roster: bool = False              # 전입/전출 중 하나라도 있으면 True
    roster_path: Optional[Path] = None
    roster_year: Optional[int] = None
    roster_info: Optional[Dict[str, Any]] = None
    roster_basis_date: Optional[date] = None  # 학생명부 기준일(파일 수정일 or 사용자가 수정한 값)

    # UI 플래그
    needs_open_date: bool = False          # 전출 있으면 True → 개학일 필요
    missing_fields: List[str] = field(default_factory=list)
    can_execute: bool = False
    can_execute_after_input: bool = False


# =========================
# Input keyword sets
# =========================
FRESHMEN_KEYWORDS = ["신입생", "신입"]
TEACHER_KEYWORDS  = ["교사", "교원"]
TRANSFER_KEYWORDS = ["전입생", "전입"]
WITHDRAW_KEYWORDS = ["전출생", "전출"]


# =========================
# Paths
# =========================
def get_project_dirs(work_root: Path) -> Dict[str, Path]:
    """
    작업 폴더(work_root) 구조:

    work_root/
      ├─ ●resources/  (또는 이름에 'resources' 포함된 아무 폴더 1개)
      │    ├─ DB/
      │    ├─ templates/
      │    └─ notices/
      ├─ A초등학교/
      ├─ B중학교/
      └─ ...
    """
    work_root = work_root.resolve()

    # 이름에 'resources' 가 들어간 폴더들을 모두 수집
    candidates = [
        p for p in work_root.iterdir()
        if p.is_dir() and "resources" in p.name.lower()
    ]

    if len(candidates) == 0:
        # 아무것도 없으면 기본값: work_root/resources
        resources_root = work_root / "resources"
    elif len(candidates) == 1:
        resources_root = candidates[0]
    else:
        # 여러 개면 애매하니까 바로 에러
        names = [p.name for p in candidates]
        raise ValueError(
            f"[오류] 작업 폴더 내에 'resources'를 포함한 폴더가 여러 개 있습니다: {names}"
        )

    return {
        "WORK_ROOT": work_root,
        "RESOURCES_ROOT": resources_root,
        "DB": resources_root / "DB",
        "TEMPLATES": resources_root / "templates",
        "NOTICES": resources_root / "notices",
        "SCHOOL_ROOT": work_root,  # 학교 폴더는 work_root 바로 아래
    }

# =========================
# File helpers
# =========================

def find_templates(format_dir: Path) -> Tuple[Optional[Path], Optional[Path], List[str]]:
    """
    [양식] 폴더 템플릿 2개 식별:
    - 등록 템플릿: 파일명에 '등록' 포함
    - 안내 템플릿: 파일명에 '안내' 포함
    """
    format_dir = Path(format_dir).resolve()
    if not format_dir.exists():
        return None, None, [f"[오류] [templates] 폴더를 찾을 수 없습니다: {format_dir}"]

    xlsx_files = [
        p for p in format_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    ]
    if not xlsx_files:
        return None, None, [f"[오류] [templates] 폴더에 .xlsx 파일이 없습니다: {format_dir}"]

    reg = [p for p in xlsx_files if "등록" in p.stem]
    notice = [p for p in xlsx_files if "안내" in p.stem]

    errors: List[str] = []
    if len(reg) == 0:
        errors.append("[오류] [templates] 폴더에서 '등록' 템플릿을 찾지 못했습니다. (파일명에 '등록' 포함)")
    elif len(reg) > 1:
        errors.append("[오류] [templates] 폴더에 '등록' 템플릿이 여러 개입니다.")

    if len(notice) == 0:
        errors.append("[오류] [templates] 폴더에서 '안내' 템플릿을 찾지 못했습니다. (파일명에 '안내' 포함)")
    elif len(notice) > 1:
        errors.append("[오류] [templates] 폴더에 '안내' 템플릿이 여러 개입니다.")

    if errors:
        return None, None, errors

    return reg[0], notice[0], []


NOTICE_ORDER = [
    "신규등록 - 메일",
    "신규등록 - 문자",
    "교직원 등록 - 메일",
    "반이동 - 메일",
    "반이동 - 메일 (신입생, 교직원 등록 & 반이동)",
    "반이동 - 문자",
    "2-6학년 명단 보내 온 경우 - 메일",
    "2-6학년 반편성 자료 재요청 - 문자",
]


def scan_work_root(work_root: Path) -> Dict[str, Any]:
    """
    작업 루트에서 resources/DB, resources/templates, resources/notices, 학교 폴더 상태를 점검한다.
    app.py는 여기서 다음 키들을 기대하고 있음:

      - ok: bool
      - errors: List[str]
      - message: str
      - school_folders: List[str]
      - notice_titles: List[str]

      - db_ok: bool
      - errors_db: List[str]
      - db_file: Optional[Path]

      - format_ok: bool
      - errors_format: List[str]
      - register_template: Optional[Path]
      - notice_template: Optional[Path]
    """
    work_root = work_root.resolve()
    dirs = get_project_dirs(work_root)

    # 전체 에러
    errors: List[str] = []

    # -------------------------
    # 0. resources 루트
    # -------------------------
    res_root = dirs["RESOURCES_ROOT"].resolve()

    # 학교 폴더 목록 (resources 폴더 제외)
    school_folders = [
        p.name
        for p in work_root.iterdir()
        if p.is_dir()
        and p.resolve() != res_root
        and not p.name.startswith(".")
    ]
    school_folders.sort()

    # -------------------------
    # 1. DB 폴더 점검
    # -------------------------
    db_ok = False
    errors_db: List[str] = []
    db_file: Optional[Path] = None

    db_dir = dirs["DB"]
    if not db_dir.exists():
        errors_db.append("[오류] resources/DB 폴더가 없습니다.")
    else:
        db_files = [
            p for p in db_dir.glob("*.xlsb")
            if "학교전체명단" in p.stem and not p.name.startswith("~$")
        ]
        if len(db_files) == 0:
            errors_db.append("[오류] DB 폴더에 '학교전체명단' xlsb 파일이 없습니다.")
        elif len(db_files) > 1:
            errors_db.append("[오류] DB 폴더에 '학교전체명단' xlsb 파일이 2개 이상입니다.")
        else:
            db_ok = True
            db_file = db_files[0]

    # -------------------------
    # 2. templates(양식) 폴더 점검
    # -------------------------
    format_ok = False
    errors_format: List[str] = []
    register_template: Optional[Path] = None
    notice_template: Optional[Path] = None

    tpl_dir = dirs["TEMPLATES"]
    if not tpl_dir.exists():
        errors_format.append("[오류] resources/templates 폴더가 없습니다.")
    else:
        reg_files = [
            p for p in tpl_dir.glob("*.xlsx")
            if "등록" in p.stem and not p.name.startswith("~$")
        ]
        notice_files = [
            p for p in tpl_dir.glob("*.xlsx")
            if "안내" in p.stem and not p.name.startswith("~$")
        ]

        if len(reg_files) != 1:
            errors_format.append("templates 폴더에 '등록' 템플릿 파일이 정확히 1개 있어야 합니다.")
        else:
            register_template = reg_files[0]

        if len(notice_files) != 1:
            errors_format.append("templates 폴더에 '안내' 템플릿 파일이 정확히 1개 있어야 합니다.")
        else:
            notice_template = notice_files[0]

        if not errors_format:
            format_ok = True

    # -------------------------
    # 3. notices 폴더 점검
    # -------------------------
    notice_dir = dirs["NOTICES"]
    notice_titles: List[str] = []

    if not notice_dir.exists():
        errors.append("[오류] resources/notices 폴더가 없습니다.")
    else:
        txt_files = [p for p in notice_dir.glob("*.txt") if p.is_file()]
        if not txt_files:
            errors.append("[오류] notices 폴더에 .txt 파일이 없습니다.")
        else:
            notice_titles = sorted({p.stem.strip() for p in txt_files})

    # -------------------------
    # 4. 전체 에러 합치기
    # -------------------------
    errors.extend(errors_db)
    errors.extend(errors_format)

    ok = len(errors) == 0
    message = (
        "[OK] resources(DB/templates/notices)가 정상적으로 준비되었습니다."
        if ok else ""
    )

    return {
        "ok": ok,
        "errors": errors,
        "message": message,
        "school_folders": school_folders,
        "notice_titles": notice_titles,

        # DB 상태 (app.py에서 사용)
        "db_ok": db_ok,
        "errors_db": errors_db,
        "db_file": db_file,

        # 양식 상태 (app.py에서 사용)
        "format_ok": format_ok,
        "errors_format": errors_format,
        "register_template": register_template,
        "notice_template": notice_template,
    }

def ensure_xlsx_only(p: Path) -> None:
    if p.suffix.lower() != ".xlsx":
        raise ValueError(f"[오류] 파일 형식이 .xlsx가 아닙니다: {p.name}")


def backup_if_exists(out_path: Path) -> Optional[Path]:
    """기존 파일이 있으면 작업/_backup으로 이동."""
    out_path = Path(out_path)
    if not out_path.exists():
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = out_path.parent / "_backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    dest = backup_dir / f"{out_path.stem}_{ts}{out_path.suffix}"
    out_path.replace(dest)
    return dest


def find_single_input_file(input_dir: Path, keywords: Sequence[str]) -> Optional[Path]:
    if not input_dir.exists():
        return None

    kw_list: List[str] = []
    for k in keywords:
        k = "" if k is None else str(k).strip()
        if k:
            kw_list.append(k)

    if not kw_list:
        return None

    candidates: List[Path] = []
    for p in input_dir.iterdir():
        if not (p.is_file() and p.suffix.lower() == ".xlsx"):
            continue
        if p.name.startswith("~$"):
            continue
        if any(text_contains(p.name, kw) for kw in kw_list):
            candidates.append(p)

    if len(candidates) == 0:
        return None
    if len(candidates) > 1:
        raise ValueError(f"[오류] {kw_list} 포함 .xlsx 파일이 2개 이상: {[c.name for c in candidates]}")
    return candidates[0]


def choose_template_register(format_dir: Path, year_str: str = "") -> Path:
    reg, notice, errors = find_templates(format_dir)
    if errors:
        raise ValueError(errors[0])
    assert reg is not None
    return reg


def choose_template_notice(format_dir: Path, year_str: str = "") -> Path:
    reg, notice, errors = find_templates(format_dir)
    if errors:
        raise ValueError(errors[-1])
    assert notice is not None
    return notice


def choose_db_xlsb(db_dir: Path) -> Path:
    if not db_dir.exists():
        raise ValueError(f"[오류] DB 폴더가 없습니다: {db_dir}")

    xlsb_files = [
        p for p in db_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsb" and not p.name.startswith("~$")
    ]
    if not xlsb_files:
        raise ValueError("[오류] DB 폴더에 .xlsb 파일이 없습니다.")
    xlsb_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return xlsb_files[0]


def search_schools_in_db(work_root: Path, keyword: str, limit: int = 30) -> List[str]:
    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)
    db_path = choose_db_xlsb(dirs["DB"])

    kw = (keyword or "").strip()
    if not kw:
        return []

    kw_norm = normalize_text(kw)

    results: List[str] = []
    seen = set()

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            return []
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 4:
                    continue
                v = row[4].v  # E열
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue

                if kw_norm and (kw_norm in normalize_text(s)) and s not in seen:
                    seen.add(s)
                    results.append(s)
                    if len(results) >= limit:
                        break

    return results


# =========================
# DB validate (xlsb)
# =========================
def school_exists_in_db(db_dir: Path, school_name: str) -> Path:
    db_path = choose_db_xlsb(db_dir)

    target = (school_name or "").strip()
    if not target:
        raise ValueError("[오류] 학교명이 비어 있습니다(DB 검증 불가).")

    target_norm = normalize_text(target)
    found = False

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            raise ValueError("[오류] DB xlsb에 시트가 없습니다.")
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 4:
                    continue
                v = row[4].v  # E열
                if v is None:
                    continue
                cell = str(v).strip()
                if not cell:
                    continue
                cell_norm = normalize_text(cell)
                if target_norm and cell_norm and (target_norm in cell_norm):
                    found = True
                    break

    if not found:
        raise ValueError(f"[오류] DB(E열 9행~)에서 학교명 '{target}' 포함 항목을 찾지 못했습니다.")

    return db_path


def _normalize_domain(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    s = re.sub(r"^https?://", "", s, flags=re.I)
    s = s.split("/")[0].strip()
    return s


def get_school_domain_from_db(db_dir: Path, school_name: str) -> Optional[str]:
    """
    DB xlsb에서:
    - E열: 학교명 매칭
    - F열: 홈페이지(리딩게이트 전용 도메인) 반환
    없으면 None
    """
    db_path = choose_db_xlsb(db_dir)
    target = (school_name or "").strip()
    if not target:
        return None
    target_norm = normalize_text(target)

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            return None
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 5:
                    continue
                ev = row[4].v  # E
                if ev is None:
                    continue
                ecell = str(ev).strip()
                if not ecell:
                    continue
                if target_norm and (target_norm in normalize_text(ecell)):
                    fv = row[5].v  # F
                    dom = _normalize_domain("" if fv is None else str(fv))
                    return dom if dom else None
    return None


# =========================
# openpyxl custom prop guard
# =========================
def safe_load_workbook(xlsx_path: Path, data_only: bool = True):
    try:
        return load_workbook(xlsx_path, data_only=data_only)
    except TypeError as e:
        msg = str(e)
        if "openpyxl.packaging.custom" not in msg or "NoneType" not in msg:
            raise

        buffer = BytesIO()
        with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(
            buffer, "w", compression=zipfile.ZIP_DEFLATED
        ) as zout:
            for item in zin.infolist():
                if item.filename == "docProps/custom.xml":
                    root = ET.fromstring(zin.read(item.filename))
                    ns = "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
                    tag = f"{{{ns}}}property"
                    for prop in list(root.findall(tag)):
                        name = prop.get("name")
                        if name is None or str(name).strip() == "":
                            root.remove(prop)
                    new_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    zout.writestr(item, new_xml)
                else:
                    zout.writestr(item, zin.read(item.filename))

        buffer.seek(0)
        return load_workbook(buffer, data_only=data_only)

    except IndexError as e:
        # 스타일 인덱스 꼬여서 나는 openpyxl 버그 회피용
        # 템플릿 저장에 쓰일 일 있는 케이스(data_only=False)는 그대로 올려보내고,
        # 인풋 읽기용(data_only=True)일 때만 read_only 모드로 다시 시도
        if not data_only:
            raise
        return load_workbook(xlsx_path, data_only=data_only, read_only=True)
    
# =========================
# name normalize + suffix
# =========================
HANGUL_RE = re.compile(r"[가-힣]")
EN_RE = re.compile(r"[A-Za-z]")


def normalize_name(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-z가-힣\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""

    has_ko = bool(HANGUL_RE.search(s))
    has_en = bool(EN_RE.search(s))

    if has_ko and not has_en:
        return s.replace(" ", "")

    if has_en and not has_ko:
        parts = [p for p in s.split(" ") if p]
        parts = [p.lower().capitalize() for p in parts]
        return "".join(parts)

    if has_ko and has_en:
        def _fix_en(m: re.Match) -> str:
            tok = m.group(0).lower()
            return tok[0].upper() + tok[1:] if tok else tok
        s2 = re.sub(r"[A-Za-z]+", _fix_en, s)
        return s2.replace(" ", "")

    return ""


def normalize_name_key(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-z가-힣\s]", "", s)
    s = re.sub(r"\s+", "", s)
    return s.casefold()


def english_casefold_key(name: str) -> str:
    if name is None:
        return ""
    return str(name).strip().casefold()


def dedup_suffix_letters(n: int) -> str:
    if n <= 0:
        return ""
    out = ""
    while n > 0:
        n -= 1
        out = chr(ord("A") + (n % 26)) + out
        n //= 26
    return out


def apply_suffix_for_duplicates(names: List[str]) -> List[str]:
    total = {}
    for nm in names:
        key = english_casefold_key(nm)
        total[key] = total.get(key, 0) + 1

    seen = {}
    out = []
    for nm in names:
        key = english_casefold_key(nm)
        if total.get(key, 0) <= 1:
            out.append(nm)
            continue
        seen[key] = seen.get(key, 0) + 1
        out.append(nm + dedup_suffix_letters(seen[key]))
    return out


# =========================
# example row detection (예시 + 데이터 시작 행)
# =========================
EXAMPLE_NAMES_RAW = ["홍길동", "이순신", "유관순", "임꺽정"]
EXAMPLE_NAMES_NORM = {normalize_text(n) for n in EXAMPLE_NAMES_RAW}
EXAMPLE_KEYWORDS = ["예시"]  # 행 안 어느 셀이라도 '예시' 포함되면 예시로 처리


def _row_is_empty(ws, row: int, max_col: Optional[int] = None) -> bool:
    if max_col is None:
        max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and str(v).strip() != "":
            return False
    return True


def _row_has_example_keyword(ws, row: int, max_col: Optional[int] = None) -> bool:
    if max_col is None:
        max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        s = normalize_text(str(v))
        if not s:
            continue
        for kw in EXAMPLE_KEYWORDS:
            if kw in s:
                return True
    return False


def _cell_is_example_name(value: Any) -> bool:
    if value is None:
        return False
    s = normalize_text(str(value))
    return bool(s) and s in EXAMPLE_NAMES_NORM


def detect_example_and_data_start(
    ws,
    header_row: int,
    name_col: int,
    max_search_row: Optional[int] = None,
    max_col: Optional[int] = None,
) -> Tuple[List[int], int]:
    """
    헤더 아래에서 예시 행(0개 이상)과 실제 데이터 시작 행을 자동 감지한다.

    - header_row 바로 아래 행부터 스캔
    - 완전 빈 행은 건너뜀
    - '예시' 키워드가 있거나 이름 칸이 예시 이름이면 → 예시 행
    - 그 외 첫 번째 비-예시 행 → 실제 데이터 시작 행
    """
    if max_search_row is None:
        max_search_row = ws.max_row

    example_rows: List[int] = []
    r = header_row + 1

    while r <= max_search_row:
        # 1) 완전 빈 행은 스킵
        if _row_is_empty(ws, r, max_col=max_col):
            r += 1
            continue

        # 2) 행 안에 '예시' 키워드 있으면 예시
        if _row_has_example_keyword(ws, r, max_col=max_col):
            example_rows.append(r)
            r += 1
            continue

        # 3) 이름 칸이 예시 이름이면 예시
        v_name = ws.cell(row=r, column=name_col).value
        if _cell_is_example_name(v_name):
            example_rows.append(r)
            r += 1
            continue

        # 4) 여기까지 안 걸리면 → 실제 데이터 시작
        return example_rows, r

    raise ValueError(
        f"[오류] 데이터 시작 행을 찾지 못했습니다. 헤더({header_row}행) 아래에 예시나 실제 데이터로 보이는 행이 없습니다."
    )

def detect_input_layout(xlsx_path: Path, kind: str) -> Dict[str, Any]:
    """
    UI에서 인풋 파일 구조를 미리 보여줄 때 사용.
    kind: 'freshmen' | 'transfer' | 'withdraw' | 'teacher'
    반환:
      {
        "header_row": int,
        "example_rows": [int, ...],
        "data_start_row": int,
      }
    """
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    kind_norm = (kind or "").strip().lower()
    if kind_norm == "freshmen":
        header_row = 2
        name_col = 5  # E열: 성명
    elif kind_norm == "transfer":
        header_row = 2
        name_col = 5  # E열: 성명
    elif kind_norm == "withdraw":
        header_row = 2
        name_col = 4  # D열: 성명
    elif kind_norm == "teacher":
        header_row = 3
        name_col = 3  # C열: 선생님 이름
    else:
        raise ValueError(f"[오류] 지원하지 않는 kind 값입니다: {kind}")

    example_rows, data_start_row = detect_example_and_data_start(
        ws,
        header_row=header_row,
        name_col=name_col,
    )

    return {
        "header_row": header_row,
        "example_rows": example_rows,
        "data_start_row": data_start_row,
    }


# =========================
# input readers
# =========================

# 신입생 파일
def read_freshmen_rows(
    xlsx_path: Path,
    header_row: int = 2,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # 자동 감지 (사용자가 직접 data_start_row를 넘기면 그 값을 우선 사용)
    if data_start_row is None:
        _, data_start_row = detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=5,  # E열: 성명
        )

    out = []
    row = data_start_row
    while True:
        grade = ws[f"B{row}"].value
        cls   = ws[f"C{row}"].value
        num   = ws[f"D{row}"].value
        name  = ws[f"E{row}"].value

        vals = [grade, cls, num, name]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        if any(v is None or str(v).strip() == "" for v in vals):
            raise ValueError(f"[오류] 신입생 파일 {row}행(B~E)에 빈 값이 있습니다.")

        grade_s = str(grade).strip()
        m = re.search(r"\d+", grade_s)
        if not m:
            raise ValueError(f"[오류] 신입생 파일 {row}행 학년(B)에서 숫자를 찾지 못했습니다: {grade_s!r}")
        grade_i = int(m.group(0))

        cls_s = str(cls).strip()
        num_s = str(num).strip()
        name_n = normalize_name(name)

        if not name_n:
            raise ValueError(f"[오류] 신입생 파일 {row}행 성명(E) 정규화 결과가 비어 있습니다.")

        out.append({"grade": grade_i, "class": cls_s, "number": num_s, "name": name_n})
        row += 1

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, x)

    out.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"])))
    return out

# 전입생 파일
def read_transfer_rows(
    xlsx_path: Path,
    header_row: int = 2,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    expected = ["no", "학년", "반", "번호", "성명", "비고"]
    for col, exp in zip(["A", "B", "C", "D", "E", "F"], expected):
        v = ws[f"{col}{header_row}"].value
        v = "" if v is None else re.sub(r"\s+", "", str(v)).lower()
        if v != re.sub(r"\s+", "", exp).lower():
            raise ValueError("[오류] 전입생 파일 헤더(2행 A~F)가 양식과 다릅니다.")

    if data_start_row is None:
        _, data_start_row = detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=5,  # E열: 성명
        )

    out = []
    row = data_start_row
    while True:
        grade = ws[f"B{row}"].value
        cls   = ws[f"C{row}"].value
        num   = ws[f"D{row}"].value
        name  = ws[f"E{row}"].value

        vals = [grade, cls, num, name]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        if any(v is None or str(v).strip() == "" for v in vals):
            raise ValueError(f"[오류] 전입생 파일 {row}행(B~E)에 빈 값이 있습니다.")

        grade_i = int(str(grade).strip())
        cls_s = str(cls).strip()
        num_s = str(num).strip()
        name_n = normalize_name(name)
        if not name_n:
            raise ValueError(f"[오류] 전입생 파일 {row}행 성명(E) 정규화 결과가 비어 있습니다.")

        out.append({"grade": grade_i, "class": cls_s, "number": num_s, "name": name_n})
        row += 1

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, x)

    out.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"])))
    return out

# 교사 아이디 파일
def read_teacher_rows(
    xlsx_path: Path,
    header_row: int = 3,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    expected = ["NO", "직위,담당", "선생님이름", "학습용ID신청", "관리용ID신청"]
    for col, exp in zip(["A", "B", "C", "D", "E"], expected):
        v = ws[f"{col}{header_row}"].value
        v = "" if v is None else re.sub(r"\s+", "", str(v))
        if v != re.sub(r"\s+", "", exp):
            raise ValueError("[오류] 교사 파일 헤더(3행 A~E)가 양식과 다릅니다.")

    if data_start_row is None:
        _, data_start_row = detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=3,  # C열: 선생님 이름
        )

    out = []
    row = data_start_row
    while True:
        b = ws[f"B{row}"].value
        c = ws[f"C{row}"].value
        d = ws[f"D{row}"].value
        e = ws[f"E{row}"].value

        if all(v is None or str(v).strip() == "" for v in [b, c, d, e]):
            break

        if c is None or str(c).strip() == "":
            row += 1
            continue

        name_n = normalize_name(c)
        if not name_n:
            row += 1
            continue

        learn_apply = not (d is None or str(d).strip() == "")
        admin_apply = not (e is None or str(e).strip() == "")

        out.append(
            {
                "position": "" if b is None else str(b).strip(),
                "name": name_n,
                "learn_apply": learn_apply,
                "admin_apply": admin_apply,
            }
        )
        row += 1

    return out

def normalize_withdraw_class(raw, grade: int) -> str:
    if raw is None:
        return ""

    s = str(raw).strip()
    if not s:
        return ""

    s = re.sub(r"\s+", "", s)

    # 1-1, 1-1반, 1학년1반 같은 케이스 정리
    m = re.search(r"(\d+)[-학년]*(\d+)", s)
    if m:
        g = int(m.group(1))
        c = int(m.group(2))
        return f"{g}-{c}반"

    # 숫자만 있는 경우
    if s.isdigit():
        return f"{grade}-{int(s)}반"

    return s


# 전출생 파일
def read_withdraw_rows(
    xlsx_path: Path,
    header_row: int = 2,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    expected = ["no", "학년", "반", "성명", "비고"]
    for col, exp in zip(["A", "B", "C", "D", "E"], expected):
        v = ws[f"{col}{header_row}"].value
        v = "" if v is None else re.sub(r"\s+", "", str(v)).lower()
        if v != re.sub(r"\s+", "", exp).lower():
            raise ValueError("[오류] 전출생 파일 헤더(2행 A~E)가 양식과 다릅니다.")

    if data_start_row is None:
        _, data_start_row = detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=4,  # D열: 성명
        )

    out = []
    row = data_start_row
    while True:
        grade = ws[f"B{row}"].value
        cls   = ws[f"C{row}"].value
        name  = ws[f"D{row}"].value

        vals = [grade, cls, name]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        if any(v is None or str(v).strip() == "" for v in vals):
            raise ValueError(f"[오류] 전출생 파일 {row}행(B~D)에 빈 값이 있습니다.")

        grade_i = int(str(grade).strip())
        cls_s = normalize_withdraw_class(cls, grade_i)
        if not cls_s:
            raise ValueError(f"[오류] 전출생 파일 {row}행 반(C) 정규화 결과가 비어 있습니다.")

        name_n = normalize_name(name)
        if not name_n:
            raise ValueError(f"[오류] 전출생 파일 {row}행 성명(D) 정규화 결과가 비어 있습니다.")

        out.append({"grade": grade_i, "class": cls_s, "name": name_n})
        row += 1

    return out

# =========================
# sheet utilities
# =========================
def header_map(ws, header_row: int = 1):
    mapping = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        key = str(cell.value)
        key = key.replace("\u00A0", " ")
        key = re.sub(r"\s+", "", key)
        key = key.replace(".", "")
        mapping[key] = cell.column
    return mapping


def find_last_data_row(ws, key_col: int, start_row: int) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is not None and str(v).strip() != "":
            last = r
    return last


def clear_sheet_rows(ws, start_row=2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def move_sheet_after(wb, sheet_name: str, after_name: str):
    if sheet_name not in wb.sheetnames or after_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    wb._sheets.remove(ws)
    idx = wb.sheetnames.index(after_name)
    wb._sheets.insert(idx + 1, ws)


def delete_rows_below(ws, last_keep_row: int):
    if ws.max_row > last_keep_row:
        ws.delete_rows(last_keep_row + 1, ws.max_row - last_keep_row)


def clear_format_workbook_from_row(wb, start_row: int = 2):
    """
    모든 시트에서:
    - start_row부터 실제 데이터가 있는 마지막 행까지 스캔
    - 그 아래 행들에 대해서만 서식(fill, border) 제거
    """
    for ws in wb.worksheets:
        last_data_row = 0
        max_row = ws.max_row
        max_col = ws.max_column or 1

        # 실제 데이터 마지막 행 찾기
        for r in range(start_row, max_row + 1):
            row_has_value = False
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip() != "":
                    row_has_value = True
                    break
            if row_has_value:
                last_data_row = r

        if last_data_row == 0:
            continue

        # 마지막 데이터 행 아래부터 서식 제거
        for r in range(last_data_row + 1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()


def reset_view_to_a1(wb):
    """
    - 모든 시트: 화면은 A1, 커서는 A2
    - 모든 시트: 1행 고정(freeze_panes = A2)
    - 모든 시트: 그룹 선택(tabSelected) 해제
    - 통합문서: 첫 번째 시트만 선택 + 활성
    """
    # 1) 공통 뷰/고정 설정
    for ws in wb.worksheets:
        sv = ws.sheet_view

        # 화면/커서
        sv.topLeftCell = "A1"
        sv.activeCell = "A2"
        sv.selection = [Selection(activeCell="A2", sqref="A2")]

        # 1행 고정
        ws.freeze_panes = "A2"

        # 시트 그룹 선택 풀기
        if hasattr(sv, "tabSelected"):
            sv.tabSelected = False

    # 2) 첫 번째 시트만 선택 + 활성
    first_ws = wb.worksheets[0]
    if hasattr(first_ws.sheet_view, "tabSelected"):
        first_ws.sheet_view.tabSelected = True

    wb.active = 0

    # 3) 통합문서 뷰도 첫 시트 기준으로 통일
    if getattr(wb, "views", None):
        views = wb.views
        if views:
            views[0].activeTab = 0
            views[0].firstSheet = 0

# =========================
# roster analyze
# =========================
def parse_roster_year_from_filename(roster_path: Path) -> Optional[int]:
    stem = roster_path.stem
    s = stem.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()

    m = re.search(r"(\d{4})\s*학\s*년도", s)
    if m:
        return int(m.group(1))

    m2 = re.search(r"(19\d{2}|20\d{2})", s)
    if m2:
        return int(m2.group(1))

    return None


def load_roster_sheet(dirs: Dict[str, Path], school_name: str):
    """
    학생명부(.xlsx, 파일명에 '학생명부' 포함)를 학교 폴더에서 찾아서
    - 첫 번째 시트를 openpyxl 워크시트로 반환
    - 파일 경로
    - 파일명 기준 추정 학년도 (없으면 None)
    를 돌려준다.
    """
    root_dir = dirs["SCHOOL_ROOT"]

    kw = (school_name or "").strip()
    if not kw:
        raise ValueError("[오류] 학교명이 비어 있어 학생명부 폴더를 찾을 수 없습니다.")

    # 🔹 학교 폴더를 포함 매칭으로 찾기
    matches = [
        p
        for p in root_dir.iterdir()
        if p.is_dir() and text_contains(p.name, kw)
    ]

    if not matches:
        raise ValueError(
            f"[오류] 학생명부를 찾을 학교 폴더를 찾지 못했습니다. "
            f"(작업 폴더 내 '{school_name}' 포함 폴더 없음)"
        )

    if len(matches) > 1:
        raise ValueError(
            f"[오류] 학생명부를 찾을 학교 폴더 후보가 여러 개입니다: "
            + ", ".join(p.name for p in matches)
        )

    school_root = matches[0]

    candidates: List[Path] = [
        p
        for p in school_root.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".xlsx"
        and "학생명부" in p.stem
        and not p.name.startswith("~$")
    ]
    if not candidates:
        raise ValueError("[오류] 학생명부(.xlsx, 파일명에 '학생명부') 파일을 찾지 못했습니다.")

    # 가장 최근 수정 파일 사용
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    roster_path = candidates[0]

    wb = safe_load_workbook(roster_path, data_only=True)
    ws = wb.worksheets[0]
    roster_year = parse_roster_year_from_filename(roster_path)

    return ws, roster_path, roster_year

def parse_class_str(s: str) -> Optional[Tuple[int, str]]:
    if s is None:
        return None
    m = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", str(s))
    if not m:
        return None
    return int(m.group(1)), m.group(2).strip()


def extract_id_prefix4(uid: str) -> Optional[int]:
    if uid is None:
        return None
    s = str(uid).strip()
    if len(s) >= 4 and s[:4].isdigit():
        return int(s[:4])
    return None


def analyze_roster_once(roster_ws, input_year: int) -> Dict:
    hm = header_map(roster_ws, 1)
    need = ["현재반", "이전반", "학생이름", "아이디"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[오류] 학생명부에 '{k}' 헤더가 없습니다.")

    c_class = hm["현재반"]
    c_name  = hm["학생이름"]
    c_id    = hm["아이디"]

    prefixes_by_grade = defaultdict(list)
    name_counter_by_grade = defaultdict(Counter)
    prefixes_grade1 = []

    for r in range(2, roster_ws.max_row + 1):
        clv = roster_ws.cell(r, c_class).value
        nmv = roster_ws.cell(r, c_name).value
        idv = roster_ws.cell(r, c_id).value
        if clv is None or nmv is None:
            continue

        parsed = parse_class_str(clv)
        if parsed is None:
            continue
        g, _cls = parsed

        nm = normalize_name(nmv)
        if not nm:
            continue
        name_counter_by_grade[g][nm] += 1

        p4 = extract_id_prefix4(idv)
        if p4 is not None:
            prefixes_by_grade[g].append(p4)
            if g == 1:
                prefixes_grade1.append(p4)

    prefix_mode_by_grade = {}
    for g, arr in prefixes_by_grade.items():
        if arr:
            prefix_mode_by_grade[g] = Counter(arr).most_common(1)[0][0]

    roster_time = "unknown"
    ref_shift = 0
    if prefixes_grade1:
        mode1 = Counter(prefixes_grade1).most_common(1)[0][0]
        if mode1 == input_year:
            roster_time = "this_year"
            ref_shift = 0
        elif mode1 == input_year - 1:
            roster_time = "last_year"
            ref_shift = -1
        else:
            roster_time = "unknown"
            ref_shift = 0

    return {
        "roster_time": roster_time,
        "ref_grade_shift": ref_shift,
        "prefix_mode_by_roster_grade": prefix_mode_by_grade,
        "name_count_by_roster_grade": name_counter_by_grade,
    }


# =========================
# transfer ids
# =========================
def build_transfer_ids(
    transfer_rows: List[Dict],
    roster_info: Dict,
    input_year: int,
) -> Tuple[List[Dict], List[Dict], Dict[int, int]]:
    shift = roster_info["ref_grade_shift"]
    prefix_mode = roster_info["prefix_mode_by_roster_grade"]
    name_counts = roster_info["name_count_by_roster_grade"]

    done: List[Dict] = []
    hold: List[Dict] = []
    final_prefix_by_current_grade: Dict[int, int] = {}
    seen_in_transfer_by_grade = defaultdict(Counter)

    grade1_rows = [tr for tr in transfer_rows if tr["grade"] == 1]
    if grade1_rows:
        g1_names = [tr["name"] for tr in grade1_rows]
        g1_names_sfx = apply_suffix_for_duplicates(g1_names)
        for tr, nm_sfx in zip(grade1_rows, g1_names_sfx):
            uid = f"{input_year}{nm_sfx}"
            done.append({**tr, "id": uid})

    other_rows = [tr for tr in transfer_rows if tr["grade"] != 1]

    for tr in other_rows:
        g_cur = tr["grade"]
        g_roster = g_cur + shift

        pref = prefix_mode.get(g_roster)
        if pref is None:
            hold.append({**tr, "보류사유": f"명부 학년({g_roster})에서 ID prefix 최빈값 산출 불가"})
            continue

        final_prefix_by_current_grade[g_cur] = pref

        nm = tr["name"]
        base_cnt = name_counts.get(g_roster, Counter()).get(nm, 0)

        seen_in_transfer_by_grade[g_cur][nm] += 1
        add_seq = seen_in_transfer_by_grade[g_cur][nm]

        need_suffix = (base_cnt > 0)
        suffix = dedup_suffix_letters(add_seq) if need_suffix else ""

        uid = f"{pref}{nm}{suffix}"
        done.append({**tr, "id": uid})

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, str(x))

    done.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))
    hold.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))

    return done, hold, final_prefix_by_current_grade


# =========================
# withdraw outputs
# =========================
def build_withdraw_outputs(
    roster_ws,
    withdraw_rows: List[Dict],
    school_start_date: date,
    work_date: date,
    roster_info: Optional[Dict] = None,
) -> Tuple[List[Dict], List[Dict]]:
    """
    학생명부 + 전출 명단 기반 퇴원/보류 리스트 생성.
    - 퇴원일자: 작업일 < 개학일 → 개학일, 그 외에는 작업일 기준
    """
    # 🔹 결과 리스트
    done: List[Dict] = []
    hold: List[Dict] = []

    # 🔹 퇴원일자 계산 (파일 전체 공통)
    eff = school_start_date if work_date < school_start_date else work_date

    hm = header_map(roster_ws, 1)
    need = ["현재반", "이전반", "학생이름", "아이디"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[오류] 학생명부에 '{k}' 헤더가 없습니다.")

    col_now   = hm["현재반"]
    col_prev  = hm["이전반"]
    col_name  = hm["학생이름"]
    col_id    = hm["아이디"]

    # scan에서 넘겨준 학년도 판정 활용
    roster_time = (roster_info or {}).get("roster_time", "this_year")
    use_prev_for_grade = (roster_time == "last_year")

    roster_map: Dict[str, List[Dict]] = {}
    roster_by_grade_name: Dict[str, List[Dict]] = {}

    # 🔹 학년+이름 인덱스 중복 방지용
    seen_grade_name_ids = set()  # (grade, name_key, id_str)

    def _index_class_map(class_val, name_key: str, idv, name_disp: str):
        """반+이름 완전 매칭용 인덱스 (현재반/이전반 둘 다)"""
        if class_val is None:
            return
        c = str(class_val).strip()
        if not c:
            return

        key1 = f"{c}|{name_key}"
        roster_map.setdefault(key1, []).append(
            {
                "class": c,
                "name_key": name_key,
                "name_disp": name_disp,
                "id": "" if idv is None else str(idv).strip(),
            }
        )

    def _index_grade_map(class_val, name_key: str, idv, name_disp: str):
        """학년+이름 fallback용 (pre/post 기준으로 고른 반만 사용)"""
        if class_val is None:
            return
        c = str(class_val).strip()
        if not c:
            return

        parsed = parse_class_str(c)
        if parsed is None:
            return
        g = parsed[0]

        id_str = "" if idv is None else str(idv).strip()
        dedup_key = (g, name_key, id_str)
        if dedup_key in seen_grade_name_ids:
            return
        seen_grade_name_ids.add(dedup_key)

        key2 = f"{g}|{name_key}"
        roster_by_grade_name.setdefault(key2, []).append(
            {
                "class": c,
                "name_key": name_key,
                "name_disp": name_disp,
                "id": id_str,
                "grade": g,
            }
        )

    # 학생명부 인덱스 생성
    for r in range(2, roster_ws.max_row + 1):
        nmv = roster_ws.cell(r, col_name).value
        if nmv is None:
            continue
        name_disp = normalize_name(nmv)
        name_key  = normalize_name_key(nmv)
        if not name_key:
            continue

        idv = roster_ws.cell(r, col_id).value
        nowv  = roster_ws.cell(r, col_now).value
        prevv = roster_ws.cell(r, col_prev).value

        # 1) 반+이름 완전 매칭용: 현재반/이전반 둘 다
        _index_class_map(nowv,  name_key, idv, name_disp)
        _index_class_map(prevv, name_key, idv, name_disp)

        # 2) 학년+이름 fallback용: pre/post 기준으로 고른 한 열만
        base_class_val = prevv if use_prev_for_grade else nowv
        _index_grade_map(base_class_val, name_key, idv, name_disp)

    # 전출 행 처리
    for w in withdraw_rows:
        g_cur = w["grade"]
        w_name_disp = w["name"]
        w_name_key  = normalize_name_key(w_name_disp)
        if not w_name_key:
            hold.append(
                {
                    "학년": g_cur,
                    "반": w["class"],
                    "성명": w_name_disp,
                    "보류사유": "성명 정규화(키) 결과가 비어 있음",
                }
            )
            continue

        # 전출 명단의 반(C열)은 이미 normalize_withdraw_class로 통일된 상태라고 가정
        w_class_full = w["class"]
        key = f"{w_class_full}|{w_name_key}"
        matches = roster_map.get(key, [])

        if len(matches) == 0:
            # 같은 학년/다음 학년에서 이름만 일치하는 후보 찾아보기
            cand0 = roster_by_grade_name.get(f"{g_cur}|{w_name_key}", [])
            cand1 = roster_by_grade_name.get(f"{g_cur+1}|{w_name_key}", [])
            cand = cand0 + cand1
            if len(cand) == 1:
                matches = cand
            else:
                if len(cand) == 0:
                    reason = (
                        "자동 제외: 학생명부에 존재하지 않는 학생 – "
                        "서버 미등록/학년 불일치 등으로 추정되며 퇴원 처리 대상에서 제외했습니다. "
                        "(반 매칭 실패, g 또는 g+1 탐색)"
                    )
                else:
                    reason = (
                        f"보류: 학년+이름 후보가 2건 이상({len(cand)}건) – 수동 확인 필요. "
                        "(반 매칭 실패, g 또는 g+1 탐색)"
                    )
                hold.append(
                    {
                        "학년": g_cur,
                        "반": w["class"],
                        "성명": w_name_disp,
                        "보류사유": reason,
                    }
                )
                continue

        if len(matches) > 1:
            hold.append(
                {
                    "학년": g_cur,
                    "반": w["class"],
                    "성명": w_name_disp,
                    "보류사유": f"중복 매칭({len(matches)}건)",
                }
            )
            continue

        m = matches[0]
        g_server = m.get("grade")
        if g_server is None:
            parsed = parse_class_str(m.get("class", ""))
            g_server = parsed[0] if parsed else g_cur

        withdraw_class = f"{g_server}-미편성반"
        done.append(
            {
                "퇴원반명": withdraw_class,
                "학생이름": w_name_disp,
                "아이디": m["id"],
                "퇴원일자": eff,
            }
        )

    return done, hold


def write_withdraw_to_register(wb, done_rows: List[Dict], hold_rows: List[Dict]):
    ws_done = wb["퇴원"] if "퇴원" in wb.sheetnames else wb.create_sheet("퇴원")
    ws_hold = wb["퇴원_보류"] if "퇴원_보류" in wb.sheetnames else wb.create_sheet("퇴원_보류")

    # 퇴원 완료 정렬 (퇴원반명 → 학생이름 오름차순)
    done_rows = sorted(
        done_rows,
        key=lambda r: (
            str(r.get("퇴원반명", "")).strip(),
            str(r.get("학생이름", "")).strip(),
        ),
    )

    # 보류 정렬 (학년 → 반 → 성명)
    hold_rows = sorted(
        hold_rows,
        key=lambda r: (
            str(r.get("학년", "")).strip(),
            str(r.get("반", "")).strip(),
            str(r.get("성명", "")).strip(),
        ),
    )

    clear_sheet_rows(ws_done, 2)
    clear_sheet_rows(ws_hold, 2)

    r = 2
    for row in done_rows:
        ws_done.cell(r, 1).value = row["퇴원반명"]
        ws_done.cell(r, 2).value = row["학생이름"]
        ws_done.cell(r, 3).value = row["아이디"]
        ws_done.cell(r, 4).value = row["퇴원일자"]
        ws_done.cell(r, 4).number_format = "yyyy-mm-dd"
        r += 1

    r = 2
    for row in hold_rows:
        ws_hold.cell(r, 1).value = row["학년"]
        ws_hold.cell(r, 2).value = row["반"]
        ws_hold.cell(r, 3).value = row["성명"]
        ws_hold.cell(r, 4).value = row["보류사유"]
        r += 1

    move_sheet_after(wb, "퇴원_보류", "퇴원")

    from openpyxl.styles import Font, Alignment

    def _format_sheet(ws):
        for rr in range(1, ws.max_row + 1):
            for cc in range(1, ws.max_column + 1):
                cell = ws.cell(rr, cc)
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    _format_sheet(ws_done)
    _format_sheet(ws_hold)


# =========================
# register fill (rebuild)
# =========================
def school_kind_from_name(school_name: str) -> Tuple[str, str]:
    s = (school_name or "").strip()
    if not s:
        return "", ""
    last = s[-1]
    if last == "초":
        return "초등부", "초"
    if last == "중":
        return "중등부", "중"
    if last == "고":
        return "고등부", "고"
    return "", ""


def write_transfer_hold_sheet(wb, hold_rows: List[Dict]):
    sheet_name = "전입생_보류"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    ws.delete_rows(1, ws.max_row)

    ws["A1"].value = "학년"
    ws["B1"].value = "반"
    ws["C1"].value = "번호"
    ws["D1"].value = "성명"
    ws["E1"].value = "보류사유"

    r = 2
    for row in hold_rows:
        ws.cell(r, 1).value = row.get("grade", "")
        ws.cell(r, 2).value = row.get("class", "")
        ws.cell(r, 3).value = row.get("number", "")
        ws.cell(r, 4).value = row.get("name", "")
        ws.cell(r, 5).value = row.get("보류사유", "")
        r += 1


def fill_register(
    template_path: Path,
    out_path: Path,
    school_name: str,
    year: str,
    freshmen_rows: List[Dict],
    transfer_done_rows: List[Dict],
    teacher_rows: List[Dict],
    transfer_hold_rows: Optional[List[Dict]] = None,
    withdraw_done_rows: Optional[List[Dict]] = None,
    withdraw_hold_rows: Optional[List[Dict]] = None,
) -> None:
    ensure_xlsx_only(template_path)

    wb = load_workbook(template_path)
    ws_students = wb["학생자료"]
    ws_staff = wb["직원정보"]
    ws_groups = wb["그룹반정보"]

    # =========================
    # [학생자료] 컬럼 매핑
    # =========================
    hm = header_map(ws_students, 1)
    need = ["No", "학생이름", "ID", "학교구분", "학교", "학년", "수강반"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[오류] 템플릿 [학생자료]에 '{k}' 헤더가 없습니다.")

    col_no = hm["No"]
    col_name = hm["학생이름"]
    col_id = hm["ID"]
    col_kind = hm["학교구분"]
    col_school = hm["학교"]
    col_grade = hm["학년"]
    col_class = hm["수강반"]

    # 기존 데이터 clear
    for r in range(2, ws_students.max_row + 1):
        for c in [col_no, col_name, col_id, col_kind, col_school, col_grade, col_class]:
            ws_students.cell(row=r, column=c).value = None

    kind_full, kind_prefix = school_kind_from_name(school_name)

    def write_student_row(r: int, no: int, name: str, uid: str, grade_i: int, cls_name: str):
        ws_students.cell(r, col_no).value = no
        ws_students.cell(r, col_name).value = name
        ws_students.cell(r, col_id).value = uid
        ws_students.cell(r, col_kind).value = kind_full if kind_full else ""
        ws_students.cell(r, col_school).value = school_name
        ws_students.cell(r, col_grade).value = f"{kind_prefix}{grade_i}" if kind_prefix else ""
        ws_students.cell(r, col_class).value = cls_name

    write_row = 2
    running_no = 1

    # 신입생 ID: 학년도 + 이름(중복 suffix 포함)
    fn_names = [r["name"] for r in freshmen_rows]
    fn_names_sfx = apply_suffix_for_duplicates(fn_names)
    fn_ids = [f"{year}{nm}" for nm in fn_names_sfx]

    for i, fr in enumerate(freshmen_rows):
        r = write_row + i
        write_student_row(
            r=r,
            no=running_no,
            name=fr["name"],
            uid=fn_ids[i],
            grade_i=fr["grade"],
            cls_name=f"{fr['grade']}-{fr['class']}",
        )
        running_no += 1
    write_row += len(freshmen_rows)

    # 전입생(완료)
    for tr in transfer_done_rows:
        r = write_row
        write_student_row(
            r=r,
            no=running_no,
            name=tr["name"],
            uid=tr["id"],
            grade_i=tr["grade"],
            cls_name=f"{tr['grade']}-{tr['class']}",
        )
        running_no += 1
        write_row += 1

    # 선생님(학습용) → 학생자료에 "선생님반"
    teachers_learn = [t for t in teacher_rows if t["learn_apply"]]
    t_names = [t["name"] for t in teachers_learn]
    t_names_sfx = apply_suffix_for_duplicates(t_names)
    t_ids = [f"{nm}1" for nm in t_names_sfx]

    for j, t in enumerate(teachers_learn):
        r = write_row + j
        write_student_row(
            r=r,
            no=running_no,
            name=t["name"],
            uid=t_ids[j],
            grade_i=1,
            cls_name="선생님반",
        )
        running_no += 1
    write_row += len(teachers_learn)

    # =========================
    # [직원정보]
    # =========================
    hm2 = header_map(ws_staff, 1)
    hm2_lower = {k.lower(): v for k, v in hm2.items()}

    need2 = ["no", "이름", "아이디", "권한부여"]
    for k in need2:
        if k.lower() not in hm2_lower:
            raise ValueError(f"[오류] 템플릿 [직원정보]에 '{k}' 헤더가 없습니다.")

    col_s_no = hm2_lower["no"]
    col_s_name = hm2_lower["이름"]
    col_s_id = hm2_lower["아이디"]
    col_s_role = hm2_lower["권한부여"]

    for r in range(2, ws_staff.max_row + 1):
        for c in [col_s_no, col_s_name, col_s_id, col_s_role]:
            ws_staff.cell(row=r, column=c).value = None

    teachers_admin = [t for t in teacher_rows if t["admin_apply"]]
    a_names = [t["name"] for t in teachers_admin]
    a_names_sfx = apply_suffix_for_duplicates(a_names)

    staff_write = 2
    for i, t in enumerate(teachers_admin):
        r = staff_write + i
        ws_staff.cell(r, col_s_no).value = i + 1
        ws_staff.cell(r, col_s_name).value = t["name"]
        ws_staff.cell(r, col_s_id).value = a_names_sfx[i]
        ws_staff.cell(r, col_s_role).value = "선생님"

    # =========================
    # [그룹반정보]
    # =========================
    hm_g = header_map(ws_groups, 1)
    need_g = ["그룹명", "반명", "수강료", "담임명", "FullMode"]
    for k in need_g:
        if k not in hm_g:
            raise ValueError(f"[오류] 템플릿 [그룹반정보]에 '{k}' 헤더가 없습니다.")

    col_g_group = hm_g["그룹명"]
    col_g_class = hm_g["반명"]
    col_g_fee = hm_g["수강료"]
    col_g_teacher = hm_g["담임명"]
    col_g_full = hm_g["FullMode"]

    for r in range(2, ws_groups.max_row + 1):
        for c in [col_g_group, col_g_class, col_g_fee, col_g_teacher, col_g_full]:
            ws_groups.cell(row=r, column=c).value = None

    class_set = set()
    last_student_row = find_last_data_row(ws_students, key_col=col_no, start_row=2)
    for r in range(2, last_student_row + 1):
        v = ws_students.cell(row=r, column=col_class).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            class_set.add(s)

    def parse_grade_prefix(class_name: str):
        m = re.match(r"^\s*(\d+)\s*-\s*(.+)\s*$", str(class_name))
        if not m:
            return None
        return int(m.group(1))

    def group_name_from_class(class_name: str) -> str:
        if class_name == "선생님반":
            return "기타그룹"
        g = parse_grade_prefix(class_name)
        if g is None:
            return "기타그룹"
        return f"{g}학년"

    def class_sort_key(class_name: str):
        if class_name == "선생님반":
            return (2, 0, "zzz")
        g = parse_grade_prefix(class_name)
        if g is None:
            return (1, 0, str(class_name))
        return (0, g, str(class_name))

    class_list = sorted(class_set, key=class_sort_key)

    start_r = 2
    for i, cls_name in enumerate(class_list):
        r = start_r + i
        ws_groups.cell(r, col_g_group).value = group_name_from_class(cls_name)
        ws_groups.cell(r, col_g_class).value = cls_name
        ws_groups.cell(r, col_g_fee).value = None
        ws_groups.cell(r, col_g_teacher).value = "선생님"
        ws_groups.cell(r, col_g_full).value = "Y"

    # 전입 보류 시트
    if transfer_hold_rows:
        write_transfer_hold_sheet(wb, transfer_hold_rows)

    # 전출 완료/보류 시트
    if (withdraw_done_rows is not None) and (withdraw_hold_rows is not None):
        write_withdraw_to_register(wb, withdraw_done_rows, withdraw_hold_rows)

    # 워크북 전체: 빈 행 아래 서식 제거 + A1로 통일
    clear_format_workbook_from_row(wb, start_row=2)
    reset_view_to_a1(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    backup_if_exists(out_path)
    wb.save(out_path)


# =========================
# NOTICE FILE (ID/PW 안내) 생성
# =========================
FILL_TRANSFER = PatternFill("solid", fgColor="F8CBAD")  # 옅은 주황
FILL_DUP      = PatternFill("solid", fgColor="FFFF00")  # 노랑
FILL_GREY     = PatternFill("solid", fgColor="D9D9D9")  # 회색


def _is_duplicate_id(uid: str) -> bool:
    if uid is None:
        return False
    s = str(uid).strip()
    if not s:
        return False
    # 동명이인: 아이디 끝이 대문자 A~Z (A, B, ..., AA 등)
    return bool(re.search(r"[A-Z]+$", s))


def _parse_grade_class_from_register(class_str: str) -> Tuple[Optional[int], str]:
    """
    register의 수강반: "1-3" 같은 형태 → (1, "3")
    실패하면 (None, 원본)
    """
    if class_str is None:
        return None, ""
    s = str(class_str).strip()
    if not s:
        return None, ""
    m = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", s)
    if not m:
        return None, s
    return int(m.group(1)), m.group(2).strip()


def build_notice_student_sheet(
    ws_notice,
    register_students_ws,
    transfer_ids: set,
):
    """
    안내파일 - 학생 ID,PW(학습용)
    헤더 3행: No., 학년, 반, 학생이름, ID, PW
    데이터 4행부터
    """
    hm_r = header_map(register_students_ws, 1)
    need_r = ["No", "학생이름", "ID", "수강반"]
    for k in need_r:
        if k not in hm_r:
            raise ValueError(f"[오류] 등록작업파일 [학생자료]에 '{k}' 헤더가 없습니다.")
    c_r_name = hm_r["학생이름"]
    c_r_id   = hm_r["ID"]
    c_r_cls  = hm_r["수강반"]
    c_r_no   = hm_r["No"]

    header_row = 3
    start_row = 4

    out_rows: List[Dict[str, Any]] = []
    last_r = find_last_data_row(register_students_ws, key_col=c_r_no, start_row=2)
    for r in range(2, last_r + 1):
        nm = register_students_ws.cell(r, c_r_name).value
        uid = register_students_ws.cell(r, c_r_id).value
        cls = register_students_ws.cell(r, c_r_cls).value

        cls_str = "" if cls is None else str(cls).strip()
        if cls_str == "선생님반":
            continue  # 학생 안내에서 제외

        if (nm is None or str(nm).strip() == "") and (uid is None or str(uid).strip() == ""):
            continue
        nm_s = "" if nm is None else str(nm).strip()
        uid_s = "" if uid is None else str(uid).strip()
        if not uid_s:
            continue

        g, cls_only = _parse_grade_class_from_register(cls)
        if g is None:
            g_disp = ""
            cls_disp = "" if cls is None else str(cls).strip()
        else:
            g_disp = g
            cls_disp = cls_only

        out_rows.append(
            {
                "name": nm_s,
                "id": uid_s,
                "grade": g_disp,
                "class": cls_disp,
                "is_transfer": (uid_s in transfer_ids),
                "is_dup": _is_duplicate_id(uid_s),
            }
        )

    r_out = start_row
    no = 1
    for item in out_rows:
        ws_notice.cell(r_out, 1).value = no
        ws_notice.cell(r_out, 2).value = item["grade"]
        ws_notice.cell(r_out, 3).value = item["class"]
        ws_notice.cell(r_out, 4).value = item["name"]
        ws_notice.cell(r_out, 5).value = item["id"]
        ws_notice.cell(r_out, 6).value = "1234" if item["id"] else ""

        fill = None
        if item["is_dup"]:
            fill = FILL_DUP
        elif item["is_transfer"]:
            fill = FILL_TRANSFER

        if fill is not None:
            for c in range(1, 7):
                ws_notice.cell(r_out, c).fill = fill

        no += 1
        r_out += 1

    delete_rows_below(ws_notice, r_out - 1)


def build_notice_teacher_sheet(
    ws_notice,
    teacher_rows: List[Dict],
):
    """
    안내파일 - 선생님ID,PW(관리용,학습용)
    헤더 3행, 데이터 4행부터.
    - No, 직위, 선생님이름: teacher_rows의 position/name 그대로
    - 관리용ID: admin_apply True → name, PW는 t1234
    - 학습용ID: learn_apply True → name+'1', PW는 1234
    - 신청 안 한 칸은 회색 처리
    """
    header_row = 3
    start_row = 4

    # 직위(B열) 컬럼 폭 확장 (긴 직위/담당 명칭 잘리지 않도록)
    try:
        ws_notice.column_dimensions["B"].width = 16.6
    except Exception:
        # 열 정보가 없거나 시트 구조가 다른 경우에도 전체 로직은 계속 진행
        pass

    r_out = start_row
    no = 1
    for t in teacher_rows:
        pos = "" if t.get("position") is None else str(t.get("position")).strip()
        nm  = "" if t.get("name") is None else str(t.get("name")).strip()
        if not nm and not pos and (not t.get("learn_apply")) and (not t.get("admin_apply")):
            continue

        admin_apply = bool(t.get("admin_apply"))
        learn_apply = bool(t.get("learn_apply"))

        admin_id = nm if admin_apply else ""
        admin_pw = "t1234" if admin_id else ""

        learn_id = f"{nm}1" if learn_apply else ""
        learn_pw = "1234" if learn_id else ""

        # A: No. / B: 직위 / C: 선생님이름 / D: 구분용 빈 칸
        # E: 관리용 ID / F: PW / G: 구분용 빈 칸 / H: 학습용 ID / I: PW
        ws_notice.cell(r_out, 1).value = no
        ws_notice.cell(r_out, 2).value = pos
        ws_notice.cell(r_out, 3).value = nm
        ws_notice.cell(r_out, 5).value = admin_id
        ws_notice.cell(r_out, 6).value = admin_pw
        ws_notice.cell(r_out, 8).value = learn_id
        ws_notice.cell(r_out, 9).value = learn_pw

        # 회색 처리(신청 안 한 영역)
        if not admin_apply:
            for c in [5, 6]:
                ws_notice.cell(r_out, c).fill = FILL_GREY

        if not learn_apply:
            for c in [8, 9]:
                ws_notice.cell(r_out, c).fill = FILL_GREY

        no += 1
        r_out += 1

    delete_rows_below(ws_notice, r_out - 1)


def build_notice_file(
    template_notice_path: Path,
    out_notice_path: Path,
    out_register_path: Path,
    teacher_file_path: Optional[Path],
    transfer_done_rows: List[Dict],
) -> None:
    ensure_xlsx_only(template_notice_path)
    ensure_xlsx_only(out_register_path)

    wb_notice = safe_load_workbook(template_notice_path, data_only=False)
    wb_reg = load_workbook(out_register_path)

    if "학생자료" not in wb_reg.sheetnames:
        raise ValueError("[오류] 등록작업파일에 '학생자료' 시트가 없습니다.")

    ws_reg_students = wb_reg["학생자료"]

    def _norm_sheetname(s: str) -> str:
        if s is None:
            return ""
        s = str(s)
        s = s.replace("\u00A0", " ")
        s = re.sub(r"\s+", "", s)
        return s

    def _pick_sheet_by_keywords(wb, keywords: List[str]) -> str:
        keys = [_norm_sheetname(k) for k in keywords]
        for name in wb.sheetnames:
            n = _norm_sheetname(name)
            if all(k in n for k in keys):
                return name
        raise ValueError(
            "[오류] 안내 템플릿에서 필요한 시트를 찾지 못했습니다.\n"
            f"- keywords: {keywords}\n"
            f"- sheetnames: {wb.sheetnames}"
        )

    sh_student = _pick_sheet_by_keywords(wb_notice, ["학생", "PW", "학습용"])
    sh_teacher = _pick_sheet_by_keywords(wb_notice, ["선생님", "PW"])

    ws_notice_students = wb_notice[sh_student]
    ws_notice_teachers = wb_notice[sh_teacher]

    transfer_ids = set()
    for tr in transfer_done_rows:
        uid = tr.get("id")
        if uid:
            transfer_ids.add(str(uid).strip())

    build_notice_student_sheet(
        ws_notice=ws_notice_students,
        register_students_ws=ws_reg_students,
        transfer_ids=transfer_ids,
    )

    teacher_rows = read_teacher_rows(teacher_file_path) if teacher_file_path else []
    build_notice_teacher_sheet(
        ws_notice=ws_notice_teachers,
        teacher_rows=teacher_rows,
    )

    out_notice_path.parent.mkdir(parents=True, exist_ok=True)
    backup_if_exists(out_notice_path)

    # 안내 파일도 워크북 공통 규칙 적용
    clear_format_workbook_from_row(wb_notice, start_row=4)
    reset_view_to_a1(wb_notice)

    wb_notice.save(out_notice_path)


# =========================
# MAIL TEMPLATE (텍스트 치환)
# =========================
def render_mail_text(
    mail_template_text: str,
    school_name: str,
    domain: str,
) -> str:
    """
    텍스트 파일 내부:
    - 'OO초'/'OO중'/'OO고' 같은 표현 → school_name
    - 'OOOOO.readinggate.com' → domain
    """
    txt = mail_template_text or ""
    if school_name:
        txt = txt.replace("OO초", school_name).replace("OO중", school_name).replace("OO고", school_name)
    if domain:
        txt = re.sub(r"[A-Za-z0-9\-]+\.readinggate\.com", domain, txt)
    return txt


def load_notice_templates(work_root: Path) -> dict[str, str]:
    dirs = get_project_dirs(work_root)
    notice_dir = dirs["NOTICES"]

    if not notice_dir.exists():
        return {}

    result = {}

    for p in notice_dir.glob("*.txt"):
        if not p.is_file():
            continue
        try:
            text = p.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = p.read_text(encoding="utf-8-sig")

        result[p.stem.strip()] = text.strip()

    return result

def domain_missing_message(school_name: str) -> str:
    _, kind_prefix = school_kind_from_name(school_name)
    kind_disp = kind_prefix if kind_prefix else "학교"
    return f"{kind_disp} (사용자가 작업중인) 의 도메인 주소가 존재하지 않습니다. 학교 전체 명단 파일을 확인하세요."


# =========================
# NEW: SCAN (pre-check)
# =========================
def scan_pipeline(
    work_root: Path,
    school_name: str,
    open_date: date,
    work_date: date,
    roster_basis_date: Optional[date] = None
) -> ScanResult:
    logs: List[str] = []

    def log(msg: str):
        logs.append(msg)

    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)

    school_name = (school_name or "").strip()
    year_str = str(open_date.year).strip()

    sr = ScanResult(
        ok=False,
        logs=logs,
        school_name=school_name,
        year_str=year_str,
        year_int=0,
        project_root=work_root,
        input_dir=Path("."),
        output_dir=Path("."),
        template_register=None,
        template_notice=None,
        db_path=None,
        freshmen_file=None,
        teacher_file=None,
        transfer_file=None,
        withdraw_file=None,
        need_roster=False,
        roster_path=None,
        roster_year=None,
        roster_info=None,
        needs_open_date=False,
        missing_fields=[],
        can_execute=False,
        can_execute_after_input=False,
    )

    try:
        if not school_name:
            raise ValueError("[오류] 학교명이 비어 있습니다.")
        year_int = int(year_str)
        sr.year_int = year_int

        db_path = school_exists_in_db(dirs["DB"], school_name)
        sr.db_path = db_path
        log(f"[OK] DB 검증 통과 | 사용 파일: {db_path.name}")

        # 🔹 학교 폴더: 선택한 학교명이 포함된 폴더 찾기 (정규화 포함 매칭)
        root_dir = dirs["SCHOOL_ROOT"]

        kw = (school_name or "").strip()
        if not kw:
            raise ValueError("[오류] 학교명이 비어 있어 학교 폴더를 찾을 수 없습니다.")

        matches = [
            p
            for p in root_dir.iterdir()
            if p.is_dir() and text_contains(p.name, kw)
        ]

        if not matches:
            raise ValueError(
                f"[오류] 작업 폴더 안에서 '{school_name}' 이(가) 포함된 학교 폴더를 찾지 못했습니다."
            )

        if len(matches) > 1:
            raise ValueError(
                f"[오류] '{school_name}' 이(가) 포함된 폴더가 여러 개입니다: "
                + ", ".join(p.name for p in matches)
            )

        school_dir = matches[0]

        # 🔹 로그: 어떤 폴더로 매칭됐는지 명확히 찍어줌
        log(f"[OK] 학교 폴더 매칭: {school_dir.name}")

        input_dir = school_dir
        output_dir = school_dir / "작업"

        sr.input_dir = input_dir
        sr.output_dir = output_dir

        # 🔹 로그: 학교 폴더 안 파일 목록 출력 (안전 처리)
        try:
            file_list = [p.name for p in input_dir.iterdir() if p.is_file()]
            log(f"[DEBUG] input files: {file_list}")
        except Exception as e:
            log(f"[WARN] 학교 폴더 파일 목록 조회 중 오류: {e}")

        freshmen_file = find_single_input_file(input_dir, FRESHMEN_KEYWORDS)
        if freshmen_file is None:
            raise ValueError("[오류] xlsx 형식의 신입생 명단 파일을 찾지 못했습니다. (키워드: 신입생/신입)")

        teacher_file  = find_single_input_file(input_dir, TEACHER_KEYWORDS)
        transfer_file = find_single_input_file(input_dir, TRANSFER_KEYWORDS)
        withdraw_file = find_single_input_file(input_dir, WITHDRAW_KEYWORDS)

        sr.freshmen_file = freshmen_file
        sr.teacher_file = teacher_file
        sr.transfer_file = transfer_file
        sr.withdraw_file = withdraw_file

        log(f"[OK] 신입생: {freshmen_file.name}")
        log(f"[OK] 교사: {teacher_file.name}" if teacher_file else "[SKIP] 교사 파일 없음 (키워드: 교사/교원)")
        log(f"[OK] 전입생: {transfer_file.name}" if transfer_file else "[SKIP] 전입생 파일 없음 (키워드: 전입생/전입)")
        log(f"[OK] 전출생: {withdraw_file.name}" if withdraw_file else "[SKIP] 전출생 파일 없음 (키워드: 전출생/전출)")

        template_register = choose_template_register(dirs["TEMPLATES"], year_str)
        sr.template_register = template_register
        log(f"[OK] 양식(등록): {template_register.name}")

        template_notice = choose_template_notice(dirs["TEMPLATES"], year_str)
        sr.template_notice = template_notice
        log(f"[OK] 양식(안내): {template_notice.name}")

        need_roster = bool(transfer_file) or bool(withdraw_file)
        sr.need_roster = need_roster

        if need_roster:
            roster_ws, roster_path, roster_year = load_roster_sheet(dirs, school_name)
            sr.roster_path = roster_path
            sr.roster_year = roster_year
            log(f"[OK] 학생명부: {roster_path.name}")

            # 1) 학생명부 마지막 수정일 → '명부 기준일'로 자동 감지
            try:
                modified_date = datetime.fromtimestamp(roster_path.stat().st_mtime).date()
                sr.roster_basis_date = modified_date

                log(
                    f"[INFO] 학생명부 마지막 수정일({modified_date.isoformat()})을 "
                    "명부 기준일로 자동 감지했습니다."
                )

                if modified_date != work_date:
                    log(
                        "[INFO] 현재 작업일과 다른 명부일입니다. "
                        "필요하면 앱에서 '명부 기준일'을 수정해 주세요."
                    )
                else:
                    log(f"[INFO] 명부 기준일이 작업일({work_date.isoformat()})과 같습니다.")
            except Exception as e:
                sr.roster_basis_date = None
                log(f"[WARN] 학생명부 마지막 수정일 확인 중 오류가 발생했습니다: {e}")

            # 2) ID prefix 기반 학년도 추정 (참고용 안내)
            try:
                expected_year = year_int
                roster_info = analyze_roster_once(roster_ws, input_year=expected_year)
                id_roster_time = roster_info.get("roster_time")  # this_year / last_year / unknown

                if id_roster_time == "this_year":
                    log(f"[INFO] 학생명부 ID 패턴 기준으로 {expected_year}학년도 명부로 추정됩니다.")
                elif id_roster_time == "last_year":
                    log(f"[INFO] 학생명부 ID 패턴 기준으로 {expected_year-1}학년도 명부로 추정됩니다.")
                else:
                    log("[INFO] 학생명부 ID 패턴 기준 학년도 추정이 불확실합니다(unknown).")

                # 3) '명부 기준일' + 개학일 기준으로 실제 사용할 학년도 결정
                #    - UI에서 사용자가 입력한 값(roster_basis_date)이 있으면 그걸 최우선으로 사용
                basis_date = roster_basis_date or sr.roster_basis_date or work_date
                sr.roster_basis_date = basis_date  # 최종 기준일을 ScanResult에도 반영

                if basis_date < open_date:
                    roster_time = "last_year"
                    ref_shift = -1
                else:
                    roster_time = "this_year"
                    ref_shift = 0

                roster_info["roster_time"] = roster_time          # 우리가 실제로 쓸 학년도
                roster_info["ref_grade_shift"] = ref_shift        # g_roster = g_cur + ref_shift
                roster_info["id_roster_time"] = id_roster_time    # ID 패턴 기준 값은 참고용

                sr.roster_info = roster_info

                log(
                    "[INFO] 명부 기준일/개학일 기준으로 "
                    f"'{ '작년' if roster_time == 'last_year' else '올해' } 학년도 명부'로 간주합니다. "
                    f"(ref_grade_shift={ref_shift})"
                )

                # ID 추정값과 실제 사용 학년도가 다르면 경고만
                if id_roster_time in ("this_year", "last_year") and id_roster_time != roster_time:
                    log(
                        "[WARN] 학생명부 ID 패턴 기준 학년도 추정이 "
                        "명부 기준일/개학일 기준 예상 학년도와 다를 수 있습니다. "
                        "명부가 최신인지 한 번 더 확인해 주세요."
                    )
            except Exception as e:
                log(f"[WARN] 학생명부 학년도/ID 패턴 추정 중 오류가 발생했습니다: {e}")
        else:
            log("[SKIP] 전입/전출 파일이 없어 학생명부 로드를 스킵")

        needs_open_date = bool(withdraw_file)
        sr.needs_open_date = needs_open_date
        if needs_open_date:
            sr.missing_fields.append("school_start_date")
            log("[INFO] 전출생 파일 감지 → 개학일(퇴원일자 계산용) 입력 필요")
        else:
            log("[INFO] 개학일 입력 불필요")

        base_ok = True
        if need_roster and sr.roster_path is None:
            base_ok = False

        sr.can_execute_after_input = base_ok
        sr.can_execute = base_ok and (len(sr.missing_fields) == 0)

        sr.ok = True
        log("[DONE] 스캔 완료")
        return sr
    
    except Exception as e:
        log(f"[ERROR] {e}")
        sr.ok = False
        return sr



def _extract_layout(layout_overrides: Dict[str, Any], kind: str, default_header: int):
    """
    layout_overrides[kind]가
      - dict: {"header_row": x, "data_start_row": y, ...}
      - int : y (data_start_row만)
      - None: 자동 감지
    이런 케이스를 모두 처리해서 (header_row, data_start_row) 튜플로 반환.
    """
    info = layout_overrides.get(kind)

    # dict 형태 (detect_input_layout 결과 그대로 들어온 경우)
    if isinstance(info, dict):
        header = info.get("header_row") or default_header
        data_start = info.get("data_start_row")
        return header, data_start

    # 숫자 하나만 들어온 경우 → header는 기본값 유지
    if isinstance(info, (int, float)):
        return default_header, int(info)

    # 아무 것도 없으면 자동 감지
    return default_header, None


# =========================
# EXECUTE: FULL REBUILD
# =========================

def execute_pipeline(
    scan: ScanResult,
    work_date: date,
    school_start_date: Optional[date] = None,
    layout_overrides: Optional[Dict[str, int]] = None,
) -> PipelineResult:
    """
    scan 결과를 기반으로 등록파일 + 안내파일을 한 번에 생성.
    - 신입생만 있어도 동작
    - 전입/전출/교사 파일이 없으면 그 부분은 자동으로 스킵
    - 전출은 학생명부 + 개학일이 모두 있어야 처리
    """
    logs: List[str] = []

    def log(msg: str):
        logs.append(msg)

    layout_overrides = layout_overrides or {}

    try:
        if not scan.ok:
            raise ValueError("[오류] scan 결과가 ok=False 입니다. 스캔 단계 오류를 먼저 확인해 주세요.")

        school_name = scan.school_name
        year_str = scan.year_str
        year_int = scan.year_int or int(year_str)

        log(f"[INFO] 실행 시작 | 학교={school_name}, 학년도={year_str}")
        log(f"[INFO] 작업 폴더: {scan.output_dir}")

        # -------------------------------------------------
        # 1) 입력 파일 존재 여부 확인
        # -------------------------------------------------
        if scan.freshmen_file is None:
            raise ValueError("[오류] 신입생 파일이 없습니다. 실행을 진행할 수 없습니다.")

        freshmen_path = scan.freshmen_file
        teacher_path = scan.teacher_file
        transfer_path = scan.transfer_file
        withdraw_path = scan.withdraw_file

        # -------------------------------------------------
        # 2) 인풋 읽기 (레이아웃 override 반영)
        # -------------------------------------------------
        # 신입생
        fr_header, fr_start = _extract_layout(layout_overrides, "freshmen", default_header=2)
        log(
            "[DEBUG] 신입생 layout: "
            f"header_row={fr_header}, data_start_row={fr_start if fr_start is not None else 'auto'}"
        )
        freshmen_rows = read_freshmen_rows(
            freshmen_path,
            header_row=fr_header,
            data_start_row=fr_start,
        )
        log(f"[OK] 신입생 {len(freshmen_rows)}명 로드")

        # 교사
        if teacher_path:
            t_header, t_start = _extract_layout(layout_overrides, "teacher", default_header=3)
            log(
                "[DEBUG] 교사 layout: "
                f"header_row={t_header}, data_start_row={t_start if t_start is not None else 'auto'}"
            )
            teacher_rows = read_teacher_rows(
                teacher_path,
                header_row=t_header,
                data_start_row=t_start,
            )
            log(f"[OK] 교사 신청 {len(teacher_rows)}건 로드")
        else:
            teacher_rows = []
            log("[INFO] 교사 파일 없음 → 교사 관련 처리는 스킵")

        # 전입
        if transfer_path:
            tr_header, tr_start = _extract_layout(layout_overrides, "transfer", default_header=2)
            log(
                "[DEBUG] 전입생 layout: "
                f"header_row={tr_header}, data_start_row={tr_start if tr_start is not None else 'auto'}"
            )
            transfer_rows = read_transfer_rows(
                transfer_path,
                header_row=tr_header,
                data_start_row=tr_start,
            )
            log(f"[OK] 전입생 {len(transfer_rows)}명 로드")
        else:
            transfer_rows = []
            log("[INFO] 전입생 파일 없음 → 전입 처리 스킵")

        # 전출
        if withdraw_path:
            wd_header, wd_start = _extract_layout(layout_overrides, "withdraw", default_header=2)
            log(
                "[DEBUG] 전출생 layout: "
                f"header_row={wd_header}, data_start_row={wd_start if wd_start is not None else 'auto'}"
            )
            withdraw_rows = read_withdraw_rows(
                withdraw_path,
                header_row=wd_header,
                data_start_row=wd_start,
            )
            log(f"[OK] 전출생 {len(withdraw_rows)}명 로드")
        else:
            withdraw_rows = []
            log("[INFO] 전출생 파일 없음 → 전출 처리 스킵")

        # -------------------------------------------------
        # 3) 전입 ID 생성 (학생명부가 있는 경우에만)
        # -------------------------------------------------
        transfer_done_rows: List[Dict] = []
        transfer_hold_rows: List[Dict] = []
        prefix_by_grade: Dict[int, int] = {}

        if transfer_rows:
            if not (scan.roster_path and scan.roster_info):
                # 이론상 scan.need_roster True면 이미 명부를 읽었어야 함
                raise ValueError("[오류] 전입생이 있는데 학생명부 정보가 없습니다. 스캔 결과를 확인하세요.")

            # 명부 워크시트 로드
            roster_wb = safe_load_workbook(scan.roster_path, data_only=True)
            sheets = roster_wb.worksheets
            if not sheets:
                raise ValueError(f"[오류] 학생명부에 시트가 없습니다: {scan.roster_path.name}")
            roster_ws = sheets[0]

            # scan에서 계산해 둔 roster_info 활용
            transfer_done_rows, transfer_hold_rows, prefix_by_grade = build_transfer_ids(
                transfer_rows=transfer_rows,
                roster_info=scan.roster_info,
                input_year=year_int,
            )
            log(f"[OK] 전입 ID 매칭 완료 | 완료 {len(transfer_done_rows)}명, 보류 {len(transfer_hold_rows)}명")
        else:
            log("[INFO] 전입생 없음 → 전입 ID 생성 스킵")

        # -------------------------------------------------
        # 4) 전출 퇴원 리스트 생성 (학생명부 + 개학일 + 작업일 필요)
        # -------------------------------------------------
        withdraw_done_rows: List[Dict] = []
        withdraw_hold_rows: List[Dict] = []

        if withdraw_rows:
            if not scan.roster_path:
                raise ValueError("[오류] 전출생이 있는데 학생명부 파일 경로가 없습니다. 스캔 결과를 확인하세요.")
            if not scan.roster_info:
                raise ValueError("[오류] 전출생이 있는데 학생명부 정보(roster_info)가 없습니다.")
            if school_start_date is None:
                raise ValueError("[오류] 전출 처리에 필요한 개학일이 입력되지 않았습니다.")

            roster_wb2 = safe_load_workbook(scan.roster_path, data_only=True)
            sheets2 = roster_wb2.worksheets
            if not sheets2:
                raise ValueError(f"[오류] 학생명부에 시트가 없습니다: {scan.roster_path.name}")
            roster_ws2 = sheets2[0]

            withdraw_done_rows, withdraw_hold_rows = build_withdraw_outputs(
                roster_ws=roster_ws2,
                withdraw_rows=withdraw_rows,
                school_start_date=school_start_date,
                work_date=work_date,
                roster_info=scan.roster_info,
            )
            log(
                f"[OK] 전출 퇴원 리스트 생성 | "
                f"퇴원 {len(withdraw_done_rows)}명, 보류 {len(withdraw_hold_rows)}명"
            )
        else:
            log("[INFO] 전출생 없음 → 퇴원 처리 스킵")

        # -------------------------------------------------
        # 5) 등록작업파일 생성
        # -------------------------------------------------
        if not scan.template_register:
            raise ValueError("[오류] 등록 템플릿 경로가 없습니다. 스캔 결과를 확인하세요.")

        out_register_path = scan.output_dir / f"★{school_name}_등록작업파일(작업용).xlsx"

        fill_register(
            template_path=scan.template_register,
            out_path=out_register_path,
            school_name=school_name,
            year=year_str,
            freshmen_rows=freshmen_rows,
            transfer_done_rows=transfer_done_rows,
            teacher_rows=teacher_rows,
            transfer_hold_rows=transfer_hold_rows if transfer_hold_rows else None,
            withdraw_done_rows=withdraw_done_rows if withdraw_done_rows else None,
            withdraw_hold_rows=withdraw_hold_rows if withdraw_hold_rows else None,
        )
        log(f"[OK] 등록작업파일 생성 완료: {out_register_path.name}")

        # -------------------------------------------------
        # 6) 안내파일 생성 (ID/PW)
        # -------------------------------------------------
        if not scan.template_notice:
            raise ValueError("[오류] 안내 템플릿 경로가 없습니다. 스캔 결과를 확인하세요.")

        out_notice_path = scan.output_dir /  f"☆{school_name}_신입생,전입생,교직원_ID,PW안내.xlsx"

        build_notice_file(
            template_notice_path=scan.template_notice,
            out_notice_path=out_notice_path,
            out_register_path=out_register_path,
            teacher_file_path=teacher_path,
            transfer_done_rows=transfer_done_rows,
        )
        log(f"[OK] 안내파일 생성 완료: {out_notice_path.name}")

        # -------------------------------------------------
        # 7) 결과 정리
        # -------------------------------------------------
        pr = PipelineResult(
            ok=True,
            outputs=[out_register_path, out_notice_path],
            logs=logs,
        )
        pr.transfer_in_done = len(transfer_done_rows)
        pr.transfer_in_hold = len(transfer_hold_rows)
        pr.transfer_out_done = len(withdraw_done_rows)
        pr.transfer_out_hold = len(withdraw_hold_rows)
        pr.transfer_out_auto_skip = 0  # build_withdraw_outputs 내부에서 자동제외 count를 따로 넘기고 싶으면 구조 확장

        log("[DONE] 실행 완료")
        return pr

    except Exception as e:
        # 여기서 에러를 한 번에 잡아 로그에 남김
        log(f"[ERROR] {e}")
        return PipelineResult(
            ok=False,
            outputs=[],
            logs=logs,
        )

# =========================
# 안내 메일 생성(텍스트)용 헬퍼
# =========================
def generate_notice_mail_text(work_root: Path, school_name: str) -> Tuple[bool, str]:
    """
    UI에서 호출해서 사용자에게 복사 가능한 텍스트를 출력할 때 사용.
    - DB F열 도메인 없으면: (False, 에러메시지)
    - 템플릿 txt 없으면: (False, 에러메시지)
    - 성공: (True, 렌더된 텍스트)

    notices 폴더 구조:
      resources/notices/
        - 신규등록 - 메일.txt
        - 신규등록 - 문자.txt
        - ...
    기본 메일 템플릿은 '신규등록 - 메일'을 사용.
    """
    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)

    # 1) 도메인 확인
    domain = get_school_domain_from_db(dirs["DB"], school_name)
    if not domain:
        return False, domain_missing_message(school_name)

    # 2) notices 템플릿 로드
    templates = load_notice_templates(work_root)
    if not templates:
        return False, "메일 템플릿(txt)을 찾지 못했습니다. resources/notices 폴더를 확인하세요."

    # 3) 기본 템플릿: '신규등록 - 메일' 우선, 없으면 첫 번째 아무거나
    tmpl_text = templates.get("신규등록 - 메일")
    if not tmpl_text:
        # 이름이 다를 수도 있으니, 그냥 첫 항목 사용
        tmpl_text = next(iter(templates.values()))

    rendered = render_mail_text(tmpl_text, school_name=school_name, domain=domain)
    return True, rendered


# =========================
# ENGINE ENTRYPOINT (compat)
# =========================
def run_pipeline(
    work_root: Path,
    school_name: str,
    open_date: date,                      # 개학일
    work_date: Optional[date] = None,     # 작업일 (None이면 오늘 날짜)
    layout_overrides: Optional[Dict[str, Dict[str, int]]] = None,
    roster_basis_date: Optional[date] = None,
) -> PipelineResult:
    if work_date is None:
        work_date = date.today()

    scan = scan_pipeline(
        work_root=work_root,
        school_name=school_name,
        open_date=open_date,
        work_date=work_date,
        roster_basis_date=roster_basis_date,
    )

    if not scan.ok:
        return PipelineResult(ok=False, outputs=[], logs=scan.logs)

    return execute_pipeline(
        scan=scan,
        school_start_date=open_date,
        layout_overrides=layout_overrides,
        work_date=work_date,
    )

def run_pipeline_partial(
    work_root: Path,
    school_name: str,
    open_date: date,
    mode: str,
) -> PipelineResult:
    """
    UI의 '부분 실행' 버튼용.
    현재는 안정성을 위해 전체 파이프라인을 재생성하는 방식으로 동작.
    mode: 'freshmen'|'teacher'|'transfer'|'withdraw'
    """
    return run_pipeline(work_root=work_root, school_name=school_name, open_date=open_date)