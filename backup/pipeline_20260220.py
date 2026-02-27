# core/pipeline.py
from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Sequence
from collections import Counter, defaultdict

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border
from openpyxl.worksheet.views import Selection
from pyxlsb import open_workbook as open_xlsb_workbook

from core.utils import normalize_text, text_contains, text_eq


# =========================
# Result types
# =========================
@dataclass
class PipelineResult:
    ok: bool
    outputs: List[Path]
    logs: List[str]


@dataclass
class ScanResult:
    ok: bool
    logs: List[str]
    school_name: str
    year_str: str
    year_int: int

    # paths
    project_root: Path
    input_dir: Path
    output_dir: Path
    template_register: Optional[Path]
    template_notice: Optional[Path]
    db_path: Optional[Path]

    # detected inputs
    freshmen_file: Optional[Path]
    teacher_file: Optional[Path]
    transfer_file: Optional[Path]
    withdraw_file: Optional[Path]

    # roster (only when transfer/withdraw exists)
    need_roster: bool
    roster_path: Optional[Path]
    roster_year: Optional[int]
    roster_info: Optional[Dict[str, Any]]

    # ui flags
    needs_open_date: bool
    missing_fields: List[str]
    can_execute: bool
    can_execute_after_input: bool


# =========================
# Input keyword sets
# =========================
FRESHMEN_KEYWORDS = ["신입생", "신입"]
TEACHER_KEYWORDS  = ["교사", "교원"]
TRANSFER_KEYWORDS = ["전입생", "전입"]
WITHDRAW_KEYWORDS = ["전출생", "전출"]


# =========================
# Paths
# =========================
def get_project_dirs(work_root: Path) -> Dict[str, Path]:
    """
    작업 폴더(work_root) 기준 경로 규약.

    work_root/
      DB/        : 학교 전체 명단(.xlsb)
      양식/      : 템플릿(등록/안내) .xlsx, 메일 txt
      <학교폴더>/: 인풋 파일들이 바로 들어있는 폴더
        작업/    : 산출물 생성 폴더(자동)
    """
    work_root = Path(work_root).resolve()
    return {
        "WORK_ROOT": work_root,
        "DB": work_root / "DB",
        "FORMAT": work_root / "양식",
        "SCHOOL_ROOT": work_root,  # 하위에 학교 폴더들이 존재
    }


# =========================
# File helpers
# =========================
def list_school_folders(work_root: Path) -> List[str]:
    root = Path(work_root).resolve()
    if not root.exists():
        return []
    skip = {"DB", "양식"}
    schools: List[str] = []
    for p in root.iterdir():
        if p.is_dir() and p.name not in skip and not p.name.startswith("."):
            schools.append(p.name)
    schools.sort()
    return schools


def find_templates(format_dir: Path) -> Tuple[Optional[Path], Optional[Path], List[str]]:
    """
    [양식] 폴더 템플릿 2개 식별:
    - 등록 템플릿: 파일명에 '등록' 포함
    - 안내 템플릿: 파일명에 '안내' 포함
    """
    format_dir = Path(format_dir).resolve()
    if not format_dir.exists():
        return None, None, [f"[오류] [양식] 폴더를 찾을 수 없습니다: {format_dir}"]

    xlsx_files = [
        p for p in format_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    ]
    if not xlsx_files:
        return None, None, [f"[오류] [양식] 폴더에 .xlsx 파일이 없습니다: {format_dir}"]

    reg = [p for p in xlsx_files if "등록" in p.stem]
    notice = [p for p in xlsx_files if "안내" in p.stem]

    errors: List[str] = []
    if len(reg) == 0:
        errors.append("[오류] [양식] 폴더에서 '등록' 템플릿을 찾지 못했습니다. (파일명에 '등록' 포함)")
    elif len(reg) > 1:
        errors.append("[오류] [양식] 폴더에 '등록' 템플릿이 여러 개입니다.")

    if len(notice) == 0:
        errors.append("[오류] [양식] 폴더에서 '안내' 템플릿을 찾지 못했습니다. (파일명에 '안내' 포함)")
    elif len(notice) > 1:
        errors.append("[오류] [양식] 폴더에 '안내' 템플릿이 여러 개입니다.")

    if errors:
        return None, None, errors

    return reg[0], notice[0], []


def scan_work_root(work_root: Path) -> Dict[str, Any]:
    """UI의 '경로 적용' 단계용 스캔.
    - DB, 양식(templates) 각각을 따로 진단해서 UI에서 별도 안내할 수 있게 한다.
    """
    work_root = Path(work_root).expanduser().resolve()
    dirs = get_project_dirs(work_root)

    errors_db: List[str] = []
    errors_format: List[str] = []

    db_file: Optional[Path] = None
    register_tmpl: Optional[Path] = None
    notice_tmpl: Optional[Path] = None

    # DB
    try:
        db_file = choose_db_xlsb(dirs["DB"])
    except Exception as e:
        errors_db.append(str(e))

    # templates
    try:
        register_tmpl, notice_tmpl, tmpl_errors = find_templates(dirs["FORMAT"])
        errors_format.extend(tmpl_errors)
    except Exception as e:
        errors_format.append(str(e))

    schools = list_school_folders(work_root)

    db_ok = len(errors_db) == 0 and db_file is not None
    format_ok = len(errors_format) == 0 and register_tmpl is not None and notice_tmpl is not None
    ok = db_ok and format_ok

    return {
        "ok": ok,
        "db_ok": db_ok,
        "format_ok": format_ok,
        "errors": [*errors_db, *errors_format],
        "errors_db": errors_db,
        "errors_format": errors_format,
        "work_root": work_root,
        "db_file": db_file,
        "register_template": register_tmpl,
        "notice_template": notice_tmpl,
        "school_folders": schools,
    }


def ensure_xlsx_only(p: Path) -> None:
    if p.suffix.lower() != ".xlsx":
        raise ValueError(f"[오류] 파일 형식이 .xlsx가 아닙니다: {p.name} (xlsx로 저장해서 넣어주세요)")


def backup_if_exists(out_path: Path) -> Optional[Path]:
    """기존 파일이 있으면 작업/_backup으로 이동."""
    out_path = Path(out_path)
    if not out_path.exists():
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = out_path.parent / "_backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    dest = backup_dir / f"{out_path.stem}_{ts}{out_path.suffix}"
    out_path.replace(dest)
    return dest


def find_single_input_file(input_dir: Path, keywords: Sequence[str]) -> Optional[Path]:
    if not input_dir.exists():
        return None

    kw_list: List[str] = []
    for k in keywords:
        k = "" if k is None else str(k).strip()
        if k:
            kw_list.append(k)

    if not kw_list:
        return None

    candidates: List[Path] = []
    for p in input_dir.iterdir():
        if not (p.is_file() and p.suffix.lower() == ".xlsx"):
            continue
        if p.name.startswith("~$"):
            continue
        if any(text_contains(p.name, kw) for kw in kw_list):
            candidates.append(p)

    if len(candidates) == 0:
        return None
    if len(candidates) > 1:
        raise ValueError(f"[오류] {kw_list} 포함 .xlsx 파일이 2개 이상: {[c.name for c in candidates]}")
    return candidates[0]


def choose_template_register(format_dir: Path, year_str: str = "") -> Path:
    reg, notice, errors = find_templates(format_dir)
    if errors:
        raise ValueError(errors[0])
    assert reg is not None
    return reg


def choose_template_notice(format_dir: Path, year_str: str = "") -> Path:
    reg, notice, errors = find_templates(format_dir)
    if errors:
        raise ValueError(errors[-1])
    assert notice is not None
    return notice


def choose_db_xlsb(db_dir: Path) -> Path:
    if not db_dir.exists():
        raise ValueError(f"[오류] DB 폴더가 없습니다: {db_dir}")

    xlsb_files = [
        p for p in db_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsb" and not p.name.startswith("~$")
    ]
    if not xlsb_files:
        raise ValueError("[오류] DB 폴더에 .xlsb 파일이 없습니다.")
    xlsb_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return xlsb_files[0]


def search_schools_in_db(work_root: Path, keyword: str, limit: int = 30) -> List[str]:
    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)
    db_path = choose_db_xlsb(dirs["DB"])

    kw = (keyword or "").strip()
    if not kw:
        return []

    kw_norm = normalize_text(kw)

    results: List[str] = []
    seen = set()

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            return []
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 4:
                    continue
                v = row[4].v  # E열
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue

                if kw_norm and (kw_norm in normalize_text(s)) and s not in seen:
                    seen.add(s)
                    results.append(s)
                    if len(results) >= limit:
                        break

    return results


# =========================
# DB validate (xlsb)
# =========================
def school_exists_in_db(db_dir: Path, school_name: str) -> Path:
    db_path = choose_db_xlsb(db_dir)

    target = (school_name or "").strip()
    if not target:
        raise ValueError("[오류] 학교명이 비어 있습니다(DB 검증 불가).")

    target_norm = normalize_text(target)
    found = False

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            raise ValueError("[오류] DB xlsb에 시트가 없습니다.")
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 4:
                    continue
                v = row[4].v  # E열
                if v is None:
                    continue
                cell = str(v).strip()
                if not cell:
                    continue
                cell_norm = normalize_text(cell)
                if target_norm and cell_norm and (target_norm in cell_norm):
                    found = True
                    break

    if not found:
        raise ValueError(f"[오류] DB(E열 9행~)에서 학교명 '{target}' 포함 항목을 찾지 못했습니다.")

    return db_path


def _normalize_domain(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    s = re.sub(r"^https?://", "", s, flags=re.I)
    s = s.split("/")[0].strip()
    return s


def get_school_domain_from_db(db_dir: Path, school_name: str) -> Optional[str]:
    """
    DB xlsb에서:
    - E열: 학교명 매칭
    - F열: 홈페이지(리딩게이트 전용 도메인) 반환
    없으면 None
    """
    db_path = choose_db_xlsb(db_dir)
    target = (school_name or "").strip()
    if not target:
        return None
    target_norm = normalize_text(target)

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            return None
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 5:
                    continue
                ev = row[4].v  # E
                if ev is None:
                    continue
                ecell = str(ev).strip()
                if not ecell:
                    continue
                if target_norm and (target_norm in normalize_text(ecell)):
                    fv = row[5].v  # F
                    dom = _normalize_domain("" if fv is None else str(fv))
                    return dom if dom else None
    return None


# =========================
# openpyxl custom prop guard
# =========================
def safe_load_workbook(xlsx_path: Path, data_only: bool = True):
    try:
        return load_workbook(xlsx_path, data_only=data_only)
    except TypeError as e:
        msg = str(e)
        if "openpyxl.packaging.custom" not in msg or "NoneType" not in msg:
            raise

        buffer = BytesIO()
        with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(
            buffer, "w", compression=zipfile.ZIP_DEFLATED
        ) as zout:
            for item in zin.infolist():
                if item.filename == "docProps/custom.xml":
                    root = ET.fromstring(zin.read(item.filename))
                    ns = "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
                    tag = f"{{{ns}}}property"
                    for prop in list(root.findall(tag)):
                        name = prop.get("name")
                        if name is None or str(name).strip() == "":
                            root.remove(prop)
                    new_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    zout.writestr(item, new_xml)
                else:
                    zout.writestr(item, zin.read(item.filename))

        buffer.seek(0)
        return load_workbook(buffer, data_only=data_only)


# =========================
# name normalize + suffix
# =========================
HANGUL_RE = re.compile(r"[가-힣]")
EN_RE = re.compile(r"[A-Za-z]")


def normalize_name(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-z가-힣\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""

    has_ko = bool(HANGUL_RE.search(s))
    has_en = bool(EN_RE.search(s))

    if has_ko and not has_en:
        return s.replace(" ", "")

    if has_en and not has_ko:
        parts = [p for p in s.split(" ") if p]
        parts = [p.lower().capitalize() for p in parts]
        return "".join(parts)

    if has_ko and has_en:
        def _fix_en(m: re.Match) -> str:
            tok = m.group(0).lower()
            return tok[0].upper() + tok[1:] if tok else tok
        s2 = re.sub(r"[A-Za-z]+", _fix_en, s)
        return s2.replace(" ", "")

    return ""


def normalize_name_key(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-z가-힣\s]", "", s)
    s = re.sub(r"\s+", "", s)
    return s.casefold()


def english_casefold_key(name: str) -> str:
    if name is None:
        return ""
    return str(name).strip().casefold()


def dedup_suffix_letters(n: int) -> str:
    if n <= 0:
        return ""
    out = ""
    while n > 0:
        n -= 1
        out = chr(ord("A") + (n % 26)) + out
        n //= 26
    return out


def apply_suffix_for_duplicates(names: List[str]) -> List[str]:
    total = {}
    for nm in names:
        key = english_casefold_key(nm)
        total[key] = total.get(key, 0) + 1

    seen = {}
    out = []
    for nm in names:
        key = english_casefold_key(nm)
        if total.get(key, 0) <= 1:
            out.append(nm)
            continue
        seen[key] = seen.get(key, 0) + 1
        out.append(nm + dedup_suffix_letters(seen[key]))
    return out


# =========================
# input readers
# =========================
def read_freshmen_rows(xlsx_path: Path) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    out = []
    row = 5
    while True:
        grade = ws[f"B{row}"].value
        cls   = ws[f"C{row}"].value
        num   = ws[f"D{row}"].value
        name  = ws[f"E{row}"].value

        vals = [grade, cls, num, name]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        if any(v is None or str(v).strip() == "" for v in vals):
            raise ValueError(f"[오류] 신입생 파일 {row}행(B~E)에 빈 값이 있습니다.")

        grade_i = int(str(grade).strip())
        cls_s = str(cls).strip()
        num_s = str(num).strip()
        name_n = normalize_name(name)
        if not name_n:
            raise ValueError(f"[오류] 신입생 파일 {row}행 성명(E) 정규화 결과가 비어 있습니다.")

        out.append({"grade": grade_i, "class": cls_s, "number": num_s, "name": name_n})
        row += 1

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, x)

    out.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"])))
    return out


def read_transfer_rows(xlsx_path: Path) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    expected = ["no", "학년", "반", "번호", "성명", "비고"]
    for col, exp in zip(["A", "B", "C", "D", "E", "F"], expected):
        v = ws[f"{col}2"].value
        v = "" if v is None else re.sub(r"\s+", "", str(v)).lower()
        if v != re.sub(r"\s+", "", exp).lower():
            raise ValueError("[오류] 전입생 파일 헤더(2행 A~F)가 양식과 다릅니다.")

    out = []
    row = 5
    while True:
        grade = ws[f"B{row}"].value
        cls   = ws[f"C{row}"].value
        num   = ws[f"D{row}"].value
        name  = ws[f"E{row}"].value

        vals = [grade, cls, num, name]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        if any(v is None or str(v).strip() == "" for v in vals):
            raise ValueError(f"[오류] 전입생 파일 {row}행(B~E)에 빈 값이 있습니다.")

        grade_i = int(str(grade).strip())
        cls_s = str(cls).strip()
        num_s = str(num).strip()
        name_n = normalize_name(name)
        if not name_n:
            raise ValueError(f"[오류] 전입생 파일 {row}행 성명(E) 정규화 결과가 비어 있습니다.")

        out.append({"grade": grade_i, "class": cls_s, "number": num_s, "name": name_n})
        row += 1

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, x)

    out.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"])))
    return out


def read_teacher_rows(xlsx_path: Path) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    expected = ["NO", "직위,담당", "선생님이름", "학습용ID신청", "관리용ID신청"]
    for col, exp in zip(["A", "B", "C", "D", "E"], expected):
        v = ws[f"{col}3"].value
        v = "" if v is None else re.sub(r"\s+", "", str(v))
        if v != re.sub(r"\s+", "", exp):
            raise ValueError("[오류] 교사 파일 헤더(3행 A~E)가 양식과 다릅니다.")

    out = []
    row = 4
    while True:
        b = ws[f"B{row}"].value
        c = ws[f"C{row}"].value
        d = ws[f"D{row}"].value
        e = ws[f"E{row}"].value

        if all(v is None or str(v).strip() == "" for v in [b, c, d, e]):
            break

        if c is None or str(c).strip() == "":
            row += 1
            continue

        name_n = normalize_name(c)
        if not name_n:
            row += 1
            continue

        learn_apply = not (d is None or str(d).strip() == "")
        admin_apply = not (e is None or str(e).strip() == "")

        out.append(
            {
                "position": "" if b is None else str(b).strip(),
                "name": name_n,
                "learn_apply": learn_apply,
                "admin_apply": admin_apply,
            }
        )
        row += 1

    return out


def normalize_withdraw_class(raw_class, grade_i: int) -> str:
    if raw_class is None:
        return ""
    s = str(raw_class).strip()
    if not s:
        return ""

    m = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", s)
    if m:
        g_in = int(m.group(1))
        tail = m.group(2).strip()
        if g_in == grade_i and tail:
            return tail
        if tail:
            return tail

    return s


def read_withdraw_rows(xlsx_path: Path) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    expected = ["no", "학년", "반", "성명", "비고"]
    for col, exp in zip(["A", "B", "C", "D", "E"], expected):
        v = ws[f"{col}2"].value
        v = "" if v is None else re.sub(r"\s+", "", str(v)).lower()
        if v != re.sub(r"\s+", "", exp).lower():
            raise ValueError("[오류] 전출생 파일 헤더(2행 A~E)가 양식과 다릅니다.")

    out = []
    row = 5
    while True:
        grade = ws[f"B{row}"].value
        cls   = ws[f"C{row}"].value
        name  = ws[f"D{row}"].value

        vals = [grade, cls, name]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        if any(v is None or str(v).strip() == "" for v in vals):
            raise ValueError(f"[오류] 전출생 파일 {row}행(B~D)에 빈 값이 있습니다.")

        grade_i = int(str(grade).strip())
        cls_s = normalize_withdraw_class(cls, grade_i)
        if not cls_s:
            raise ValueError(f"[오류] 전출생 파일 {row}행 반(C) 정규화 결과가 비어 있습니다.")

        name_n = normalize_name(name)
        if not name_n:
            raise ValueError(f"[오류] 전출생 파일 {row}행 성명(D) 정규화 결과가 비어 있습니다.")

        out.append({"grade": grade_i, "class": cls_s, "name": name_n})
        row += 1

    return out


# =========================
# sheet utilities
# =========================
def header_map(ws, header_row: int = 1):
    mapping = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        key = str(cell.value)
        key = key.replace("\u00A0", " ")
        key = re.sub(r"\s+", "", key)
        key = key.replace(".", "")
        mapping[key] = cell.column
    return mapping


def find_last_data_row(ws, key_col: int, start_row: int) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is not None and str(v).strip() != "":
            last = r
    return last


def clear_sheet_rows(ws, start_row=2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def move_sheet_after(wb, sheet_name: str, after_name: str):
    if sheet_name not in wb.sheetnames or after_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    wb._sheets.remove(ws)
    idx = wb.sheetnames.index(after_name)
    wb._sheets.insert(idx + 1, ws)


def delete_rows_below(ws, last_keep_row: int):
    if ws.max_row > last_keep_row:
        ws.delete_rows(last_keep_row + 1, ws.max_row - last_keep_row)


def clear_format_workbook_from_row(wb, start_row: int = 2):
    """
    모든 시트에서:
    - start_row부터 실제 데이터가 있는 마지막 행까지 스캔
    - 그 아래 행들에 대해서만 서식(fill, border) 제거
    """
    for ws in wb.worksheets:
        last_data_row = 0
        max_row = ws.max_row
        max_col = ws.max_column or 1

        # 실제 데이터 마지막 행 찾기
        for r in range(start_row, max_row + 1):
            row_has_value = False
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip() != "":
                    row_has_value = True
                    break
            if row_has_value:
                last_data_row = r

        if last_data_row == 0:
            continue

        # 마지막 데이터 행 아래부터 서식 제거
        for r in range(last_data_row + 1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()

from openpyxl.worksheet.views import Selection  # 위쪽에 이미 있을 거야. 없으면 추가.

from openpyxl.worksheet.views import Selection  # 위에 이미 있으면 중복 추가 X

def reset_view_to_a1(wb):
    """
    - 모든 시트: 화면은 A1, 커서는 A2
    - 모든 시트: 1행 고정(freeze_panes = A2)
    - 모든 시트: 그룹 선택(tabSelected) 해제
    - 통합문서: 첫 번째 시트만 선택 + 활성
    """
    # 1) 공통 뷰/고정 설정
    for ws in wb.worksheets:
        sv = ws.sheet_view

        # 화면/커서
        sv.topLeftCell = "A1"
        sv.activeCell = "A2"
        sv.selection = [Selection(activeCell="A2", sqref="A2")]

        # 1행 고정
        ws.freeze_panes = "A2"

        # 시트 그룹 선택 풀기
        if hasattr(sv, "tabSelected"):
            sv.tabSelected = False

    # 2) 첫 번째 시트만 선택 + 활성
    first_ws = wb.worksheets[0]
    if hasattr(first_ws.sheet_view, "tabSelected"):
        first_ws.sheet_view.tabSelected = True

    wb.active = 0

    # 3) 통합문서 뷰도 첫 시트 기준으로 통일
    if getattr(wb, "views", None):
        views = wb.views
        if views:
            views[0].activeTab = 0
            views[0].firstSheet = 0

# =========================
# roster analyze
# =========================
def parse_roster_year_from_filename(roster_path: Path) -> Optional[int]:
    stem = roster_path.stem
    s = stem.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()

    m = re.search(r"(\d{4})\s*학\s*년도", s)
    if m:
        return int(m.group(1))

    m2 = re.search(r"(19\d{2}|20\d{2})", s)
    if m2:
        return int(m2.group(1))

    return None


def parse_class_str(s: str) -> Optional[Tuple[int, str]]:
    if s is None:
        return None
    m = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", str(s))
    if not m:
        return None
    return int(m.group(1)), m.group(2).strip()


def extract_id_prefix4(uid: str) -> Optional[int]:
    if uid is None:
        return None
    s = str(uid).strip()
    if len(s) >= 4 and s[:4].isdigit():
        return int(s[:4])
    return None


def load_roster_sheet(dirs: Dict[str, Path], school_name: str) -> Tuple[object, Path, int]:
    # 학생명부는 SCHOOL_ROOT/학교명 바로 아래에서 찾는다
    roster_dir = dirs["SCHOOL_ROOT"] / school_name
    if not roster_dir.exists():
        raise ValueError(f"[오류] 학교 폴더가 없습니다: {roster_dir}")

    candidates = [p for p in roster_dir.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")]
    if not candidates:
        raise ValueError(f"[오류] 학교 폴더에 xlsx가 없습니다: {roster_dir}")

    # 파일명 키워드로 판별: '학생명부' 포함 파일 1개여야 함
    named = [p for p in candidates if "학생명부" in p.name]
    if len(named) == 1:
        roster_path = named[0]
    elif len(named) == 0:
        raise ValueError(
            "[오류] 학생명부 파일을 찾지 못했습니다. (파일명에 '학생명부' 포함 필수)\n"
            f"- 위치: {roster_dir}"
        )
    else:
        raise ValueError(
            f"[오류] '학생명부' 포함 파일이 2개 이상입니다. {[p.name for p in named]}"
        )

    roster_year = parse_roster_year_from_filename(roster_path)
    if roster_year is None:
        raise ValueError(
            "[오류] 학생명부 파일명에 학년도 정보(YYYY학년도)가 없습니다.\n"
            f"- 파일: {roster_path.name}\n"
            "학생명부 파일 이름은 학년도를 필수로 포함해야 합니다.\n"
        )

    wb = safe_load_workbook(roster_path, data_only=True)
    roster_ws = wb.worksheets[0]
    return roster_ws, roster_path, roster_year


def analyze_roster_once(roster_ws, input_year: int) -> Dict:
    hm = header_map(roster_ws, 1)
    need = ["현재반", "이전반", "학생이름", "아이디"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[오류] 학생명부에 '{k}' 헤더가 없습니다.")

    c_class = hm["현재반"]
    c_name  = hm["학생이름"]
    c_id    = hm["아이디"]

    prefixes_by_grade = defaultdict(list)
    name_counter_by_grade = defaultdict(Counter)
    prefixes_grade1 = []

    for r in range(2, roster_ws.max_row + 1):
        clv = roster_ws.cell(r, c_class).value
        nmv = roster_ws.cell(r, c_name).value
        idv = roster_ws.cell(r, c_id).value
        if clv is None or nmv is None:
            continue

        parsed = parse_class_str(clv)
        if parsed is None:
            continue
        g, _cls = parsed

        nm = normalize_name(nmv)
        if not nm:
            continue
        name_counter_by_grade[g][nm] += 1

        p4 = extract_id_prefix4(idv)
        if p4 is not None:
            prefixes_by_grade[g].append(p4)
            if g == 1:
                prefixes_grade1.append(p4)

    prefix_mode_by_grade = {}
    for g, arr in prefixes_by_grade.items():
        if arr:
            prefix_mode_by_grade[g] = Counter(arr).most_common(1)[0][0]

    roster_time = "unknown"
    ref_shift = 0
    if prefixes_grade1:
        mode1 = Counter(prefixes_grade1).most_common(1)[0][0]
        if mode1 == input_year:
            roster_time = "this_year"
            ref_shift = 0
        elif mode1 == input_year - 1:
            roster_time = "last_year"
            ref_shift = -1
        else:
            roster_time = "unknown"
            ref_shift = 0

    return {
        "roster_time": roster_time,
        "ref_grade_shift": ref_shift,
        "prefix_mode_by_roster_grade": prefix_mode_by_grade,
        "name_count_by_roster_grade": name_counter_by_grade,
    }


# =========================
# transfer ids
# =========================
def build_transfer_ids(
    transfer_rows: List[Dict],
    roster_info: Dict,
    input_year: int,
) -> Tuple[List[Dict], List[Dict], Dict[int, int]]:
    shift = roster_info["ref_grade_shift"]
    prefix_mode = roster_info["prefix_mode_by_roster_grade"]
    name_counts = roster_info["name_count_by_roster_grade"]

    done: List[Dict] = []
    hold: List[Dict] = []
    final_prefix_by_current_grade: Dict[int, int] = {}
    seen_in_transfer_by_grade = defaultdict(Counter)

    grade1_rows = [tr for tr in transfer_rows if tr["grade"] == 1]
    if grade1_rows:
        g1_names = [tr["name"] for tr in grade1_rows]
        g1_names_sfx = apply_suffix_for_duplicates(g1_names)
        for tr, nm_sfx in zip(grade1_rows, g1_names_sfx):
            uid = f"{input_year}{nm_sfx}"
            done.append({**tr, "id": uid})

    other_rows = [tr for tr in transfer_rows if tr["grade"] != 1]

    for tr in other_rows:
        g_cur = tr["grade"]
        g_roster = g_cur + shift

        pref = prefix_mode.get(g_roster)
        if pref is None:
            hold.append({**tr, "보류사유": f"명부 학년({g_roster})에서 ID prefix 최빈값 산출 불가"})
            continue

        final_prefix_by_current_grade[g_cur] = pref

        nm = tr["name"]
        base_cnt = name_counts.get(g_roster, Counter()).get(nm, 0)

        seen_in_transfer_by_grade[g_cur][nm] += 1
        add_seq = seen_in_transfer_by_grade[g_cur][nm]

        need_suffix = (base_cnt > 0)
        suffix = dedup_suffix_letters(add_seq) if need_suffix else ""

        uid = f"{pref}{nm}{suffix}"
        done.append({**tr, "id": uid})

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, str(x))

    done.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))
    hold.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))

    return done, hold, final_prefix_by_current_grade


# =========================
# withdraw outputs
# =========================
def build_withdraw_outputs(
    roster_ws,
    withdraw_rows: List[Dict],
    year_int: int,
    school_start_date: date,
) -> Tuple[List[Dict], List[Dict]]:
    hm = header_map(roster_ws, 1)
    need = ["현재반", "이전반", "학생이름", "아이디"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[오류] 학생명부에 '{k}' 헤더가 없습니다.")

    col_now   = hm["현재반"]
    col_prev  = hm["이전반"]
    col_name  = hm["학생이름"]
    col_id    = hm["아이디"]

    done: List[Dict] = []
    hold: List[Dict] = []

    today = date.today()
    eff = school_start_date if today < school_start_date else today

    roster_map: Dict[str, List[Dict]] = {}
    roster_by_grade_name: Dict[str, List[Dict]] = {}

    def _index_one(class_val, name_key: str, idv, name_disp: str):
        if class_val is None:
            return
        c = str(class_val).strip()
        if not c:
            return

        key1 = f"{c}|{name_key}"
        roster_map.setdefault(key1, []).append(
            {
                "class": c,
                "name_key": name_key,
                "name_disp": name_disp,
                "id": "" if idv is None else str(idv).strip(),
            }
        )

        parsed = parse_class_str(c)
        if parsed is None:
            return
        g = parsed[0]
        key2 = f"{g}|{name_key}"
        roster_by_grade_name.setdefault(key2, []).append(
            {
                "class": c,
                "name_key": name_key,
                "name_disp": name_disp,
                "id": "" if idv is None else str(idv).strip(),
                "grade": g,
            }
        )

    for r in range(2, roster_ws.max_row + 1):
        nmv = roster_ws.cell(r, col_name).value
        if nmv is None:
            continue
        name_disp = normalize_name(nmv)
        name_key  = normalize_name_key(nmv)
        if not name_key:
            continue

        idv = roster_ws.cell(r, col_id).value
        nowv  = roster_ws.cell(r, col_now).value
        prevv = roster_ws.cell(r, col_prev).value

        _index_one(nowv,  name_key, idv, name_disp)
        _index_one(prevv, name_key, idv, name_disp)

    for w in withdraw_rows:
        g_cur = w["grade"]
        w_name_disp = w["name"]
        w_name_key  = normalize_name_key(w_name_disp)
        if not w_name_key:
            hold.append(
                {
                    "학년": g_cur,
                    "반": w["class"],
                    "성명": w_name_disp,
                    "보류사유": "성명 정규화(키) 결과가 비어 있음",
                }
            )
            continue

        w_class_full = f"{g_cur}-{w['class']}"
        key = f"{w_class_full}|{w_name_key}"
        matches = roster_map.get(key, [])

        if len(matches) == 0:
            cand0 = roster_by_grade_name.get(f"{g_cur}|{w_name_key}", [])
            cand1 = roster_by_grade_name.get(f"{g_cur+1}|{w_name_key}", [])
            cand = cand0 + cand1
            if len(cand) == 1:
                matches = cand
            else:
                reason = "학년+이름 백업 실패(0건)" if len(cand) == 0 else f"학년+이름 후보가 2건 이상({len(cand)}건)"
                hold.append(
                    {
                        "학년": g_cur,
                        "반": w["class"],
                        "성명": w_name_disp,
                        "보류사유": f"반 매칭 실패 + {reason} (g 또는 g+1 탐색)",
                    }
                )
                continue

        if len(matches) > 1:
            hold.append(
                {
                    "학년": g_cur,
                    "반": w["class"],
                    "성명": w_name_disp,
                    "보류사유": f"중복 매칭({len(matches)}건)",
                }
            )
            continue

        m = matches[0]
        g_server = m.get("grade")
        if g_server is None:
            parsed = parse_class_str(m.get("class", ""))
            g_server = parsed[0] if parsed else g_cur

        withdraw_class = f"{g_server}-미편성반"
        done.append(
            {
                "퇴원반명": withdraw_class,
                "학생이름": w_name_disp,
                "아이디": m["id"],
                "퇴원일자": eff,
            }
        )

    return done, hold


def write_withdraw_to_register(wb, done_rows: List[Dict], hold_rows: List[Dict]):
    ws_done = wb["퇴원"] if "퇴원" in wb.sheetnames else wb.create_sheet("퇴원")
    ws_hold = wb["퇴원_보류"] if "퇴원_보류" in wb.sheetnames else wb.create_sheet("퇴원_보류")

    # 퇴원 완료 정렬 (퇴원반명 → 학생이름 오름차순)
    done_rows = sorted(
        done_rows,
        key=lambda r: (
            str(r.get("퇴원반명", "")).strip(),
            str(r.get("학생이름", "")).strip(),
        ),
    )

    # 보류 정렬 (학년 → 반 → 성명)
    hold_rows = sorted(
        hold_rows,
        key=lambda r: (
            str(r.get("학년", "")).strip(),
            str(r.get("반", "")).strip(),
            str(r.get("성명", "")).strip(),
        ),
    )

    clear_sheet_rows(ws_done, 2)
    clear_sheet_rows(ws_hold, 2)

    r = 2
    for row in done_rows:
        ws_done.cell(r, 1).value = row["퇴원반명"]
        ws_done.cell(r, 2).value = row["학생이름"]
        ws_done.cell(r, 3).value = row["아이디"]
        ws_done.cell(r, 4).value = row["퇴원일자"]
        ws_done.cell(r, 4).number_format = "yyyy-mm-dd"
        r += 1

    r = 2
    for row in hold_rows:
        ws_hold.cell(r, 1).value = row["학년"]
        ws_hold.cell(r, 2).value = row["반"]
        ws_hold.cell(r, 3).value = row["성명"]
        ws_hold.cell(r, 4).value = row["보류사유"]
        r += 1

    move_sheet_after(wb, "퇴원_보류", "퇴원")

    from openpyxl.styles import Font, Alignment

    def _format_sheet(ws):
        for rr in range(1, ws.max_row + 1):
            for cc in range(1, ws.max_column + 1):
                cell = ws.cell(rr, cc)
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    _format_sheet(ws_done)
    _format_sheet(ws_hold)


# =========================
# register fill (rebuild)
# =========================
def school_kind_from_name(school_name: str) -> Tuple[str, str]:
    s = (school_name or "").strip()
    if not s:
        return "", ""
    last = s[-1]
    if last == "초":
        return "초등부", "초"
    if last == "중":
        return "중등부", "중"
    if last == "고":
        return "고등부", "고"
    return "", ""


def write_transfer_hold_sheet(wb, hold_rows: List[Dict]):
    sheet_name = "전입생_보류"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    ws.delete_rows(1, ws.max_row)

    ws["A1"].value = "학년"
    ws["B1"].value = "반"
    ws["C1"].value = "번호"
    ws["D1"].value = "성명"
    ws["E1"].value = "보류사유"

    r = 2
    for row in hold_rows:
        ws.cell(r, 1).value = row.get("grade", "")
        ws.cell(r, 2).value = row.get("class", "")
        ws.cell(r, 3).value = row.get("number", "")
        ws.cell(r, 4).value = row.get("name", "")
        ws.cell(r, 5).value = row.get("보류사유", "")
        r += 1


def fill_register(
    template_path: Path,
    out_path: Path,
    school_name: str,
    year: str,
    freshmen_rows: List[Dict],
    transfer_done_rows: List[Dict],
    teacher_rows: List[Dict],
    transfer_hold_rows: Optional[List[Dict]] = None,
    withdraw_done_rows: Optional[List[Dict]] = None,
    withdraw_hold_rows: Optional[List[Dict]] = None,
) -> None:
    ensure_xlsx_only(template_path)

    wb = load_workbook(template_path)
    ws_students = wb["학생자료"]
    ws_staff = wb["직원정보"]
    ws_groups = wb["그룹반정보"]

    # =========================
    # [학생자료] 컬럼 매핑
    # =========================
    hm = header_map(ws_students, 1)
    need = ["No", "학생이름", "ID", "학교구분", "학교", "학년", "수강반"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[오류] 템플릿 [학생자료]에 '{k}' 헤더가 없습니다.")

    col_no = hm["No"]
    col_name = hm["학생이름"]
    col_id = hm["ID"]
    col_kind = hm["학교구분"]
    col_school = hm["학교"]
    col_grade = hm["학년"]
    col_class = hm["수강반"]

    # 기존 데이터 clear
    for r in range(2, ws_students.max_row + 1):
        for c in [col_no, col_name, col_id, col_kind, col_school, col_grade, col_class]:
            ws_students.cell(row=r, column=c).value = None

    kind_full, kind_prefix = school_kind_from_name(school_name)

    def write_student_row(r: int, no: int, name: str, uid: str, grade_i: int, cls_name: str):
        ws_students.cell(r, col_no).value = no
        ws_students.cell(r, col_name).value = name
        ws_students.cell(r, col_id).value = uid
        ws_students.cell(r, col_kind).value = kind_full if kind_full else ""
        ws_students.cell(r, col_school).value = school_name
        ws_students.cell(r, col_grade).value = f"{kind_prefix}{grade_i}" if kind_prefix else ""
        ws_students.cell(r, col_class).value = cls_name

    write_row = 2
    running_no = 1

    # 신입생 ID: 학년도 + 이름(중복 suffix 포함)
    fn_names = [r["name"] for r in freshmen_rows]
    fn_names_sfx = apply_suffix_for_duplicates(fn_names)
    fn_ids = [f"{year}{nm}" for nm in fn_names_sfx]

    for i, fr in enumerate(freshmen_rows):
        r = write_row + i
        write_student_row(
            r=r,
            no=running_no,
            name=fr["name"],
            uid=fn_ids[i],
            grade_i=fr["grade"],
            cls_name=f"{fr['grade']}-{fr['class']}",
        )
        running_no += 1
    write_row += len(freshmen_rows)

    # 전입생(완료)
    for tr in transfer_done_rows:
        r = write_row
        write_student_row(
            r=r,
            no=running_no,
            name=tr["name"],
            uid=tr["id"],
            grade_i=tr["grade"],
            cls_name=f"{tr['grade']}-{tr['class']}",
        )
        running_no += 1
        write_row += 1

    # 선생님(학습용) → 학생자료에 "선생님반"
    teachers_learn = [t for t in teacher_rows if t["learn_apply"]]
    t_names = [t["name"] for t in teachers_learn]
    t_names_sfx = apply_suffix_for_duplicates(t_names)
    t_ids = [f"{nm}1" for nm in t_names_sfx]

    for j, t in enumerate(teachers_learn):
        r = write_row + j
        write_student_row(
            r=r,
            no=running_no,
            name=t["name"],
            uid=t_ids[j],
            grade_i=1,
            cls_name="선생님반",
        )
        running_no += 1
    write_row += len(teachers_learn)

    # =========================
    # [직원정보]
    # =========================
    hm2 = header_map(ws_staff, 1)
    hm2_lower = {k.lower(): v for k, v in hm2.items()}

    need2 = ["no", "이름", "아이디", "권한부여"]
    for k in need2:
        if k.lower() not in hm2_lower:
            raise ValueError(f"[오류] 템플릿 [직원정보]에 '{k}' 헤더가 없습니다.")

    col_s_no = hm2_lower["no"]
    col_s_name = hm2_lower["이름"]
    col_s_id = hm2_lower["아이디"]
    col_s_role = hm2_lower["권한부여"]

    for r in range(2, ws_staff.max_row + 1):
        for c in [col_s_no, col_s_name, col_s_id, col_s_role]:
            ws_staff.cell(row=r, column=c).value = None

    teachers_admin = [t for t in teacher_rows if t["admin_apply"]]
    a_names = [t["name"] for t in teachers_admin]
    a_names_sfx = apply_suffix_for_duplicates(a_names)

    staff_write = 2
    for i, t in enumerate(teachers_admin):
        r = staff_write + i
        ws_staff.cell(r, col_s_no).value = i + 1
        ws_staff.cell(r, col_s_name).value = t["name"]
        ws_staff.cell(r, col_s_id).value = a_names_sfx[i]
        ws_staff.cell(r, col_s_role).value = "선생님"

    # =========================
    # [그룹반정보]
    # =========================
    hm_g = header_map(ws_groups, 1)
    need_g = ["그룹명", "반명", "수강료", "담임명", "FullMode"]
    for k in need_g:
        if k not in hm_g:
            raise ValueError(f"[오류] 템플릿 [그룹반정보]에 '{k}' 헤더가 없습니다.")

    col_g_group = hm_g["그룹명"]
    col_g_class = hm_g["반명"]
    col_g_fee = hm_g["수강료"]
    col_g_teacher = hm_g["담임명"]
    col_g_full = hm_g["FullMode"]

    for r in range(2, ws_groups.max_row + 1):
        for c in [col_g_group, col_g_class, col_g_fee, col_g_teacher, col_g_full]:
            ws_groups.cell(row=r, column=c).value = None

    class_set = set()
    last_student_row = find_last_data_row(ws_students, key_col=col_no, start_row=2)
    for r in range(2, last_student_row + 1):
        v = ws_students.cell(row=r, column=col_class).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            class_set.add(s)

    def parse_grade_prefix(class_name: str):
        m = re.match(r"^\s*(\d+)\s*-\s*(.+)\s*$", str(class_name))
        if not m:
            return None
        return int(m.group(1))

    def group_name_from_class(class_name: str) -> str:
        if class_name == "선생님반":
            return "기타그룹"
        g = parse_grade_prefix(class_name)
        if g is None:
            return "기타그룹"
        return f"{g}학년"

    def class_sort_key(class_name: str):
        if class_name == "선생님반":
            return (2, 0, "zzz")
        g = parse_grade_prefix(class_name)
        if g is None:
            return (1, 0, str(class_name))
        return (0, g, str(class_name))

    class_list = sorted(class_set, key=class_sort_key)

    start_r = 2
    for i, cls_name in enumerate(class_list):
        r = start_r + i
        ws_groups.cell(r, col_g_group).value = group_name_from_class(cls_name)
        ws_groups.cell(r, col_g_class).value = cls_name
        ws_groups.cell(r, col_g_fee).value = None
        ws_groups.cell(r, col_g_teacher).value = "선생님"
        ws_groups.cell(r, col_g_full).value = "Y"

    # 전입 보류 시트
    if transfer_hold_rows:
        write_transfer_hold_sheet(wb, transfer_hold_rows)

    # 전출 완료/보류 시트
    if (withdraw_done_rows is not None) and (withdraw_hold_rows is not None):
        write_withdraw_to_register(wb, withdraw_done_rows, withdraw_hold_rows)

    # 워크북 전체: 빈 행 아래 서식 제거 + A1로 통일
    clear_format_workbook_from_row(wb, start_row=2)
    reset_view_to_a1(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    backup_if_exists(out_path)
    wb.save(out_path)


# =========================
# NOTICE FILE (ID/PW 안내) 생성
# =========================
FILL_TRANSFER = PatternFill("solid", fgColor="F8CBAD")  # 옅은 주황
FILL_DUP      = PatternFill("solid", fgColor="FFFF00")  # 노랑
FILL_GREY     = PatternFill("solid", fgColor="D9D9D9")  # 회색


def _is_duplicate_id(uid: str) -> bool:
    if uid is None:
        return False
    s = str(uid).strip()
    if not s:
        return False
    # 동명이인: 아이디 끝이 대문자 A~Z (A, B, ..., AA 등)
    return bool(re.search(r"[A-Z]+$", s))


def _parse_grade_class_from_register(class_str: str) -> Tuple[Optional[int], str]:
    """
    register의 수강반: "1-3" 같은 형태 → (1, "3")
    실패하면 (None, 원본)
    """
    if class_str is None:
        return None, ""
    s = str(class_str).strip()
    if not s:
        return None, ""
    m = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", s)
    if not m:
        return None, s
    return int(m.group(1)), m.group(2).strip()


def build_notice_student_sheet(
    ws_notice,
    register_students_ws,
    transfer_ids: set,
):
    """
    안내파일 - 학생 ID,PW(학습용)
    헤더 3행: No., 학년, 반, 학생이름, ID, PW
    데이터 4행부터
    """
    hm_r = header_map(register_students_ws, 1)
    need_r = ["No", "학생이름", "ID", "수강반"]
    for k in need_r:
        if k not in hm_r:
            raise ValueError(f"[오류] 등록작업파일 [학생자료]에 '{k}' 헤더가 없습니다.")
    c_r_name = hm_r["학생이름"]
    c_r_id   = hm_r["ID"]
    c_r_cls  = hm_r["수강반"]
    c_r_no   = hm_r["No"]

    header_row = 3
    start_row = 4

    out_rows: List[Dict[str, Any]] = []
    last_r = find_last_data_row(register_students_ws, key_col=c_r_no, start_row=2)
    for r in range(2, last_r + 1):
        nm = register_students_ws.cell(r, c_r_name).value
        uid = register_students_ws.cell(r, c_r_id).value
        cls = register_students_ws.cell(r, c_r_cls).value

        cls_str = "" if cls is None else str(cls).strip()
        if cls_str == "선생님반":
            continue  # 학생 안내에서 제외

        if (nm is None or str(nm).strip() == "") and (uid is None or str(uid).strip() == ""):
            continue
        nm_s = "" if nm is None else str(nm).strip()
        uid_s = "" if uid is None else str(uid).strip()
        if not uid_s:
            continue

        g, cls_only = _parse_grade_class_from_register(cls)
        if g is None:
            g_disp = ""
            cls_disp = "" if cls is None else str(cls).strip()
        else:
            g_disp = g
            cls_disp = cls_only

        out_rows.append(
            {
                "name": nm_s,
                "id": uid_s,
                "grade": g_disp,
                "class": cls_disp,
                "is_transfer": (uid_s in transfer_ids),
                "is_dup": _is_duplicate_id(uid_s),
            }
        )

    r_out = start_row
    no = 1
    for item in out_rows:
        ws_notice.cell(r_out, 1).value = no
        ws_notice.cell(r_out, 2).value = item["grade"]
        ws_notice.cell(r_out, 3).value = item["class"]
        ws_notice.cell(r_out, 4).value = item["name"]
        ws_notice.cell(r_out, 5).value = item["id"]
        ws_notice.cell(r_out, 6).value = "1234" if item["id"] else ""

        fill = None
        if item["is_dup"]:
            fill = FILL_DUP
        elif item["is_transfer"]:
            fill = FILL_TRANSFER

        if fill is not None:
            for c in range(1, 7):
                ws_notice.cell(r_out, c).fill = fill

        no += 1
        r_out += 1

    delete_rows_below(ws_notice, r_out - 1)


def build_notice_teacher_sheet(
    ws_notice,
    teacher_rows: List[Dict],
):
    """
    안내파일 - 선생님ID,PW(관리용,학습용)
    헤더 3행, 데이터 4행부터.
    - No, 직위, 선생님이름: teacher_rows의 position/name 그대로
    - 관리용ID: admin_apply True → name, PW는 t1234
    - 학습용ID: learn_apply True → name+'1', PW는 1234
    - 신청 안 한 칸은 회색 처리
    """
    header_row = 3
    start_row = 4

    r_out = start_row
    no = 1
    for t in teacher_rows:
        pos = "" if t.get("position") is None else str(t.get("position")).strip()
        nm  = "" if t.get("name") is None else str(t.get("name")).strip()
        if not nm and not pos and (not t.get("learn_apply")) and (not t.get("admin_apply")):
            continue

        admin_apply = bool(t.get("admin_apply"))
        learn_apply = bool(t.get("learn_apply"))

        admin_id = nm if admin_apply else ""
        admin_pw = "t1234" if admin_id else ""

        learn_id = f"{nm}1" if learn_apply else ""
        learn_pw = "1234" if learn_id else ""

        # A: No. / B: 직위 / C: 선생님이름 / D: 구분용 빈 칸
        # E: 관리용 ID / F: PW / G: 구분용 빈 칸 / H: 학습용 ID / I: PW
        ws_notice.cell(r_out, 1).value = no
        ws_notice.cell(r_out, 2).value = pos
        ws_notice.cell(r_out, 3).value = nm
        ws_notice.cell(r_out, 5).value = admin_id
        ws_notice.cell(r_out, 6).value = admin_pw
        ws_notice.cell(r_out, 8).value = learn_id
        ws_notice.cell(r_out, 9).value = learn_pw

        # 회색 처리(신청 안 한 영역)
        if not admin_apply:
            for c in [5, 6]:
                ws_notice.cell(r_out, c).fill = FILL_GREY

        if not learn_apply:
            for c in [8, 9]:
                ws_notice.cell(r_out, c).fill = FILL_GREY

        no += 1
        r_out += 1

    delete_rows_below(ws_notice, r_out - 1)


def build_notice_file(
    template_notice_path: Path,
    out_notice_path: Path,
    out_register_path: Path,
    teacher_file_path: Optional[Path],
    transfer_done_rows: List[Dict],
) -> None:
    ensure_xlsx_only(template_notice_path)
    ensure_xlsx_only(out_register_path)

    wb_notice = safe_load_workbook(template_notice_path, data_only=False)
    wb_reg = load_workbook(out_register_path)

    if "학생자료" not in wb_reg.sheetnames:
        raise ValueError("[오류] 등록작업파일에 '학생자료' 시트가 없습니다.")

    ws_reg_students = wb_reg["학생자료"]

    def _norm_sheetname(s: str) -> str:
        if s is None:
            return ""
        s = str(s)
        s = s.replace("\u00A0", " ")
        s = re.sub(r"\s+", "", s)
        return s

    def _pick_sheet_by_keywords(wb, keywords: List[str]) -> str:
        keys = [_norm_sheetname(k) for k in keywords]
        for name in wb.sheetnames:
            n = _norm_sheetname(name)
            if all(k in n for k in keys):
                return name
        raise ValueError(
            "[오류] 안내 템플릿에서 필요한 시트를 찾지 못했습니다.\n"
            f"- keywords: {keywords}\n"
            f"- sheetnames: {wb.sheetnames}"
        )

    sh_student = _pick_sheet_by_keywords(wb_notice, ["학생", "PW", "학습용"])
    sh_teacher = _pick_sheet_by_keywords(wb_notice, ["선생님", "PW"])

    ws_notice_students = wb_notice[sh_student]
    ws_notice_teachers = wb_notice[sh_teacher]

    transfer_ids = set()
    for tr in transfer_done_rows:
        uid = tr.get("id")
        if uid:
            transfer_ids.add(str(uid).strip())

    build_notice_student_sheet(
        ws_notice=ws_notice_students,
        register_students_ws=ws_reg_students,
        transfer_ids=transfer_ids,
    )

    teacher_rows = read_teacher_rows(teacher_file_path) if teacher_file_path else []
    build_notice_teacher_sheet(
        ws_notice=ws_notice_teachers,
        teacher_rows=teacher_rows,
    )

    out_notice_path.parent.mkdir(parents=True, exist_ok=True)
    backup_if_exists(out_notice_path)

    # 안내 파일도 워크북 공통 규칙 적용
    clear_format_workbook_from_row(wb_notice, start_row=4)
    reset_view_to_a1(wb_notice)

    wb_notice.save(out_notice_path)


# =========================
# MAIL TEMPLATE (텍스트 치환)
# =========================
def render_mail_text(
    mail_template_text: str,
    school_name: str,
    domain: str,
) -> str:
    """
    텍스트 파일 내부:
    - 'OO초'/'OO중'/'OO고' 같은 표현 → school_name
    - 'OOOOO.readinggate.com' → domain
    """
    txt = mail_template_text or ""
    if school_name:
        txt = txt.replace("OO초", school_name).replace("OO중", school_name).replace("OO고", school_name)
    if domain:
        txt = re.sub(r"[A-Za-z0-9\-]+\.readinggate\.com", domain, txt)
    return txt


def load_mail_template_text(work_root: Path) -> Optional[str]:
    """
    work_root/양식 폴더 안에서
    '메일' 또는 '문자' + '내용' 키워드 포함 txt 1개를 찾는다.
    없으면 None
    """
    dirs = get_project_dirs(work_root)
    fmt = dirs["FORMAT"]
    if not fmt.exists():
        return None

    cands = []
    for p in fmt.glob("*.txt"):
        if not p.is_file():
            continue
        nm = p.name
        if ("메일" in nm or "문자" in nm) and ("내용" in nm):
            cands.append(p)

    if not cands:
        return None

    cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    try:
        return cands[0].read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return cands[0].read_text(encoding="utf-8-sig")


def domain_missing_message(school_name: str) -> str:
    _, kind_prefix = school_kind_from_name(school_name)
    kind_disp = kind_prefix if kind_prefix else "학교"
    return f"{kind_disp} (사용자가 작업중인) 의 도메인 주소가 존재하지 않습니다. 학교 전체 명단 파일을 확인하세요."


# =========================
# NEW: SCAN (pre-check)
# =========================
def scan_pipeline(
    work_root: Path,
    school_name: str,
    open_date: date,
) -> ScanResult:
    logs: List[str] = []

    def log(msg: str):
        logs.append(msg)

    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)

    school_name = (school_name or "").strip()
    year_str = str(open_date.year).strip()

    sr = ScanResult(
        ok=False,
        logs=logs,
        school_name=school_name,
        year_str=year_str,
        year_int=0,
        project_root=work_root,
        input_dir=Path("."),
        output_dir=Path("."),
        template_register=None,
        template_notice=None,
        db_path=None,
        freshmen_file=None,
        teacher_file=None,
        transfer_file=None,
        withdraw_file=None,
        need_roster=False,
        roster_path=None,
        roster_year=None,
        roster_info=None,
        needs_open_date=False,
        missing_fields=[],
        can_execute=False,
        can_execute_after_input=False,
    )

    try:
        if not school_name:
            raise ValueError("[오류] 학교명이 비어 있습니다.")
        year_int = int(year_str)
        sr.year_int = year_int

        db_path = school_exists_in_db(dirs["DB"], school_name)
        sr.db_path = db_path
        log(f"[OK] DB 검증 통과 | 사용 파일: {db_path.name}")

        school_dir = dirs["SCHOOL_ROOT"] / school_name
        input_dir = school_dir
        output_dir = school_dir / "작업"
        sr.input_dir = input_dir
        sr.output_dir = output_dir

        if not input_dir.exists():
            raise ValueError(f"[오류] 설정한 작업 폴더 안에 해당 학교 폴더가 없습니다: {school_dir}")

        log(f"[DEBUG] input files: {[p.name for p in input_dir.iterdir() if p.is_file()]}")

        freshmen_file = find_single_input_file(input_dir, FRESHMEN_KEYWORDS)
        if freshmen_file is None:
            raise ValueError("[오류] xlsx 형식의 신입생 명단 파일을 찾지 못했습니다. (키워드: 신입생/신입)")

        teacher_file  = find_single_input_file(input_dir, TEACHER_KEYWORDS)
        transfer_file = find_single_input_file(input_dir, TRANSFER_KEYWORDS)
        withdraw_file = find_single_input_file(input_dir, WITHDRAW_KEYWORDS)

        sr.freshmen_file = freshmen_file
        sr.teacher_file = teacher_file
        sr.transfer_file = transfer_file
        sr.withdraw_file = withdraw_file

        log(f"[OK] 신입생: {freshmen_file.name}")
        log(f"[OK] 교사: {teacher_file.name}" if teacher_file else "[SKIP] 교사 파일 없음 (키워드: 교사/교원)")
        log(f"[OK] 전입생: {transfer_file.name}" if transfer_file else "[SKIP] 전입생 파일 없음 (키워드: 전입생/전입)")
        log(f"[OK] 전출생: {withdraw_file.name}" if withdraw_file else "[SKIP] 전출생 파일 없음 (키워드: 전출생/전출)")

        template_register = choose_template_register(dirs["FORMAT"], year_str)
        sr.template_register = template_register
        log(f"[OK] 템플릿(등록): {template_register.name}")

        template_notice = choose_template_notice(dirs["FORMAT"], year_str)
        sr.template_notice = template_notice
        log(f"[OK] 템플릿(안내): {template_notice.name}")

        need_roster = bool(transfer_file) or bool(withdraw_file)
        sr.need_roster = need_roster

        if need_roster:
            roster_ws, roster_path, roster_year = load_roster_sheet(dirs, school_name)
            sr.roster_path = roster_path
            sr.roster_year = roster_year
            log(f"[OK] 학생명부: {roster_path.name} | 학년도: {roster_year}")

            roster_info = analyze_roster_once(roster_ws, input_year=year_int)
            sr.roster_info = roster_info
            log(f"[OK] 명부 판정: {roster_info['roster_time']} (shift={roster_info['ref_grade_shift']})")
        else:
            log("[SKIP] 전입/전출 파일이 없어 학생명부 로드를 스킵")

        needs_open_date = bool(withdraw_file)
        sr.needs_open_date = needs_open_date
        if needs_open_date:
            sr.missing_fields.append("school_start_date")
            log("[INFO] 전출생 파일 감지 → 개학일(퇴원일자 계산용) 입력 필요")
        else:
            log("[INFO] 개학일 입력 불필요")

        base_ok = True
        if sr.db_path is None:
            base_ok = False
        if sr.template_register is None:
            base_ok = False
        if sr.template_notice is None:
            base_ok = False
        if sr.freshmen_file is None:
            base_ok = False
        if need_roster and (sr.roster_path is None or sr.roster_info is None):
            base_ok = False

        sr.can_execute_after_input = base_ok
        sr.can_execute = base_ok and (len(sr.missing_fields) == 0)

        sr.ok = True
        log("[DONE] 스캔 완료")
        return sr

    except Exception as e:
        log(f"[ERROR] {e}")
        sr.ok = False
        sr.can_execute = False
        sr.can_execute_after_input = False
        return sr


# =========================
# EXECUTE: FULL REBUILD
# =========================
def execute_pipeline(
    scan: ScanResult,
    school_start_date: Optional[date] = None,
) -> PipelineResult:
    logs: List[str] = list(scan.logs)

    def log(msg: str):
        logs.append(msg)

    try:
        if not scan.ok:
            raise ValueError("[오류] 스캔 결과가 ok=False 입니다. 스캔을 먼저 통과해야 실행할 수 있습니다.")

        school_name = scan.school_name
        year_str = scan.year_str
        year_int = scan.year_int

        if scan.freshmen_file is None:
            raise ValueError("[오류] 신입생 파일 경로가 없습니다(스캔 결과 이상).")
        if scan.template_register is None:
            raise ValueError("[오류] 등록 템플릿 경로가 없습니다(스캔 결과 이상).")
        if scan.template_notice is None:
            raise ValueError("[오류] 안내 템플릿 경로가 없습니다(스캔 결과 이상).")

        if scan.needs_open_date and school_start_date is None:
            raise ValueError("[오류] 전출생 파일이 있어 개학일이 필요하지만 입력되지 않았습니다.")

        output_dir = scan.output_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        freshmen_file = scan.freshmen_file
        teacher_file = scan.teacher_file
        transfer_file = scan.transfer_file
        withdraw_file = scan.withdraw_file

        roster_ws = None
        roster_info = scan.roster_info
        if scan.need_roster:
            if scan.roster_path is None:
                raise ValueError("[오류] 전입/전출이 있는데 학생명부 경로가 없습니다(스캔 결과 이상).")
            wb = safe_load_workbook(scan.roster_path, data_only=True)
            roster_ws = wb.worksheets[0]
            log(f"[OK] 학생명부 재로드: {scan.roster_path.name}")

        freshmen_rows = read_freshmen_rows(freshmen_file)
        teacher_rows = read_teacher_rows(teacher_file) if teacher_file else []

        transfer_done_rows: List[Dict] = []
        transfer_hold_rows: List[Dict] = []
        if transfer_file:
            if roster_info is None:
                raise ValueError("[오류] 전입생 파일이 있는데 학생명부 분석 정보(roster_info)가 없습니다.")
            transfer_rows = read_transfer_rows(transfer_file)
            transfer_done_rows, transfer_hold_rows, _ = build_transfer_ids(
                transfer_rows=transfer_rows,
                roster_info=roster_info,
                input_year=year_int,
            )
            log(f"[OK] 전입생: 완료 {len(transfer_done_rows)} / 보류 {len(transfer_hold_rows)}")
        else:
            log("[SKIP] 전입생 처리 스킵")

        withdraw_done_rows = None
        withdraw_hold_rows = None
        if withdraw_file:
            if roster_ws is None:
                raise ValueError("[오류] 전출생 파일이 있는데 학생명부 시트가 로드되지 않았습니다.")
            withdraw_rows = read_withdraw_rows(withdraw_file)
            withdraw_done_rows, withdraw_hold_rows = build_withdraw_outputs(
                roster_ws=roster_ws,
                withdraw_rows=withdraw_rows,
                year_int=year_int,
                school_start_date=school_start_date,  # type: ignore[arg-type]
            )
            log(f"[OK] 전출생: 퇴원 {len(withdraw_done_rows)} / 보류 {len(withdraw_hold_rows)}")
        else:
            log("[SKIP] 전출생 처리 스킵")

        out_register = output_dir / f"★등록작업파일(작업용){year_str}.xlsx"

        fill_register(
            template_path=scan.template_register,
            out_path=out_register,
            school_name=school_name,
            year=year_str,
            freshmen_rows=freshmen_rows,
            transfer_done_rows=transfer_done_rows,
            teacher_rows=teacher_rows,
            transfer_hold_rows=transfer_hold_rows,
            withdraw_done_rows=withdraw_done_rows,
            withdraw_hold_rows=withdraw_hold_rows,
        )
        log(f"[DONE] 생성 완료: {out_register}")

        out_notice = output_dir / f"☆{school_name}_{year_str}신입생,전입생,교직원_ID,PW안내.xlsx"
        build_notice_file(
            template_notice_path=scan.template_notice,
            out_notice_path=out_notice,
            out_register_path=out_register,
            teacher_file_path=teacher_file,
            transfer_done_rows=transfer_done_rows,
        )
        log(f"[DONE] 생성 완료: {out_notice}")

        kind_full, _ = school_kind_from_name(school_name)
        if not kind_full:
            log("[WARN] 학교명 끝 글자가 초/중/고가 아니라 학교구분/학년 표기가 빈칸으로 들어갔을 수 있음")

        return PipelineResult(ok=True, outputs=[out_register, out_notice], logs=logs)

    except Exception as e:
        log(f"[ERROR] {e}")
        return PipelineResult(ok=False, outputs=[], logs=logs)


# =========================
# 안내 메일 생성(텍스트)용 헬퍼
# =========================
def generate_notice_mail_text(work_root: Path, school_name: str) -> Tuple[bool, str]:
    """
    UI에서 호출해서 사용자에게 복사 가능한 텍스트를 출력할 때 사용.
    - DB F열 도메인 없으면: (False, 에러메시지)
    - 템플릿 txt 없으면: (False, 에러메시지)
    - 성공: (True, 렌더된 텍스트)
    """
    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)

    domain = get_school_domain_from_db(dirs["DB"], school_name)
    if not domain:
        return False, domain_missing_message(school_name)

    tmpl = load_mail_template_text(work_root)
    if not tmpl:
        return False, "메일 템플릿(txt)을 찾지 못했습니다. [양식] 폴더에 텍스트 파일을 넣어주세요."

    rendered = render_mail_text(tmpl, school_name=school_name, domain=domain)
    return True, rendered


# =========================
# ENGINE ENTRYPOINT (compat)
# =========================
def run_pipeline(
    work_root: Path,
    school_name: str,
    open_date: date,
) -> PipelineResult:
    """UI의 '전체 실행' 버튼용."""
    scan = scan_pipeline(work_root=work_root, school_name=school_name, open_date=open_date)
    if not scan.ok:
        return PipelineResult(ok=False, outputs=[], logs=scan.logs)
    return execute_pipeline(scan=scan, school_start_date=open_date)


def run_pipeline_partial(
    work_root: Path,
    school_name: str,
    open_date: date,
    mode: str,
) -> PipelineResult:
    """
    UI의 '부분 실행' 버튼용.
    현재는 안정성을 위해 전체 파이프라인을 재생성하는 방식으로 동작.
    mode: 'freshmen'|'teacher'|'transfer'|'withdraw'
    """
    return run_pipeline(work_root=work_root, school_name=school_name, open_date=open_date)